import requests
import os
import time
import shutil
import zipfile
from glob import glob
import datetime
import dataEntryFunction as df


def setfoldername1():




#########################################################################################################
#########################################################################################################

#                     *  검수 완료되면 실행
#                     *  1,2 하나로 합쳤습니다.
#                     *  다운받고 > 압축풀고 > 분류하는거까지 실행하는 코드입니다.
#
# 월요일에는 바탕화면 폴더를 토요일 날짜에서 일요일로 바꾸고 main.py와 open_C_or_E_files.py의 날짜를 하루전으로 변경해야 함

###################################################################################################
###################################################################################################
                                                                                                ###
                                                                                                ###
                    filename = "TOT_240306_01"  ## 가운데 오늘날짜 쓰기     ex) TOT_221221_01    ###
                    objdate  = ""                ## -1로 할거면 ""로 두기   ex) 20221215         ###
                                                                                                ###
                                                                                                ###
###################################################################################################
###################################################################################################



                    today = filename.split("_")[1][2:]
                    return today, filename, objdate


url1 = "http://10.28.78.30:8081"
url2 = "http://10.28.78.30:8082"

if not df.compare_server_data(url1, url2):
    exit()
else:
    print("PASS")

작업파일경로 = "C:/Users/user/Documents/GitHub/lotteplaza/작업파일/"
결과폴더경로= "C:/Users/user/Documents/GitHub/webdataentry/result/"
today,filename,objdate = setfoldername1()
## 다운로드
url=f"http://10.28.78.30:8889/export_result/-1/{filename}/{objdate}/-1"
if objdate == "":
    url = f"http://10.28.78.30:8889/export_result/-1/{filename}/-1/-1"
else:
    url = f"http://10.28.78.30:8889/export_result/-1/{filename}/{objdate}/-1"
print(url)
response = requests.get(url)
print("\tresponse :", response.status_code)
time.sleep(2)
df.makedir(작업파일경로 + today)
df.makedir(작업파일경로 + today + "/제외") ## 월요일에는 제외를 합쳐야되기 때문에 그부분 처리하기
df.makedir(작업파일경로 + today + "/3차")
df.makedir(작업파일경로 + today + "/4차")
df.makedir(작업파일경로 + today + "/결과") ## 월요일에는 주말꺼엔 결과 없어도 됨.
df.makedir(작업파일경로 + today + "/결과/MARGIN_DIFF")
shutil.copy2(결과폴더경로 + filename + "_except.zip",
             작업파일경로 + today + "/" + filename + "_except.zip")
shutil.copy2(결과폴더경로 + filename + "_result.xlsx",
             작업파일경로 + today + "/3차" + "/" + filename + "_result.xlsx")

todayinteger = int("1"+today) ## 오늘날짜변환 (int로 만들기위해서 10000더함)
curfolder = 0
for i in os.listdir(작업파일경로):
    if i.isdigit(): ## 하위폴더 모두 검사해서 숫자인경우만
        iterInteger = int("1"+i) #앞에 1 추가(나중에 다시 뺄것임)
        if iterInteger < todayinteger and curfolder < iterInteger: ## 오늘날짜보다 작은것중에 가장 큰 폴더
            curfolder = iterInteger
curfolder = str(curfolder)[1:] ## 다시폴더명으로 변환

shutil.copy2(작업파일경로 + curfolder + "/결과/Data_Entry_작업량_기록_" + curfolder + "2024.xlsx",
             작업파일경로 + today + "/결과/Data_Entry_작업량_기록_" + today + "2024.xlsx")

shutil.copy2(작업파일경로 + curfolder + "/결과/Daily Note_" + curfolder + "2024.xlsx",
             작업파일경로 + today + "/결과/Daily Note_" + today + "2024.xlsx")

result_3차파일경로 = 작업파일경로 + today + "/3차" + "/" + filename + "_result.xlsx"

zipf = zipfile.ZipFile(작업파일경로 + today + "/" + filename + "_except.zip")
zipf.extractall(작업파일경로 + today + "/제외")
zipf.close()
print("다운로드 및 파일 이동 완료.")
print("제외 파일 분류를 시작합니다.")

## 제외 파일 분류
제외폴더경로=f"C:/Users/user/Documents/GitHub/lotteplaza/작업파일/{today}/제외/"
제외pdf경로템플릿= f"C:/Users/user/Documents/GitHub/lotteplaza/작업파일/{today}/제외/*.pdf"
dict={"업체코드 검색 불가" : [], "입력 대상 제품 없음" : [], "제품코드, UPC 부재 & 한글 Description & 수기" : [], "파본" : []}
totalnum=len(glob(제외pdf경로템플릿))
print("==============================")
print("작업 폴더 :", today)
print("==============================")
print("전체 제외 파일 :", totalnum)

# 각 요소가 리스트면 and조건 적용, 각 요소들끼리는 or조건
제외키워드 ={
     "업체코드 검색 불가" : [["업체","검색"],["업체","불가"],
                                "업체코드", "미등록업체", "미등록업체","업체명확인불가"],
     "입력 대상 제품 없음" : [
                             ["제외","업체"] ,           "rheebros", "sungwon", "성원", "리브로", "패킹", "단가"
                             , ["제외","업종"],          "단가", "charge", "recap", "stale", "weekly", "주간"
                             , ["내용","없음"],          "리스트", "리포", "report", "상세내역없음", "위클리", "리퐅","리폿"
                             , ["내역","없음"],          "state", "statm", "집계", "stock", "receip", "중복"
                             , ["입력", "없음"]          ,"매장이동", "크레딧", "크래딧", "credit", "crdit","credot"
                                                         ,"이동", "거래", "REUTN", "리턴", "flxjs", "return" ],
     "제품코드, UPC 부재 & 한글 Description & 수기" : ["코드없음", "수기", "제품번호", "upc", "한글"],
     "파본" : ["파본", "vkqhs"]
     }

for x in glob(제외pdf경로템플릿):
    분류성공 = False
    fname=os.path.basename(x)
    rst = fname.translate(str.maketrans('1234567890ABCDEFGHIJKLMNOPQRSTUVWXYZ', '..........abcdefghijklmnopqrstuvwxyz')).split(".")[-2].replace(" ","")
    for category, keywords in 제외키워드.items():
        for keyword in keywords:
            if isinstance(keyword, list):
                if all(word in rst for word in keyword):
                    dict[category].append(fname)
                    분류성공 = True
                    break
            elif keyword in rst:
                dict[category].append(fname)
                분류성공 = True
                break
        if 분류성공:
            break
    if not 분류성공:
        print(fname)

print("------------------------------")
print(dict)
print("분류 완료, 파일 이동...")
filemovenum=0
for i in dict:
    if len(dict[i]) > 0:
        df.makedir(제외폴더경로+i)
        print("\t",i,":",len(dict[i]))
        filemovenum+=len(dict[i])
    for k in dict[i]:
        shutil.move(os.path.join(제외폴더경로,k), os.path.join(제외폴더경로+i,k))
print("------------------------------")
print(filemovenum,"개 파일 이동 완료")


print("제외 폴더 압축 시작")
zip_path=f"C:/Users/user/Documents/GitHub/lotteplaza/작업파일/{today}/제외"
result_path=f"C:/Users/user/Documents/GitHub/lotteplaza/작업파일/{today}/결과/제외인보이스_{today+str(datetime.datetime.now().year)}"
shutil.make_archive(result_path,'zip',zip_path)
print("제외 폴더 압축 완료")

if filemovenum==totalnum:
    print("제외 폴더 수동 확인이 필요없습니다")
else:
    print("!!!!!!!!전체 제외 파일과 이동된 파일의 개수가 다릅니다.")
    print("!!!!!!!!제외 폴더 수동 확인이 필요합니다.")


check_today = datetime.datetime.now()
year=str(check_today.year)[2:4]
day=str(check_today.day).zfill(2)
month=str(check_today.month).zfill(2)

if year+month+day != filename.split("_")[1] :
    print(filename.split("_")[1],"폴더로 작업했습니다. 재확인 바랍니다.")

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# 엑셀 파일 로드
workbook = load_workbook(filename= result_3차파일경로)

# 시트 선택
sheet = workbook.active

# 데이터가 저장된 행의 수 계산
total_rows = sheet.max_row

# 필터링할 범위 선택
data_range = sheet['A1':'AN' + str(total_rows)]
b_column_index = column_index_from_string('B')
v_column_index = column_index_from_string('V')
x_column_index = column_index_from_string('X')
y_column_index = column_index_from_string('Y')
z_column_index = column_index_from_string('Z')
ab_column_index = column_index_from_string('AB')
ac_column_index = column_index_from_string('AC')

# V, X, Y, Z, AB, AC 열에서 빈 텍스트 또는 빈 데이터인 row 수 계산
empty_row_count = 0

for row in data_range:
    b_value = str(row[b_column_index - 1].value).strip()
    v_value = str(row[v_column_index - 1].value).strip()
    x_value = str(row[x_column_index - 1].value).strip()
    y_value = str(row[y_column_index - 1].value).strip()
    z_value = str(row[z_column_index - 1].value).strip()
    ab_value = str(row[ab_column_index - 1].value).strip()
    ac_value = str(row[ac_column_index - 1].value).strip()

    values_to_check = [v_value, x_value, y_value, z_value, ab_value, ac_value]
    empty_value_names = ["v_value", "x_value", "y_value", "z_value", "ab_value", "ac_value"]

    values_to_check = {
        "v_value": v_value,
        "x_value": x_value,
        "y_value": y_value,
        "z_value": z_value,
        "ab_value": ab_value,
        "ac_value": ac_value
    }
    for name, value in values_to_check.items():
        if value in [None, '', "None"]:
            empty_row_count += 1
            print(f"{b_value} : Empty or 'None' value found in: {name}")

if empty_row_count==0:
    print("컬럼 누락이 없습니다.")
    print(filename+".xlsx : 3차 ===> 4차 변환중... ")
    data = {'file_result_name': filename}  # 추가적인 폼 데이터
    url = 'http://10.28.78.30:8081/file_input_upload_step_4'  # 1번서버
    files = {'file_input': (filename+".xlsx", open(result_3차파일경로, 'rb'))}  # 1번서버
    response = requests.post(url, files=files, data=data)

    print("Server 1 response : ", response)

    down_url = f"http://10.28.78.30:8081/down_file/result_step_4/{filename}.xlsx"
    down_response = requests.get(down_url)
    save_path = os.path.join("C:\\Users\\user\\Downloads\\", f"{filename}.xlsx")
    with open(save_path, 'wb') as f:
        f.write(down_response.content)
    print(save_path,"에 4차 파일 저장 완료")

    import subprocess
    bat_file_path = r'C:\Users\user\Desktop\data_entry_ocr.bat' ## ocr 실행
    subprocess.run([bat_file_path])
    exec(open('open_C_or_E_files.py', encoding="utf-8").read())
else:
    print("!!!!!!!일부 컬럼이 누락되었습니다. 확인바랍니다.")
    print("Description(V), F19(X, 입수량), F38(Y, 단가), F39(Z, 날짜), QTY(AB), Amount(AC) 컬럼들 [필드 값 없음] 항목 찾아서 수기 입력")
