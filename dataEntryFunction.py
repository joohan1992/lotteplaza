import os
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import requests
from PIL import Image
import shutil
import datetime
import pymysql
import time

from selenium import webdriver
import sys

##날짜 string 구하기
today = datetime.datetime.now() - datetime.timedelta(days=2) # 토요일 당일에 오후 배분 못했을 때 월요일에 다운받는 경우
today = datetime.datetime.now()
today_weekday = (datetime.datetime.now() - datetime.timedelta(days=2)).weekday() # 토요일 당일에 오후 배분 못했을 때 월요일에 다운받는 경우
today_weekday = datetime.datetime.now().weekday()

# YYMMDD 형식
yymmdd = today.strftime("%y%m%d")
# MMDDYY 형식
mmddyy = today.strftime("%m%d%y")
# YYYY-MM-DD 형식
yyyy_mm_dd = today.strftime("%Y-%m-%d")

# 2자리로 맞춤
year = today.strftime("%y")
month = today.strftime("%m")
day = today.strftime("%d")

def set_filename():
#           *  downAndUploadInv.py 실행 전 수정합니다.
#           *  supervalu 입력 후 1,2번 서버에서 처리하고 다운로드 받은 직후 수정합니다.
#           *  filename, objdate 수정하고 executeThisAfterInputSupervalue.py만 실행하면 됩니다.

    filename = ""       ## 오늘날짜 쓰기   ex) 221221
    objdate  = []       ## _뒤에 적힐것    ex) ["01","02"]

#           *   filename = "", objdate = []로 놓으면,
#           *   기본값(오늘날짜, ["01", "02"]로 세팅됩니다.

        ##### 기본값 셋팅 ######
    if filename == "" or filename is None:
        filename = yymmdd
    if objdate == "" or objdate == [] or objdate is None:
        objdate = ["01", "02"]
        ########################

    return filename, objdate

g_data = {
    'base_address': 'http://10.28.78.30:8889',
    'tmp_uri': './tmp',
    'host': 'localhost',
    'port': 3307,
    'user': 'root',
    'password': 'a4s5d6f7!@',
    'db': 'dew',
    'charset': 'utf8',
    'list_new_store': []   # 신규 매장 목록, 있으면 해당 매장 데이터는 data_lookup시 다른 매장 데이터도 참조함
}


def get_conn(g_data):
    return pymysql.connect(host=g_data['host'], port=g_data['port'], user=g_data['user'], password=g_data['password'], db=g_data['db'], charset=g_data['charset'])

def execute_r_query(g_data, query):
    conn = get_conn(g_data)
    curs = conn.cursor(pymysql.cursors.DictCursor)
    curs.execute(query)
    result = curs.fetchall()
    conn.close()
    return result

def get_first_workdata(text):
    text = text.replace(" ","") # 스페이스바 제거
    text = text.replace(":","") # : 제거
    text = text.replace("/","") # / 제거
    result = {'이지원' :[0,0]} # 이지원님은 항상 있어야 함.
    for line in text.split('\n'):
        name_and_scores = line.split('R1')
        if len(name_and_scores) == 2:
            if name_and_scores[0] == "이지워":
                name = "이지원"
            else:
                name = name_and_scores[0]
            r1 = name_and_scores[1].split('R2')[0]
            r2 = name_and_scores[1].split('R2')[1]
            result[name.strip()] = [int(r1), int(r2)]
    return result

def get_second_workdata () :
    conn = get_conn(g_data)
    curs = conn.cursor(pymysql.cursors.DictCursor)
    query = f"""SELECT 
                   AU.USER_KEY, 
                   AU.USER_NM, 
                   DII.LOOKUP_BASE,
                   COUNT(*) AS CNT 
                FROM dd_inv_item DII 
                INNER JOIN dd_inv DI 
                   ON DI.INV_NO = DII.INV_NO 
                   AND DI.USE_YN = 'Y' 
                INNER JOIN dd_work_schd DWS 
                   ON DWS.WORK_SCHD_NO = DI.WORK_SCHD_NO
                INNER JOIN am_user AU
                   ON AU.USER_NO = DWS.WORK_USER_NO 
                WHERE	1=1 
                   AND DWS.WORK_DATE = '{yyyy_mm_dd}'
                   AND DI.INV_STATUS IN ('C', 'P', 'R', 'TEST_C', 'TEST_P', 'TEST_R') 
                   AND DII.USE_YN = 'Y' 
                   AND DWS.USE_YN = 'Y' 
                GROUP BY AU.USER_KEY, DII.LOOKUP_BASE """
    curs.execute(query)
    results = curs.fetchall()
    result_dict = {}
    # LOOKUP_BASE : N - rank2 / else - rank1
    for i in results:
        if i["USER_NM"] not in result_dict :
            result_dict[i["USER_NM"]] = [0, 0]
        if i["LOOKUP_BASE"] == 'N':
            result_dict[i["USER_NM"]][1] =  i["CNT"]
        else :
            result_dict[i["USER_NM"]][0] += i["CNT"]

    return result_dict



def get_previous_day(yymmdd):

    date_obj = datetime.datetime.strptime(yymmdd, '%y%m%d')
    previous_day = date_obj - datetime.timedelta(days=1)
    print(previous_day)
    return previous_day.strftime('%y%m%d')


def fetch_filename_from_link(server_url):
    try:
        response = requests.post(server_url)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        links = soup.find_all('a', href=True)
        for link in links:
            if "DB.xlsx" in link['href']:
                return link.text.strip()
    except requests.RequestException as error:
        print(f"Error fetching data from the server: {error}")
        return None
    return None

def compare_server_data(url1,url2):
    print("================ 서버점검결과 ================")
    result1 = fetch_filename_from_link(url1)
    result2 = fetch_filename_from_link(url2)
    ret_val = False

    if result1 is None:
        print(f"{url1} : Error. 점검요망")
    elif result2 is None:
        print(f"{url2} : Error. 점검요망")
    elif result1 == result2:
        print(f'{result1} : [{url1}], [{url2}] 기초데이터가 같습니다.')
        ret_val = True
    else:
        print(f"{url1} : {result1}")
        print(f"{url2} : {result2}")
        print("서버를 재시작해서 기초데이터를 동기화 하세요.")
    print("=============================================")
    return ret_val

def move_files_with_rename(src_folder, dest_folder):
    if not os.path.exists(src_folder):
        print(f"{src_folder}폴더가 없습니다.")
        return False
    if not os.path.exists(dest_folder):
        print(f"{dest_folder}폴더가 없습니다.")
        return False
    # 백업폴더 생성
    shutil.copytree(src_folder, dest_folder.rstrip('\\') + "_합치기전backup")
    # 대상 폴더 내에 동일한 구조의 폴더 생성
    for dirpath, dirnames, filenames in os.walk(src_folder):
        for dirname in dirnames:
            dest_dir = os.path.join(dest_folder, os.path.relpath(os.path.join(dirpath, dirname), src_folder))
            if not os.path.exists(dest_dir):
                os.makedirs(dest_dir)
        # 파일 이동
        for filename in filenames:
            src_file = os.path.join(dirpath, filename)
            dest_file = os.path.join(dest_folder, os.path.relpath(src_file, src_folder))
            base, extension = os.path.splitext(dest_file)
            counter = 1
            # 파일명 중복 체크 및 수정
            while os.path.exists(dest_file):
                dest_file = f"{base}({counter}){extension}"
                counter += 1
            shutil.move(src_file, dest_file)


def image_to_pdf(image_path):
    with Image.open(image_path) as img:
        if img.mode == 'RGBA':
            img = img.convert('RGB')
        # PDF 파일 경로 생성 (동일한 이름의 .pdf 확장자)
        pdf_path = f"{os.path.splitext(image_path)[0]}.pdf"
        img.save(pdf_path, 'PDF', resolution=100.0)
    os.remove(image_path)

def get_weekday_from_yymmdd(yymmdd):
    try:
        date_obj = datetime.datetime.strptime(yymmdd, "%y%m%d")
        weekday = date_obj.weekday()
        weekdays = ["월요일", "화요일", "수요일", "목요일", "금요일", "토요일", "일요일"]
        return weekdays[weekday]
    except ValueError:
        return "유효하지 않은 날짜 형식입니다."

def getDownloadObjDates(current_weekday = today_weekday ):
    yesterday = today - datetime.timedelta(days=1)
    two_days_ago = today - datetime.timedelta(days=2)
    weekdays = {
        0: [two_days_ago.strftime("%y%m%d")],
        1: [two_days_ago.strftime("%y%m%d"), yesterday.strftime("%y%m%d")],
    }

    ## dic에서 0,1 가져오고 없으면 어제날짜
    return weekdays.get(current_weekday, [yesterday.strftime("%y%m%d")])


def progress_bar(current, total, title = "Download", bar_length=40):
    progress = current / total
    completed_length = int(bar_length * progress)
    bar = '|'+'█' * completed_length + ' ' * (bar_length - completed_length) + '| '
    percentage = str("%3d" % int(progress * 100)) + "%"
    if progress == 1:
        bar = bar + 'Done'
    print("\r", title,":", percentage, bar,  end="", flush=True)


def isMonday():
    if today_weekday == 0:
        print("오늘은 월요일 입니다.")
        return True
    else:
        return False

def getAbsTodayYYMMDD():
    return yymmdd

def postprocessing_date_process():
    filename, objdate = set_filename()
    today_mmdd = filename[2:6]
    today_yy = filename[0:2]
    return str(today_mmdd) + str(today_yy) ## mmddyy

def makedir(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
            print("폴더생성 : ",directory)
        else:
            pass
    except OSError:
        print("Error: Failed to create the directory.")

def delete_columns(ws, columns_to_delete):
    for column in sorted(columns_to_delete, reverse=True):
        ws.delete_cols(column)


def count_pdf_files(path): ## path 하위의 pdf파일 개수 세는 함수
    count = 0
    for root, dirs, files in os.walk(path):
        for file in files:
            if file.lower().endswith('.pdf'):
                count += 1
    return count


def getDownloadUrl(download_urls, driver, dept, obj_yymmdd):
    yyyy = "20" + str(obj_yymmdd[0:2])
    mm = obj_yymmdd[2:4]
    dd = obj_yymmdd[4:6]

    print(dept,"다운로드 경로 수집 중....")
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    divs = soup.find_all('div', {'class': 'card-header'})
    cnt_file = 0
    for div in divs:
        if mm + "-" + dd + "-"+yyyy in div.text:
            # 해당 div의 하위 ul > li > a를 찾는다
            ul_tag = div.find_next('ul')
            li_tags = ul_tag.find_all('li')
            for li in li_tags:
                a_tag = li.find_next('a')
                inv_no = a_tag["href"].rsplit('/', 2)[1]
                download_url = "http://134.209.127.114/invoice/download_all_files_contractor/"+inv_no
                download_urls.append(download_url)
                cnt_file += 1
    print(dept, "다운로드 대상 파일 :", cnt_file, "개")
    return download_urls


def getDownloadUrl2(pages, stores, departs, driver, dept, obj_yymmdd):
    yyyy = "20" + str(obj_yymmdd[0:2])
    mm = obj_yymmdd[2:4]
    dd = obj_yymmdd[4:6]

    print(dept,"다운로드 경로 수집 중....")
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    divs = soup.find_all('div', {'class': 'card-header'})
    cnt_file = 0
    for div in divs:
        if mm + "-" + dd + "-"+yyyy in div.text:
            # 해당 div의 다음 ul > li > a를 찾는다
            ul_tag = div.find_next('ul')
            li_tags = ul_tag.find_all('li')
            for li in li_tags:
                a_tag = li.find_next('a')
                store = a_tag.text.split(' ')[0]
                inv_no = a_tag["href"].rsplit('/', 2)[1]
                page = "http://134.209.127.114/invoice/inv_contractor_history/"+inv_no
                pages.append(page)
                departs.append(dept)
                stores.append(store)
                cnt_file += 1
    print(dept, "다운로드 대상 파일 :", cnt_file, "개")
    return pages, stores, departs


def download(url, cur_dt, store, dict_result):
    try:
        file_name = url.rsplit('/', 1)[1]
        file_name_ext = file_name.rsplit('.', 1)
        idx_dup_file = 0
        while os.path.isfile(f'C:/Users/user/Downloads/{cur_dt}/{store}/{file_name_ext[0]}{"_"+str(idx_dup_file) if idx_dup_file > 0 else ""}.{file_name_ext[1]}'):
            idx_dup_file += 1
        isDone = False
        try_cnt = 0
        while not isDone and try_cnt < 5:
            isDone = True
            try:
                with open(f'C:/Users/user/Downloads/{cur_dt}/{store}/{file_name_ext[0]}{"_"+str(idx_dup_file) if idx_dup_file > 0 else ""}.{file_name_ext[1]}', "wb") as file:   # open in binary mode
                    response = requests.get(url)               # get request
                    file.write(response.content)
            except:
                try_cnt += 1
                time.sleep(3)
                isDone = False
        if try_cnt == 5:
            dict_result['url'] = url
            dict_result['try_cnt'] = try_cnt
    except:
        dict_result['url'] = url
        dict_result['try_cnt'] = -1

def getClassifiedRows(filepath):
    # 엑셀 파일 로드
    workbook = load_workbook(filename=filepath)
    sheet = workbook.active
    total_rows = sheet.max_row
    data_range = sheet['A1':'AN' + str(total_rows)]
    c_column_index = column_index_from_string('C')
    # 빈 텍스트와 None 값을 제외한 C열의 데이터 수 계산
    non_empty_c_count = 0
    for row in data_range:
        c_value = row[c_column_index - 1].value
        if c_value is not None and c_value != '':
            non_empty_c_count += 1
    return non_empty_c_count - 1

def 본문1():
    본문 = '''
안녕하세요, 리테일앤인사이트 성주한입니다.


금일 전체 인보이스는 총 #{TOT_CNT}건이었습니다.

제외 인보이스 #{EXT_CNT}건을 제외한 #{VAL_CNT}건을 작업하였습니다.


제외 인보이스는 아래와 같으며 파일로 압축하여 보내드립니다.
'''
    return 본문

def 본문2():
    본문 = '''

감사합니다.

성주한 올림.'''
    return 본문
