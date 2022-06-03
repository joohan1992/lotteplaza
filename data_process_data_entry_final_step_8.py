from openpyxl import load_workbook, Workbook
from os import listdir, mkdir
from os.path import isfile, join, isdir
import data_process_data_entry_lib_8 as jhlib


def dict_to_file(dict_link, dict_vendor_link, ws_link):

    for u_key in dict_link:
        arr2 = []
        idx3 = -1
        for u_key2 in dict_link[u_key]:
            idx3 += 1
            if idx3 == 1:
                arr2.append(u_key2)
                if dict_link[u_key][13] == 'N':
                    if int(dict_link[u_key][3]) > 1 and (dict_link[u_key][6] is None or dict_link[u_key][6] == ''):
                        arr2.append('업체+제품 검색 시 다수이나 인보이스 내 UPC 없음')
                        arr2.append('')
                        arr2.append('Description 및 F19 입력')
                        arr2.append('')
                    elif (dict_link[u_key][12] is None or dict_link[u_key][12] == '') and (
                            dict_link[u_key][6] is None or dict_link[u_key][6] == ''):
                        arr2.append('인보이스에 제품코드와 UPC 둘 다 없음')
                        arr2.append('')
                        arr2.append('제품 코드, UPC 없는지 확인 후 Description 및 F19 입력')
                        arr2.append('')
                    elif dict_link[u_key][10] not in dict_vendor_link:
                        arr2.append('매장 데이터에 등록되지 않은 업체 코드')
                        arr2.append('')
                        arr2.append('업체 코드 재확인 후 Description 및 F19 입력')
                        arr2.append('')
                    elif int(dict_link[u_key][3]) == 0:
                        arr2.append('신상품')
                        arr2.append('')
                        arr2.append('제품 코드 및 업체 코드 재확인 후 Description 및 F19 입력')
                        arr2.append('')
                    else:
                        arr2.append('업체+제품 검색 시 다수이나 인보이스의 UPC와 상이함')
                        arr2.append('')
                        arr2.append('UPC 재확인 후 Description 및 F19 입력')
                        arr2.append('')
                elif dict_link[u_key][13] == 'I' and dict_link[u_key][5] == 0 and dict_link[u_key][6] is not None and \
                        dict_link[u_key][6] != '':
                    arr2.append('업체+제품으로 검색된 UPC가 단일이지만 인보이스의 UPC와 상이')
                    arr2.append('')
                    arr2.append('UPC 재확인')
                    arr2.append('')
                else:
                    arr2.append('')
                    arr2.append('')
                    arr2.append('')
                    arr2.append('')
            elif idx3 == 17:
                if dict_link[u_key][23] == 1:
                    arr2.append(str(dict_link[u_key][21] / dict_link[u_key][20]))
                else:
                    arr2.append(u_key2)
            elif idx3 == 20 or idx3 == 21:
                arr2.append(str(u_key2))
            elif u_key2 is None:
                arr2.append('')
            else:
                arr2.append(u_key2)
        ws_link.append(arr2)


len_last = 0
len_tmp = 0

last_ic = ''
last_vc = ''

file_dir_target = './input_step_2'
if not isdir(file_dir_target):
    mkdir(file_dir_target)
file_error_log = './error/log.txt'
if not isdir('./error'):
    mkdir('./error')
output_file_nm = 'result_step_2'

f_error_log = open(file_error_log, 'a')
f_error_log.write('\nSTART\n')
f_error_log.close()

dict_vendor = jhlib.load_store_db()

'''
for vc in dict_vendor:
    for ic in dict_vendor[vc]:
        for sc in dict_vendor[vc][ic]:
            if sc == 'n_upc':
                file_test.write(vc+' '+ic+' '+sc+' number:'+str(dict_vendor[vc][ic][sc])+'\n')
            else:
                for upc in dict_vendor[vc][ic][sc]:
                    file_test.write(vc+' '+ic+' '+sc+' '+upc+'\n')
file_test.close()
'''

wb_n = Workbook()
ws_n = wb_n.active
ws_n.title = 'Sheet1'

title_flag = 1
primary_key = 0
file_no = 0
total_file = len(listdir(file_dir_target))
# 여러 파일일 경우 하나씩 순회
for f in listdir(file_dir_target):
    file_targ = join(file_dir_target, f)
    file_no += 1
    # print('processing {0} of {1}\n'.format(file_no, total_file))
    if isfile(file_targ):
        file_name = f.rsplit('.', 1)[0]
        file_ext = f.rsplit('.', 1)[1]

        dict_error_flag = dict()

        dict_unique = dict()

        wb_t = load_workbook(file_targ)
        # print(file_targ)
        ws_t = wb_t['Sheet1']

        last_date = ''

        idx1 = -1
        try:
            for row in ws_t.rows:
                idx1 += 1

                if idx1 == 0:
                    if title_flag == 1:
                        arr1 = ['PK', '파일명', '분류', '세부 분류', '작업 내용', '처리 내용']
                        idx2 = 0
                        for col in row:
                            arr1.append(col.value)
                            if idx2 == 0:
                                arr1.append('UPC check')
                                arr1.append('lookup UPC')
                                arr1.append('UPC 동일 체크')
                                arr1.append('원본 UPC')
                                arr1.append('후보 UPC')
                            if idx2 == 5:
                                arr1.append('Description Lookup 기준')
                            idx2 += 1
                        arr1.append('DupFlag')
                        arr1.append('OLD CASE COST')
                        arr1.append('NEW CASE COST')
                        arr1.append('PRICE')
                        arr1.append('OLD MARGIN')
                        arr1.append('NEW MARGIN')
                        arr1.append('SRP')
                        ws_n.append(arr1)
                        title_flag = 0
                    continue
                elif idx1 == 1:
                    continue

                tmp_upc = str(row[0].value)
                if row[0].value is None:
                    tmp_upc = ''
                tmp_sc = str(row[2].value)
                if row[2].value is None:
                    tmp_sc = ''
                tmp_vc = str(row[3].value)
                if row[3].value is None:
                    tmp_vc = ''
                tmp_ic = str(row[5].value)
                if row[5].value is None:
                    tmp_ic = ''
                tmp_desc = row[6].value
                tmp_csize = ''
                tmp_qty = row[12].value
                tmp_ncost = row[9].value
                tmp_date = row[10].value
                tmp_upc_cand2 = []
                if last_date != tmp_date:
                    last_date = tmp_date
                    dict_to_file(dict_unique, dict_vendor, ws_n)
                    dict_unique = dict()
                if (tmp_vc is None or tmp_vc == '') and (tmp_ic is None or tmp_ic == '') and (tmp_sc is None or tmp_sc == '') and (tmp_upc is None or tmp_upc == ''):
                    continue
                tmp_vat = jhlib.getVAT(tmp_vc, tmp_sc)

                result_upc, result_ic, result_desc, result_csize, result_ccost, result_upc_check, result_upc_same_check, result_desc_lookup_base, result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp, tmp_upc_cand2 = jhlib.data_lookup(tmp_upc, tmp_sc, tmp_vc, tmp_ic, tmp_csize, tmp_ncost, tmp_upc_cand2, tmp_vat, dict_error_flag, dict_vendor)

                # arr1 = [PK, 파일명,
                # 0       F01, UPC check, lookup UPC, UPC 동일여부, 원본 UPC, 후보 UPC,
                # 1~3     F902, F1000, F27
                # 4       F1184(CASE)
                # 5       F26, Description Lookup 기준
                # 6       Description
                # 7       F1001
                # 8       F19
                # 9~10    F38, F39
                # 11      F90(1)
                # 12, 13  QTY, Amount
                # 14      F1122]
                arr1 = [primary_key, file_name]
                primary_key += 1
                idx2 = 0
                for col in row:
                    if idx2 == 0:
                        if result_desc_lookup_base == 'N':
                            tmp_upc_3, tmp_upc_cand = jhlib.getFormalUPC(tmp_upc, tmp_vc)
                            arr1.append(tmp_upc_3)
                        else:
                            arr1.append(result_upc)
                        arr1.append(result_upc_check)
                        arr1.append(result_upc)
                        tmp_upc_3, tmp_upc_cand = jhlib.getFormalUPC(tmp_upc, tmp_vc)
                        if tmp_upc_3 == result_upc or result_upc in tmp_upc_cand:
                            arr1.append(1)
                        else:
                            arr1.append(0)
                        arr1.append(tmp_upc)
                        arr1.append(str(tmp_upc_cand2))
                    elif idx2 == 4:
                        arr1.append('CASE')
                    elif idx2 == 5:
                        arr1.append(tmp_ic)
                        arr1.append(result_desc_lookup_base)
                    elif idx2 == 6:
                        if result_desc_lookup_base != 'N':
                            arr1.append(result_desc)
                        else:
                            arr1.append(tmp_desc)
                    elif idx2 == 7:
                        arr1.append('1')
                    elif idx2 == 8:
                        if result_desc_lookup_base != 'N':
                            arr1.append(result_csize)
                        else:
                            arr1.append(tmp_csize)
                    elif idx2 == 11:
                        arr1.append('1')
                    elif idx2 == 12 or idx2 == 13:
                        if col.value is None:
                            arr1.append(0.0)
                        else:
                            arr1.append(float(str(col.value).replace(',', '')))
                    elif idx2 == 14:
                        arr1.append(tmp_vat)
                    else:
                        arr1.append(col.value)
                    idx2 += 1
                # print('#2-3')
                u_vc = tmp_vc
                if u_vc is None:
                    u_vc = ''
                u_ic = tmp_ic
                if u_ic is None:
                    u_ic = ''
                u_upc = tmp_upc
                if u_upc is None:
                    u_upc = ''
                u_tmp_key = u_vc + u_ic + u_upc
                if (u_ic is None or u_ic == '') and (u_upc is None or u_upc == ''):
                    u_tmp_key = 'nokey_' + str(idx1)
                if u_tmp_key in dict_unique:
                    if int(arr1[21]) == 0:
                        dict_unique[u_tmp_key][20] = dict_unique[u_tmp_key][20] + arr1[20]
                        dict_unique[u_tmp_key][21] = dict_unique[u_tmp_key][21] + arr1[21]
                        dict_unique[u_tmp_key][23] = 1
                        len_tmp = len(dict_unique[u_tmp_key])
                        if len_tmp != len_last:
                            print(len_tmp)
                            len_tmp = len_last
                    else:
                        idx_u_key = 0
                        u_tmp_key2 = u_tmp_key+'_'+str(idx_u_key)
                        while u_tmp_key2 in dict_unique:
                            idx_u_key += 1
                            u_tmp_key2 = u_tmp_key+'_'+str(idx_u_key)
                        dict_unique[u_tmp_key2] = arr1
                        dict_unique[u_tmp_key2].append(2)
                        dict_unique[u_tmp_key2].append(result_old_case_cost)
                        dict_unique[u_tmp_key2].append(result_new_case_cost)
                        dict_unique[u_tmp_key2].append(result_price)
                        dict_unique[u_tmp_key2].append(result_old_margin)
                        dict_unique[u_tmp_key2].append(result_new_margin)
                        dict_unique[u_tmp_key2].append(result_srp)
                        len_tmp = len(dict_unique[u_tmp_key2])
                        if len_tmp != len_last:
                            print(len_tmp)
                            len_tmp = len_last
                else:
                    dict_unique[u_tmp_key] = arr1
                    dict_unique[u_tmp_key].append(0)
                    dict_unique[u_tmp_key].append(result_old_case_cost)
                    dict_unique[u_tmp_key].append(result_new_case_cost)
                    dict_unique[u_tmp_key].append(result_price)
                    dict_unique[u_tmp_key].append(result_old_margin)
                    dict_unique[u_tmp_key].append(result_new_margin)
                    dict_unique[u_tmp_key].append(result_srp)
                    len_tmp = len(dict_unique[u_tmp_key])
                    if len_tmp != len_last:
                        print(len_tmp)
                        len_tmp = len_last

            # print('#3')
            dict_to_file(dict_unique, dict_vendor, ws_n)
            dict_unique = dict()
            wb_t.close()

            # print('#4')

        except Exception as ex:
            f_error_log = open(file_error_log, 'a')
            f_error_log.write(f+'\t'+str(idx1)+'\t'+str(ex)+'\n')
            f_error_log.close()

if not isdir('./result_step_2'):
    mkdir('./result_step_2')
wb_n.save('./result_step_2/'+output_file_nm+'.xlsx')

'''
        for key1 in dict_error_flag:
            if key1 == 'disable':
                file_origin = open(file_processed, 'rb')
                file_error = open(join(file_dir_disable, f), 'wb')
                file_error.writelines(file_origin.readlines())
                file_origin.close()
                file_error.close()
            elif key1 == 'duplicate':
                file_origin = open(file_processed, 'rb')
                file_error = open(join(file_dir_duplicate, f), 'wb')
                file_error.writelines(file_origin.readlines())
                file_origin.close()
                file_error.close()
            elif key1 == 'new':
                file_origin = open(file_processed, 'rb')
                file_error = open(join(file_dir_new, f), 'wb')
                file_error.writelines(file_origin.readlines())
                file_origin.close()
                file_error.close()
            elif key1 == 'lessthanone':
                file_origin = open(file_processed, 'rb')
                file_error = open(join(file_dir_lessthanone, f), 'wb')
                file_error.writelines(file_origin.readlines())
                file_origin.close()
                file_error.close()
            elif key1 == 'multicode':
                file_origin = open(file_processed, 'rb')
                file_error = open(join(file_dir_multicode, f), 'wb')
                file_error.writelines(file_origin.readlines())
                file_origin.close()
                file_error.close()
'''
