import re
from openpyxl import load_workbook


def getFormalUPC(upc, vc):
    if upc is None:
        ret_upc = ''
    elif len(upc) == 12:
        ret_upc = '00' + upc[:-1]
    elif len(upc) == 13:
        ret_upc = '0' + upc[:-1]
    else:
        ret_upc = upc

    cand_upc = []
    if upc is not None and len(upc) > 1 and vc is not None:
        if vc == '0179':
            if len(upc) == 14:
                cand_upc.append('0' + upc[1:-1])
        elif vc == '0501':
            if len(upc) == 11:
                cand_upc.append('000' + upc[:-1])
        elif vc == '1004':
            cand_upc.append('00' + upc[:-1])
            cand_upc.append('0' + upc[:-1])
        elif vc == '1213':
            if len(upc) == 10:
                cand_upc.append('000' + upc)
                cand_upc.append('007' + upc)
        elif vc == '1229':
            if len(upc) == 13:
                cand_upc.append(upc)
            elif len(upc) == 11:
                cand_upc.append('00' + upc)
        elif vc == '1231':
            tmp_len = len(upc)
            suff = ''
            for tmp_idx in range(13 - tmp_len):
                suff += '0'
            cand_upc.append(suff + upc)
        elif vc == '1251':
            cand_upc.append('0' + upc)

    return ret_upc, cand_upc


def getFormalAMT(csize, ccost):
    if csize is not None and csize != '' and re.match(r"(?<![-.])\b[0-9]+\b(?!\.[0-9])", str(csize)) and ccost is not None and ccost != '' and re.match(r"(?<![-.])\b[0-9]+\b(?!\.[0-9])", str(ccost)):
        ret_amt = str(float(csize)*float(ccost))
    else:
        ret_amt = '0'
    return ret_amt


def getVAT(vc, sc):
    if vc is not None and vc == '1229':
        return 11.5
    elif sc is not None and sc in ['011', '015']:
        return 5
    else:
        return 0


def load_store_db():
    file_result = './processed_store_data/11-07-2019_DB.xlsx.xlsx'
    wb_r = load_workbook(file_result)
    ws_r = wb_r['Sheet1']

    dict_vendor = dict()

    idx1 = 0
    for row in ws_r.rows:
        # 첫 행(컬럼 명) 통과
        if idx1 == 0:
            idx1 += 1
            continue

        tmp_vc = row[1].value
        tmp_ic = row[2].value
        tmp_sc = row[0].value
        tmp_upc = row[3].value
        tmp_desc = row[5].value
        tmp_csize = row[6].value
        tmp_ccost = row[7].value
        tmp_price = row[8].value

        if tmp_vc not in dict_vendor:
            dict_vendor[tmp_vc] = dict()
        if tmp_ic not in dict_vendor[tmp_vc]:
            dict_vendor[tmp_vc][tmp_ic] = dict()
            dict_vendor[tmp_vc][tmp_ic]['upc_dict'] = dict()
        if tmp_sc not in dict_vendor[tmp_vc][tmp_ic]:
            dict_vendor[tmp_vc][tmp_ic][tmp_sc] = dict()
        if tmp_upc not in dict_vendor[tmp_vc][tmp_ic][tmp_sc]:
            dict_vendor[tmp_vc][tmp_ic][tmp_sc][tmp_upc] = dict()
            dict_vendor[tmp_vc][tmp_ic][tmp_sc][tmp_upc]['desc'] = tmp_desc
            dict_vendor[tmp_vc][tmp_ic][tmp_sc][tmp_upc]['csize'] = tmp_csize
            dict_vendor[tmp_vc][tmp_ic][tmp_sc][tmp_upc]['ccost'] = tmp_ccost
            dict_vendor[tmp_vc][tmp_ic][tmp_sc][tmp_upc]['price'] = tmp_price
            dict_vendor[tmp_vc][tmp_ic]['upc_dict'][tmp_upc] = 1

    wb_r.close()
    print('DB file is loaded')

    return dict_vendor


def load_base_db():
    file_result = './base_file/110419 INVOICE FORMAT.xlsx'
    wb_r = load_workbook(file_result)

    ws_r = wb_r['점포번호']
    dict_store = dict()
    idx1 = 0
    for row in ws_r.rows:
        # 첫 행(컬럼 명) 통과
        if idx1 == 0:
            idx1 += 1
            continue
        store_no = str(row[0].value)
        store_nm = str(row[1].value)
        store_cd = str(row[2].value)
        store_addr = str(row[3].value)
        dict_store[store_no] = dict()
        dict_store[store_no]['name'] = store_nm
        dict_store[store_no]['code'] = store_cd
        dict_store[store_no]['addr'] = store_addr

    ws_r = wb_r['업체별번호']
    dict_vendor = dict()
    idx1 = 0
    for row in ws_r.rows:
        # 첫 행(컬럼 명) 통과
        if idx1 == 0:
            idx1 += 1
            continue
        vendor_no = row[0].value
        vendor_nm = row[1].value
        vendor_pre = row[2].value
        dict_vendor[vendor_no] = dict()
        dict_vendor[vendor_no]['name'] = vendor_nm
        dict_vendor[vendor_no]['prefix'] = vendor_pre

    wb_r.close()

    print('Base DB file is loaded')

    return dict_store, dict_vendor


def data_lookup(tmp_upc, tmp_sc, tmp_vc, tmp_ic, tmp_csize, tmp_ncost, tmp_upc_cand2, tmp_vat, dict_error_flag, dict_vendor):
    # print('#2')

    result_upc = ''
    result_ic = ''
    result_desc = ''
    result_csize = ''
    result_ccost = ''
    result_upc_check = '0'
    result_upc_same_check = '0'
    result_desc_lookup_base = 'N'
    result_old_case_cost = ''
    result_new_case_cost = ''
    result_price = ''
    result_old_margin = ''
    result_new_margin = ''
    result_srp = ''

    if tmp_csize is None or tmp_csize == '' or float(tmp_csize) <= 0:
        dict_error_flag['lessthanone'] = 0

    # L0 업체 코드 없는 경우
    if tmp_vc is None:
        dict_error_flag['disable'] = 0

    # L0 업체 코드 있고 매장 데이터에 등록되어있는 경우
    elif tmp_vc in dict_vendor:

        # L1 아이템 코드가 인보이스에 없는 경우
        if tmp_ic is None:

            # L1 UPC도 없는 경우
            if tmp_upc is None:
                dict_error_flag['disable'] = 0

            # L2 UPC 코드는 입력된 경우
            else:

                if tmp_sc is None:
                    tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                    # UPC 기준으로 탐색
                    flag_exist_upc = False
                    flag_exist_upc_cand = False
                    flag_exist_exact = False
                    tmp_ic2 = ''
                    last_sc = ''
                    idx_upc_cand = -1
                    for key1 in dict_vendor[tmp_vc]:
                        for key2 in dict_vendor[tmp_vc][key1]:
                            if key2 == 'upc_dict':
                                continue
                            for key3 in dict_vendor[tmp_vc][key1][key2]:
                                if key3 == tmp_upc2:
                                    flag_exist_upc = True
                                    tmp_ic2 = key1
                                    last_sc = key2
                                    break
                                elif key3 in tmp_upc_cand:
                                    flag_exist_upc_cand = True
                                    idx_upc_cand = tmp_upc_cand.index(key3)
                                    tmp_ic2 = key1
                                    last_sc = key2
                                    break
                            if flag_exist_upc or flag_exist_upc_cand:
                                break
                        if flag_exist_upc or flag_exist_upc_cand:
                            break

                    # UPC가 없는 경우
                    if not flag_exist_upc and not flag_exist_upc_cand:
                        dict_error_flag['new'] = 1
                    else:
                        result_desc_lookup_base = 'U'
                        if flag_exist_upc:
                            result_upc = tmp_upc2
                        else:
                            result_upc = tmp_upc_cand[idx_upc_cand]
                        result_desc = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['desc']
                        result_csize = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['csize']
                        result_old_case_cost = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['ccost']
                        result_new_case_cost = tmp_ncost
                        result_price = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['price']
                        if result_csize is None or result_old_case_cost is None or result_new_case_cost is None or result_price is None:
                            result_old_case_cost = ''
                            result_new_case_cost = ''
                            result_price = ''
                            result_old_margin = ''
                            result_new_margin = ''
                            result_new_srp = ''
                        else:
                            result_old_margin = (float(result_price) - (
                                        (float(result_old_case_cost) / float(result_csize)) * (
                                            1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                            result_new_margin = (float(result_price) - (
                                        (float(result_new_case_cost) / float(result_csize)) * (
                                            1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                            if result_old_case_cost < float(result_new_case_cost):
                                result_srp = round(((float(result_new_case_cost) / float(result_csize)) * (
                                            1 + (float(tmp_vat) / 100.0))) / (1.0 - float(result_old_margin)), 1) - 0.01
                            else:
                                result_srp = ''

                else:
                    tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                    # UPC 기준으로 탐색
                    flag_exist_upc = False
                    flag_exist_upc_cand = False
                    flag_exist_exact = False
                    tmp_ic2 = ''
                    last_sc = ''
                    idx_upc_cand = -1
                    for key1 in dict_vendor[tmp_vc]:
                        for key2 in dict_vendor[tmp_vc][key1]:
                            if key2 == 'upc_dict':
                                continue
                            for key3 in dict_vendor[tmp_vc][key1][key2]:
                                if key3 == tmp_upc2:
                                    flag_exist_upc = True
                                    tmp_ic2 = key1
                                    last_sc = key2
                                    if key2 == tmp_sc:
                                        flag_exist_exact = True
                                        break
                                elif key3 in tmp_upc_cand:
                                    flag_exist_upc_cand = True
                                    idx_upc_cand = tmp_upc_cand.index(key3)
                                    tmp_ic2 = key1
                                    last_sc = key2
                                    if key2 == tmp_sc:
                                        flag_exist_exact = True
                                        break
                            if flag_exist_exact:
                                break
                        if flag_exist_exact:
                            break

                    # UPC가 없는 경우
                    if not flag_exist_upc and not flag_exist_upc_cand:
                        dict_error_flag['new'] = 1
                    else:
                        result_desc_lookup_base = 'U'
                        if flag_exist_upc:
                            result_upc = tmp_upc2
                        else:
                            result_upc = tmp_upc_cand[idx_upc_cand]
                        result_desc = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['desc']
                        result_csize = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['csize']
                        result_old_case_cost = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['ccost']
                        result_new_case_cost = tmp_ncost
                        result_price = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['price']
                        if result_csize is None or result_old_case_cost is None or result_new_case_cost is None or result_price is None:
                            result_old_case_cost = ''
                            result_new_case_cost = ''
                            result_price = ''
                            result_old_margin = ''
                            result_new_margin = ''
                            result_new_srp = ''
                        else:
                            result_old_margin = (float(result_price) - (
                                        (float(result_old_case_cost) / float(result_csize)) * (
                                            1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                            result_new_margin = (float(result_price) - (
                                        (float(result_new_case_cost) / float(result_csize)) * (
                                            1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                            if result_old_case_cost < float(result_new_case_cost):
                                result_srp = round(((float(result_new_case_cost) / float(result_csize)) * (
                                            1 + (float(tmp_vat) / 100.0))) / (1.0 - float(result_old_margin)), 1) - 0.01
                            else:
                                result_srp = ''

        # L1 기 등록된 상품 (Vendor Code & Item Code 기준)
        elif tmp_ic in dict_vendor[tmp_vc]:
            # print('#2-1')
            # print(dict_vendor[tmp_vc][tmp_ic])
            # print(tmp_sc)

            # L2 Store Code가 없는 경우
            if tmp_sc is None or tmp_sc not in dict_vendor[tmp_vc][tmp_ic]:

                # UPC is sole
                if len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']) == 1:
                    # print('#2-1-1-1')
                    result_upc_check = '1'
                    itr = iter(dict_vendor[tmp_vc][tmp_ic])
                    next(itr)
                    tmp_sc2 = next(itr)
                    result_upc = next(iter(dict_vendor[tmp_vc][tmp_ic][tmp_sc2]))
                    result_desc_lookup_base = 'I'
                    result_desc = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['desc']
                    result_csize = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['csize']
                    result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['ccost']
                    result_new_case_cost = tmp_ncost
                    result_price = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['price']
                    if result_csize is None or result_old_case_cost is None or result_new_case_cost is None or result_price is None:
                        result_old_case_cost = ''
                        result_new_case_cost = ''
                        result_price = ''
                        result_old_margin = ''
                        result_new_margin = ''
                        result_new_srp = ''
                    else:
                        result_old_margin = (float(result_price) - (
                                    (float(result_old_case_cost) / float(result_csize)) * (
                                        1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                        result_new_margin = (float(result_price) - (
                                    (float(result_new_case_cost) / float(result_csize)) * (
                                        1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                        if result_old_case_cost < float(result_new_case_cost):
                            result_srp = round(((float(result_new_case_cost) / float(result_csize)) * (
                                        1 + (float(tmp_vat) / 100.0))) / (1.0 - float(result_old_margin)), 1) - 0.01
                        else:
                            result_srp = ''
                    tmp_upc_3, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)
                    if result_upc == tmp_upc_3 or result_upc in tmp_upc_cand:
                        result_upc_same_check = 1
                    else:
                        result_upc_same_check = 0

                # UPC is duplicated
                elif len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']) > 1:
                    # print('#2-1-1-2')
                    result_upc_check = str(len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']))
                    # print('1111')

                    if tmp_upc is None:
                        # print('2222')
                        dict_error_flag['duplicate'] = 1

                    else:
                        tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)
                        # print('3333')

                        # UPC 기준으로 탐색
                        flag_exist_upc = False
                        flag_exist_upc_cand = False
                        last_sc = ''
                        last_sc_cand = ''
                        idx_upc_cand = -1
                        for key1 in dict_vendor[tmp_vc][tmp_ic]:
                            if key1 == 'upc_dict':
                                continue
                            for key2 in dict_vendor[tmp_vc][tmp_ic][key1]:
                                if key2 == tmp_upc2:
                                    flag_exist_upc = True
                                    last_sc = key1
                                    break
                                if key2 in tmp_upc_cand:
                                    flag_exist_upc_cand = True
                                    idx_upc_cand = tmp_upc_cand.index(key2)
                                    last_sc_cand = key1
                                    break
                            if flag_exist_upc or flag_exist_upc_cand:
                                break
                        # print('4444')

                        # UPC가 없는 경우
                        if not flag_exist_upc and not flag_exist_upc_cand:
                            # print('5555')
                            tmp_upc_cand_list = []
                            for cand in dict_vendor[tmp_vc][tmp_ic]['upc_dict']:
                                tmp_upc_cand_list.append(cand)
                            tmp_upc_cand2.extend(tmp_upc_cand_list)
                            dict_error_flag['new'] = 1
                        else:
                            # print('6666')
                            # print(dict_vendor[tmp_vc][tmp_ic])
                            # print(last_sc)
                            result_desc_lookup_base = 'U'
                            # print('7777')
                            # print(tmp_upc_cand)
                            # print(idx_upc_cand)
                            if flag_exist_upc:
                                result_upc = tmp_upc2
                            else:
                                result_upc = tmp_upc_cand[idx_upc_cand]
                                last_sc = last_sc_cand
                            # print(result_upc)
                            # print('8888')
                            result_desc = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['desc']
                            # print('9999')
                            result_csize = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['csize']
                            # print('0000')
                            result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['ccost']
                            result_new_case_cost = tmp_ncost
                            result_price = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['price']
                            if result_csize is None or result_old_case_cost is None or result_new_case_cost is None or result_price is None:
                                result_old_case_cost = ''
                                result_new_case_cost = ''
                                result_price = ''
                                result_old_margin = ''
                                result_new_margin = ''
                                result_new_srp = ''
                            else:
                                result_old_margin = (float(result_price) - (
                                            (float(result_old_case_cost) / float(result_csize)) * (
                                                1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                                result_new_margin = (float(result_price) - (
                                            (float(result_new_case_cost) / float(result_csize)) * (
                                                1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                                if result_old_case_cost < float(result_new_case_cost):
                                    result_srp = round(((float(result_new_case_cost) / float(result_csize)) * (
                                                1 + (float(tmp_vat) / 100.0))) / (1.0 - float(result_old_margin)),
                                                       1) - 0.01
                                else:
                                    result_srp = ' '

            # L2 Store Code가 있는 경우
            elif tmp_sc is not None and tmp_sc in dict_vendor[tmp_vc][tmp_ic]:

                # Store Code에 중복없이 하나의 UPC만 있는 경우
                if len(dict_vendor[tmp_vc][tmp_ic][tmp_sc]) == 1:
                    # print('#2-1-2-1')
                    result_upc = next(iter(dict_vendor[tmp_vc][tmp_ic][tmp_sc]))
                    result_desc = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['desc']
                    result_csize = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['csize']
                    result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['ccost']
                    result_new_case_cost = tmp_ncost
                    result_price = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['price']
                    if result_csize is None or result_old_case_cost is None or result_new_case_cost is None or result_price is None:
                        result_old_case_cost = ''
                        result_new_case_cost = ''
                        result_price = ''
                        result_old_margin = ''
                        result_new_margin = ''
                        result_new_srp = ''
                    else:
                        result_old_margin = (float(result_price) - (
                                    (float(result_old_case_cost) / float(result_csize)) * (
                                        1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                        result_new_margin = (float(result_price) - (
                                    (float(result_new_case_cost) / float(result_csize)) * (
                                        1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                        if result_old_case_cost < float(result_new_case_cost):
                            result_srp = round(((float(result_new_case_cost) / float(result_csize)) * (
                                        1 + (float(tmp_vat) / 100.0))) / (1.0 - float(result_old_margin)), 1) - 0.01
                        else:
                            result_srp = ' '
                    result_upc_check = '1'
                    result_desc_lookup_base = 'I'

                    tmp_upc_3, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)
                    if tmp_upc_3 == result_upc or result_upc in tmp_upc_cand:
                        result_upc_same_check = '1'

                elif len(dict_vendor[tmp_vc][tmp_ic][tmp_sc]) > 1:
                    # print('#2-1-2-2')
                    result_upc_check = str(len(dict_vendor[tmp_vc][tmp_ic][tmp_sc]))

                    if tmp_upc is not None:
                        tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                        # UPC 기준으로 탐색
                        flag_exist_upc = False
                        flag_exist_upc_cand = False
                        flag_exist_exact = False
                        tmp_ic2 = ''
                        last_sc = ''
                        idx_upc_cand = -1
                        for key1 in dict_vendor[tmp_vc]:
                            for key2 in dict_vendor[tmp_vc][key1]:
                                if key2 == 'upc_dict':
                                    continue
                                for key3 in dict_vendor[tmp_vc][key1][key2]:
                                    if key3 == tmp_upc2:
                                        flag_exist_upc = True
                                        tmp_ic2 = key1
                                        last_sc = key2
                                        if key2 == tmp_sc:
                                            flag_exist_exact = True
                                            break
                                    if key3 in tmp_upc_cand:
                                        flag_exist_upc_cand = True
                                        idx_upc_cand = tmp_upc_cand.index(key3)
                                        tmp_ic2 = key1
                                        last_sc = key2
                                        if key2 == tmp_sc:
                                            flag_exist_exact = True
                                            break
                                if flag_exist_exact:
                                    break
                            if flag_exist_exact:
                                break

                        # UPC가 Store 내에 없는 경우
                        if not flag_exist_upc and not flag_exist_upc_cand:
                            tmp_upc_cand_list = []
                            for cand in dict_vendor[tmp_vc][tmp_ic][tmp_sc]:
                                tmp_upc_cand_list.append(cand)
                            tmp_upc_cand2.extend(tmp_upc_cand_list)
                            if len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']) > 0:

                                # UPC is sole
                                if len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']) == 1:
                                    result_upc_check = '1'
                                    itr = iter(dict_vendor[tmp_vc][tmp_ic])
                                    next(itr)
                                    tmp_sc2 = next(itr)
                                    result_upc = next(iter(dict_vendor[tmp_vc][tmp_ic][tmp_sc2]))
                                    result_desc_lookup_base = 'I'
                                    result_desc = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['desc']
                                    result_csize = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['csize']
                                    result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['ccost']
                                    result_new_case_cost = tmp_ncost
                                    result_price = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['price']
                                    if result_csize is None or result_old_case_cost is None or result_new_case_cost is None or result_price is None:
                                        result_old_case_cost = ''
                                        result_new_case_cost = ''
                                        result_price = ''
                                        result_old_margin = ''
                                        result_new_margin = ''
                                        result_new_srp = ''
                                    else:
                                        result_old_margin = (float(result_price) - (
                                                    (float(result_old_case_cost) / float(result_csize)) * (
                                                        1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                                        result_new_margin = (float(result_price) - (
                                                    (float(result_new_case_cost) / float(result_csize)) * (
                                                        1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                                        if result_old_case_cost < float(result_new_case_cost):
                                            result_srp = round(((float(result_new_case_cost) / float(result_csize)) * (
                                                        1 + (float(tmp_vat) / 100.0))) / (
                                                                           1.0 - float(result_old_margin)), 1) - 0.01
                                        else:
                                            result_srp = ' '
                                    tmp_upc3, tmp_upc_cand = getFormalUPC(result_upc, tmp_vc)
                                    if tmp_upc == result_upc or result_upc in tmp_upc_cand:
                                        result_upc_same_check = 1
                                    else:
                                        result_upc_same_check = 0

                                # UPC is duplicated
                                elif len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']) > 1:

                                    if tmp_upc is None:
                                        dict_error_flag['duplicate'] = 0

                                    else:
                                        tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                                        # UPC 기준으로 탐색
                                        flag_exist_upc = False
                                        flag_exist_upc_cand = False
                                        last_sc = ''
                                        idx_upc_cand = -1
                                        for key1 in dict_vendor[tmp_vc][tmp_ic]:
                                            if key1 == 'upc_dict':
                                                continue
                                            for key2 in dict_vendor[tmp_vc][tmp_ic][key1]:
                                                if key2 == tmp_upc2:
                                                    flag_exist_upc = True
                                                    last_sc = key1
                                                    break
                                                if key2 in tmp_upc_cand:
                                                    flag_exist_upc_cand = True
                                                    idx_upc_cand = tmp_upc_cand.index(key2)
                                                    last_sc = key1
                                                    break
                                            if flag_exist_upc or flag_exist_upc_cand:
                                                break

                                        # UPC가 없는 경우
                                        if not flag_exist_upc and not flag_exist_upc_cand:
                                            tmp_upc_cand_list = []
                                            for cand in dict_vendor[tmp_vc][tmp_ic]['upc_dict']:
                                                tmp_upc_cand_list.append(cand)
                                            tmp_upc_cand2.extend(tmp_upc_cand_list)
                                            dict_error_flag['new'] = 1
                                        else:
                                            result_desc_lookup_base = 'U'
                                            if flag_exist_upc:
                                                result_upc = tmp_upc2
                                            else:
                                                result_upc = tmp_upc_cand[idx_upc_cand]
                                            result_desc = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['desc']
                                            result_csize = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['csize']
                                            result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc][
                                                'ccost']
                                            result_new_case_cost = tmp_ncost
                                            result_price = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['price']
                                            if result_csize is None or result_old_case_cost is None or result_new_case_cost is None or result_price is None:
                                                result_old_case_cost = ''
                                                result_new_case_cost = ''
                                                result_price = ''
                                                result_old_margin = ''
                                                result_new_margin = ''
                                                result_new_srp = ''
                                            else:
                                                result_old_margin = (float(result_price) - (
                                                            (float(result_old_case_cost) / float(result_csize)) * (
                                                                1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                                                result_new_margin = (float(result_price) - (
                                                            (float(result_new_case_cost) / float(result_csize)) * (
                                                                1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                                                if result_old_case_cost < float(result_new_case_cost):
                                                    result_srp = round(((float(result_new_case_cost) / float(
                                                        result_csize)) * (1 + (float(tmp_vat) / 100.0))) / (
                                                                                   1.0 - float(result_old_margin)),
                                                                       1) - 0.01
                                                else:
                                                    result_srp = ' '
                                else:
                                    dict_error_flag['new'] = 1
                        else:
                            result_desc_lookup_base = 'U'
                            if flag_exist_upc:
                                result_upc = tmp_upc2
                            else:
                                result_upc = tmp_upc_cand[idx_upc_cand]
                            result_desc = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['desc']
                            result_csize = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['csize']
                            result_old_case_cost = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['ccost']
                            result_new_case_cost = tmp_ncost
                            result_price = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['price']
                            if result_csize is None or result_old_case_cost is None or result_new_case_cost is None or result_price is None:
                                result_old_case_cost = ''
                                result_new_case_cost = ''
                                result_price = ''
                                result_old_margin = ''
                                result_new_margin = ''
                                result_new_srp = ''
                            else:
                                result_old_margin = (float(result_price) - (
                                            (float(result_old_case_cost) / float(result_csize)) * (
                                                1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                                result_new_margin = (float(result_price) - (
                                            (float(result_new_case_cost) / float(result_csize)) * (
                                                1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                                if result_old_case_cost < float(result_new_case_cost):
                                    result_srp = round(((float(result_new_case_cost) / float(result_csize)) * (
                                                1 + (float(tmp_vat) / 100.0))) / (1.0 - float(result_old_margin)),
                                                       1) - 0.01
                                else:
                                    result_srp = ' '

                    else:
                        dict_error_flag['duplicate'] = 1

                # Store Code에는 없고 Item Code로는 UPC가 조회된 경우
                else:
                    # print('#2-1-2-3')
                    if len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']) > 0:

                        # UPC is sole
                        if len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']) == 1:
                            result_upc_check = '1'
                            itr = iter(dict_vendor[tmp_vc][tmp_ic])
                            next(itr)
                            tmp_sc2 = next(itr)
                            result_upc = next(iter(dict_vendor[tmp_vc][tmp_ic][tmp_sc2]))
                            result_desc_lookup_base = 'I'
                            result_desc = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['desc']
                            result_csize = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['csize']
                            result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['ccost']
                            result_new_case_cost = tmp_ncost
                            result_price = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['price']
                            if result_csize is None or result_old_case_cost is None or result_new_case_cost is None or result_price is None:
                                result_old_case_cost = ''
                                result_new_case_cost = ''
                                result_price = ''
                                result_old_margin = ''
                                result_new_margin = ''
                                result_new_srp = ''
                            else:
                                result_old_margin = (float(result_price) - (
                                            (float(result_old_case_cost) / float(result_csize)) * (
                                                1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                                result_new_margin = (float(result_price) - (
                                            (float(result_new_case_cost) / float(result_csize)) * (
                                                1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                                if result_old_case_cost < float(result_new_case_cost):
                                    result_srp = round(((float(result_new_case_cost) / float(result_csize)) * (
                                                1 + (float(tmp_vat) / 100.0))) / (1.0 - float(result_old_margin)),
                                                       1) - 0.01
                                else:
                                    result_srp = ' '
                            tmp_upc3, tmp_upc_cand = getFormalUPC(result_upc, tmp_vc)
                            if tmp_upc == result_upc or result_upc in tmp_upc_cand:
                                result_upc_same_check = 1
                            else:
                                result_upc_same_check = 0

                        # UPC is duplicated
                        elif len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']) > 1:

                            if tmp_upc is None:
                                dict_error_flag['duplicate'] = 0

                            else:
                                tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                                # UPC 기준으로 탐색
                                flag_exist_upc = False
                                flag_exist_upc_cand = False
                                last_sc = ''
                                idx_upc_cand = -1
                                for key1 in dict_vendor[tmp_vc][tmp_ic]:
                                    if key1 == 'upc_dict':
                                        continue
                                    for key2 in dict_vendor[tmp_vc][tmp_ic][key1]:
                                        if key2 == tmp_upc2:
                                            flag_exist_upc = True
                                            last_sc = key1
                                            break
                                        if key2 in tmp_upc_cand:
                                            flag_exist_upc_cand = True
                                            idx_upc_cand = tmp_upc_cand.index(key2)
                                            last_sc = key1
                                            break
                                    if flag_exist_upc or flag_exist_upc_cand:
                                        break

                                # UPC가 없는 경우
                                if not flag_exist_upc and not flag_exist_upc_cand:
                                    tmp_upc_cand_list = []
                                    for cand in dict_vendor[tmp_vc][tmp_ic]['upc_dict']:
                                        tmp_upc_cand_list.append(cand)
                                    tmp_upc_cand2.extend(tmp_upc_cand_list)
                                    dict_error_flag['new'] = 1
                                else:
                                    result_desc_lookup_base = 'U'
                                    if flag_exist_upc:
                                        result_upc = tmp_upc2
                                    else:
                                        result_upc = tmp_upc_cand[idx_upc_cand]
                                    result_desc = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['desc']
                                    result_csize = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['csize']
                                    result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['ccost']
                                    result_new_case_cost = tmp_ncost
                                    result_price = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['price']
                                    if result_csize is None or result_old_case_cost is None or result_new_case_cost is None or result_price is None:
                                        result_old_case_cost = ''
                                        result_new_case_cost = ''
                                        result_price = ''
                                        result_old_margin = ''
                                        result_new_margin = ''
                                        result_new_srp = ''
                                    else:
                                        result_old_margin = (float(result_price) - (
                                                    (float(result_old_case_cost) / float(result_csize)) * (
                                                        1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                                        result_new_margin = (float(result_price) - (
                                                    (float(result_new_case_cost) / float(result_csize)) * (
                                                        1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                                        if result_old_case_cost < float(result_new_case_cost):
                                            result_srp = round(((float(result_new_case_cost) / float(result_csize)) * (
                                                        1 + (float(tmp_vat) / 100.0))) / (
                                                                           1.0 - float(result_old_margin)), 1) - 0.01
                                        else:
                                            result_srp = ' '
                        else:
                            dict_error_flag['new'] = 1


        # L1 미 등록된 상품 중 인보이스에 UPC가 있는 경우 (신상품)
        elif tmp_upc is not None:
            # print('#2-2')

            if tmp_sc is None:
                tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                # UPC 기준으로 탐색
                flag_exist_upc = False
                flag_exist_upc_cand = False
                tmp_ic2 = ''
                last_sc = ''
                idx_upc_cand = -1
                for key1 in dict_vendor[tmp_vc]:
                    for key2 in dict_vendor[tmp_vc][key1]:
                        if key2 == 'upc_dict':
                            continue
                        for key3 in dict_vendor[tmp_vc][key1][key2]:
                            if key3 == tmp_upc2:
                                flag_exist_upc = True
                                tmp_ic2 = key1
                                last_sc = key2
                                break
                            if key3 in tmp_upc_cand:
                                flag_exist_upc_cand = True
                                idx_upc_cand = tmp_upc_cand.index(key3)
                                tmp_ic2 = key1
                                last_sc = key2
                                break
                        if flag_exist_upc or flag_exist_upc_cand:
                            break
                    if flag_exist_upc or flag_exist_upc_cand:
                        break

                # UPC가 없는 경우
                if not flag_exist_upc and not flag_exist_upc_cand:
                    dict_error_flag['new'] = 1
                else:
                    result_desc_lookup_base = 'U'
                    if flag_exist_upc:
                        result_upc = tmp_upc2
                    else:
                        result_upc = tmp_upc_cand[idx_upc_cand]
                    result_desc = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['desc']
                    result_csize = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['csize']
                    result_old_case_cost = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['ccost']
                    result_new_case_cost = tmp_ncost
                    result_price = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['price']
                    if result_csize is None or result_old_case_cost is None or result_new_case_cost is None or result_price is None:
                        result_old_case_cost = ''
                        result_new_case_cost = ''
                        result_price = ''
                        result_old_margin = ''
                        result_new_margin = ''
                        result_new_srp = ''
                    else:
                        result_old_margin = (float(result_price) - (
                                    (float(result_old_case_cost) / float(result_csize)) * (
                                        1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                        result_new_margin = (float(result_price) - (
                                    (float(result_new_case_cost) / float(result_csize)) * (
                                        1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                        if result_old_case_cost < float(result_new_case_cost):
                            result_srp = round(((float(result_new_case_cost) / float(result_csize)) * (
                                        1 + (float(tmp_vat) / 100.0))) / (1.0 - float(result_old_margin)), 1) - 0.01
                        else:
                            result_srp = ' '

            else:
                tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                # UPC 기준으로 탐색
                flag_exist_upc = False
                flag_exist_upc_cand = False
                flag_exist_exact = False
                tmp_ic2 = ''
                last_sc = ''
                tmp_ic2_cand = ''
                last_sc_cand = ''
                idx_upc_cand = -1
                # print('#2-2-1')
                for key1 in dict_vendor[tmp_vc]:
                    for key2 in dict_vendor[tmp_vc][key1]:
                        if key2 == 'upc_dict':
                            continue
                        for key3 in dict_vendor[tmp_vc][key1][key2]:
                            if key3 == tmp_upc2:
                                flag_exist_upc = True
                                tmp_ic2 = key1
                                last_sc = key2
                                if key2 == tmp_sc:
                                    flag_exist_exact = True
                                    break
                            if key3 in tmp_upc_cand:
                                flag_exist_upc_cand = True
                                idx_upc_cand = tmp_upc_cand.index(key3)
                                tmp_ic2_cand = key1
                                last_sc_cand = key2
                                if key2 == tmp_sc:
                                    flag_exist_exact = True
                                    break
                        if flag_exist_exact:
                            break
                    if flag_exist_exact:
                        break

                # print('#2-2-2')
                # UPC가 없는 경우
                if not flag_exist_upc and not flag_exist_upc_cand:
                    dict_error_flag['new'] = 1
                else:
                    result_desc_lookup_base = 'U'
                    # print(flag_exist_upc)
                    # print(tmp_upc2)
                    # print(idx_upc_cand)
                    # print(tmp_upc_cand)
                    if flag_exist_upc:
                        result_upc = tmp_upc2
                    else:
                        result_upc = tmp_upc_cand[idx_upc_cand]
                        tmp_ic2 = tmp_ic2_cand
                        last_sc = last_sc_cand
                    # print(result_upc)
                    # print(last_sc)
                    # print(dict_vendor[tmp_vc][tmp_ic2])
                    result_desc = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['desc']
                    result_csize = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['csize']
                    result_old_case_cost = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['ccost']
                    result_new_case_cost = tmp_ncost
                    result_price = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['price']
                    if result_csize is None or result_old_case_cost is None or result_new_case_cost is None or result_price is None:
                        result_old_case_cost = ''
                        result_new_case_cost = ''
                        result_price = ''
                        result_old_margin = ''
                        result_new_margin = ''
                        result_new_srp = ''
                    else:
                        result_old_margin = (float(result_price) - (
                                    (float(result_old_case_cost) / float(result_csize)) * (
                                        1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                        result_new_margin = (float(result_price) - (
                                    (float(result_new_case_cost) / float(result_csize)) * (
                                        1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
                        # print(result_new_case_cost)
                        if result_old_case_cost < float(result_new_case_cost):
                            result_srp = round(((float(result_new_case_cost) / float(result_csize)) * (
                                        1 + (float(tmp_vat) / 100.0))) / (1.0 - float(result_old_margin)), 1) - 0.01
                        else:
                            result_srp = ' '
                # print('#2-2-3')

    # print('#2-3')

    return result_upc, result_ic, result_desc, result_csize, result_ccost, result_upc_check, result_upc_same_check, result_desc_lookup_base, result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp, tmp_upc_cand2

