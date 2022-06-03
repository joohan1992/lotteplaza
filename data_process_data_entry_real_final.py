from openpyxl import load_workbook, Workbook
from os import listdir, mkdir
from os.path import isfile, join, isdir
import data_process_data_entry_lib_8 as jhlib


list_item = []
dict_srp_store_date = dict()
header = []


dict_vendor = jhlib.load_store_db()
dict_store, dict_vend = jhlib.load_base_db()

output_file_nm = 'final_result_1108'

file_dir_target = './input_step_4'
idx0 = -1
for f in listdir(file_dir_target):
    idx0 += 1
    file_targ = join(file_dir_target, f)
    if isfile(file_targ):
        wb_t = load_workbook(file_targ)
        ws_t = wb_t['Sheet1']
        idx1 = -1
        print(f)
        for row in ws_t.rows:
            idx1 += 1
            print('idx1: '+str(idx1))
            # print(idx1, sep=' ')
            if idx1 == 0:
                print('\tidx2: '+str(idx0))
                if idx0 == 0:
                    for col in row:
                        header.append(col.value)
                continue
            idx2 = -1
            arr1 = []
            for col in row:
                arr1.append(col.value)

            tmp_upc = str(row[10].value)
            if row[10].value is None:
                tmp_upc = ''
            result_upc = str(row[6].value)
            if row[6].value is None:
                result_upc = ''
            tmp_sc = str(row[13].value)
            if row[13].value is None:
                tmp_sc = ''
            tmp_vc = str(row[14].value)
            if row[14].value is None:
                tmp_vc = ''
            tmp_ic = str(row[16].value)
            if row[16].value is None:
                tmp_ic = ''
            tmp_desc = row[18].value
            tmp_csize = ''
            tmp_qty = row[24].value
            tmp_ncost = row[21].value
            tmp_date = row[22].value
            tmp_upc_cand2 = []
            tmp_vat = jhlib.getVAT(tmp_vc, tmp_sc)
            tmp_srp = row[33].value
            result_old_case_cost = row[28].value
            if row[28].value is None:
                result_old_case_cost = ''
            result_new_case_cost = row[29].value
            if row[29].value is None:
                result_new_case_cost = ''
            result_price = row[30].value
            if row[30].value is None:
                result_price = ''
            result_old_margin = row[31].value
            if row[31].value is None:
                result_old_margin = ''
            result_new_margin = row[32].value
            if row[32].value is None:
                result_new_margin = ''

            # 처리 내용이 있으면 다시 lookup
            if arr1[5] is not None and arr1[5] != '':
                if not((tmp_vc is None or tmp_vc == '') and (tmp_ic is None or tmp_ic == '') and (tmp_sc is None or tmp_sc == '') and (tmp_upc is None or tmp_upc == '')):
                    dict_error_flag = dict()
                    result_upc, result_ic, result_desc, result_csize, result_ccost, result_upc_check, result_upc_same_check, result_desc_lookup_base, result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp, tmp_upc_cand2 = jhlib.data_lookup(tmp_upc, tmp_sc, tmp_vc, tmp_ic, tmp_csize, tmp_ncost, tmp_upc_cand2, tmp_vat, dict_error_flag, dict_vendor)

                    if result_desc_lookup_base == 'N':
                        tmp_upc_3, tmp_upc_cand = jhlib.getFormalUPC(tmp_upc, tmp_vc)
                        arr1[6] = tmp_upc_3
                    else:
                        arr1[6] = result_upc
                    arr1[7] = result_upc_check
                    arr1[8] = result_upc
                    tmp_upc_3, tmp_upc_cand = jhlib.getFormalUPC(tmp_upc, tmp_vc)
                    if tmp_upc_3 == result_upc or result_upc in tmp_upc_cand:
                        arr1[9] = 1
                    else:
                        arr1[9] = 0
                    arr1[10] = tmp_upc
                    arr1[11] = str(tmp_upc_cand2)
                    arr1[13] = tmp_sc
                    arr1[14] = tmp_vc
                    arr1[16] = tmp_ic
                    arr1[26] = tmp_vat

                    if result_desc_lookup_base != 'N':
                        arr1[17] = result_desc_lookup_base
                        arr1[18] = result_desc
                        arr1[20] = result_csize
                        if len(arr1) < 34:
                            for idx3 in range(34-len(arr1)):
                                arr1.append(' ')
                        arr1[28] = result_old_case_cost
                        arr1[29] = result_new_case_cost
                        arr1[30] = result_price
                        arr1[31] = result_old_margin
                        arr1[32] = result_new_margin
                        arr1[33] = result_srp

                    if result_desc_lookup_base == 'N':
                        if int(result_upc_check) > 1 and (tmp_upc is None or tmp_upc == ''):
                            arr1[2] = '업체+제품 검색 시 다수이나 인보이스 내 UPC 없음'
                            arr1[3] = ''
                            arr1[4] = 'Description 및 F19 입력'
                        elif (tmp_ic is None or tmp_ic == '') and (tmp_upc is None or tmp_upc  == ''):
                            arr1[2] = '인보이스에 제품코드와 UPC 둘 다 없음'
                            arr1[3] = ''
                            arr1[4] = '제품 코드, UPC 없는지 확인 후 Description 및 F19 입력'
                        elif tmp_vc not in dict_vendor:
                            arr1[2] = '매장 데이터에 등록되지 않은 업체 코드'
                            arr1[3] = ''
                            arr1[4] = '업체 코드 재확인 후 Description 및 F19 입력'
                        elif int(result_upc_check) == 0:
                            arr1[2] = '신상품'
                            arr1[3] = ''
                            arr1[4] = '제품 코드 및 업체 코드 재확인 후 Description 및 F19 입력'
                        else:
                            arr1[2] = '업체+제품 검색 시 다수이나 인보이스의 UPC와 상이함'
                            arr1[3] = ''
                            arr1[4] = 'UPC 재확인 후 Description 및 F19 입력'
                    else:
                        if result_desc_lookup_base == 'I' and result_upc_same_check == 0 and tmp_upc is not None and tmp_upc != '':
                            arr1[2] = '업체+제품으로 검색된 UPC가 단일이지만 인보이스의 UPC와 상이'
                            arr1[3] = ''
                            arr1[4] = 'UPC 재확인'
                        else:
                            arr1[2] = ''
                            arr1[3] = ''
                            arr1[4] = ''

                    if result_srp != None and result_srp != '' and result_srp != ' ':
                        if str(tmp_sc)+'_'+str(tmp_date) not in dict_srp_store_date:
                            dict_srp_store_date[str(tmp_sc)+'_'+str(tmp_date)] = []
                        dict_srp_store_date[str(tmp_sc)+'_'+str(tmp_date)].append([result_upc, '', tmp_sc, '1', result_srp, '1', tmp_vc, tmp_ic, tmp_date, result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin])
            else:
                if tmp_srp != None and tmp_srp != '' and tmp_srp != ' ':
                    if str(tmp_sc)+'_'+str(tmp_date) not in dict_srp_store_date:
                        dict_srp_store_date[str(tmp_sc)+'_'+str(tmp_date)] = []
                    dict_srp_store_date[str(tmp_sc)+'_'+str(tmp_date)].append([result_upc, '', tmp_sc, '1', tmp_srp, '1', tmp_vc, tmp_ic, tmp_date, result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin])
            list_item.append(arr1)

wb_n = Workbook()
ws_n = wb_n.active
ws_n.title = 'Sheet1'
ws_n.append(header)
for idx1 in range(len(list_item)):
    ws_n.append(list_item[idx1])

if not isdir('./result_step_4/'):
    mkdir('./result_step_4/')
wb_n.save('./result_step_4/'+output_file_nm+'.xlsx')

if not isdir('./srp_result/'):
    mkdir('./srp_result/')

for idx_s in dict_srp_store_date:
    tmp_sc, tmp_date = idx_s.split('_')
    tmp_date = tmp_date.split('/')
    for idx_d in range(2):
        if len(tmp_date[idx_d]) == 1:
            tmp_date[idx_d] = '0'+tmp_date[idx_d]
    tmp_date[2] = tmp_date[2][-2:]
    file_nm = tmp_date[0]+tmp_date[1]+tmp_date[2]+'_'+dict_store[tmp_sc]['code']+'_PB.xlsx'

    wb_n = Workbook()
    ws_n = wb_n.active
    ws_n.title = 'Sheet1'
    ws_n.append(['F01','F902','F1000','F126','F30','F1001','VENDOR ID','VENDOR CODE','DATE','Old Case Cost','New Case Cost','Active Price','Old Margin','New Margin'])
    for item in dict_srp_store_date[idx_s]:
        ws_n.append(item)

    wb_n.save('./srp_result/' + file_nm)
