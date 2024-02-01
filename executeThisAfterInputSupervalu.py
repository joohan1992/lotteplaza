import shutil
from openpyxl import load_workbook
import dataEntryFunction as df
import os
import requests

##################### 이 파일만 실행하면 됩니다. ####################
df.set_filename()  ##   <<<   날짜수정은 여기서

filename, objdate = df.set_filename()
today_mmdd = filename[2:6]
today_yy=filename[0:2]
filename_list = []
for i in objdate:
    if df.isMonday():
        filename_list.append(filename + "_TOT_" + i)
    else:
        filename_list.append(filename+"_"+i)
print(filename_list)

download_path = "C:/Users/user/Downloads/"
srp_path = "C:/Users/user/Documents/GitHub/lotteplaza/srp_result/"
work_path = "C:/Users/user/Documents/GitHub/lotteplaza/작업파일/"

url1 = 'http://10.28.78.30:8081/file_input_upload_step_4'  # 1번서버
url2 = 'http://10.28.78.30:8082/file_input_upload_step_4'  # 2번서버

if df.isMonday():
#   월요일의 경우 날짜별 4차 파일을
    filepath_4th_1 = os.path.join(f"C:/Users/user/Documents/GitHub/lotteplaza/작업파일/{df.get_previous_day(filename)[2:6]}/4차",
                                f"TOT_{df.get_previous_day(filename)}_01.xlsx")
    filepath_4th_2 = os.path.join(f"C:/Users/user/Documents/GitHub/lotteplaza/작업파일/{today_mmdd}/4차",
                                  f"TOT_{filename}_01.xlsx")
    print("1번 서버로 각각 돌리고 (“221218_01”, “221219_01”)")
    data1 = {'file_result_name': f"{df.get_previous_day(filename)}_{objdate[0]}"}
    data2 = {'file_result_name': f"{filename}_{objdate[0]}"}

    files1 = {'file_input': (f"TOT_{df.get_previous_day(filename)}_{objdate[0]}.xlsx", open(filepath_4th_1, 'rb'))}  # 1번서버
    files2 = {'file_input': (f"TOT_{filename}_{objdate[0]}.xlsx", open(filepath_4th_2, 'rb'))}  # 2번서버

    response1 = requests.post(url1, files=files1, data=data1)
    response2 = requests.post(url1, files=files2, data=data2)
#  1번 서버 응답
    print("Server 1 response : ", response1)
    print("Server 2 response : ", response2)
    print("작업파일 > “1219” 폴더로 모아서")
    down_url1 = f"http://10.28.78.30:8081/down_file/result_step_4/{df.get_previous_day(filename)}_{objdate[0]}.xlsx"
    down_url2 = f"http://10.28.78.30:8081/down_file/result_step_4/{filename}_{objdate[0]}.xlsx"
    down_response1 = requests.get(down_url1)
    down_response2 = requests.get(down_url2)
    print(r"C:\Users\user\Documents\GitHub\lotteplaza\작업파일\1013 로(경로는 예시)")
    save_path1 = os.path.join(f"C:/Users/user/Documents/GitHub/lotteplaza/작업파일/{today_mmdd}", f"{df.get_previous_day(filename)}_{objdate[0]}.xlsx")
    save_path2 = os.path.join(f"C:/Users/user/Documents/GitHub/lotteplaza/작업파일/{today_mmdd}", f"{filename}_{objdate[0]}.xlsx")
    with open(save_path1, 'wb') as f:
        f.write(down_response1.content)
    with open(save_path2, 'wb') as f:
        f.write(down_response2.content)
    print("모아서 다시 1번, 2번 서버 돌림")
# 모아서 다시 1번, 2번 서버 돌림(“221219_TOT_01", "221219_TOT_02")
    files1 = [
        ('file_input', (f"{df.get_previous_day(filename)}_{objdate[0]}.xlsx", open(save_path1, 'rb'))),
        ('file_input', (f"{filename}_{objdate[0]}.xlsx", open(save_path2, 'rb'))),
    ]
    files2 = [
        ('file_input', (f"{df.get_previous_day(filename)}_{objdate[0]}.xlsx", open(save_path1, 'rb'))),
        ('file_input', (f"{filename}_{objdate[0]}.xlsx", open(save_path2, 'rb'))),
    ]

else:
    filepath_4th = os.path.join(f"C:/Users/user/Documents/GitHub/lotteplaza/작업파일/{today_mmdd}/4차",f"TOT_{filename}_01.xlsx")
    files1 = {'file_input': (f"TOT_{filename}_{objdate[0]}.xlsx", open(filepath_4th, 'rb'))}  # 1번서버
    files2 = {'file_input': (f"TOT_{filename}_{objdate[0]}.xlsx", open(filepath_4th, 'rb'))}  # 1번서버

data1 = {'file_result_name': filename_list[0]}
data2 = {'file_result_name': filename_list[1]}
response1 = requests.post(url1, files=files1, data=data1)
response2 = requests.post(url2, files=files2, data=data2)

print("Server 1 response : ", response1)
print("Server 2 response : ", response2)

down_url1 = f"http://10.28.78.30:8081/down_file/result_step_4/{filename_list[0]}.xlsx"
down_url2 = f"http://10.28.78.30:8082/down_file/result_step_4/{filename_list[1]}.xlsx"
down_response1 = requests.get(down_url1)
down_response2 = requests.get(down_url2)

save_path1 = os.path.join(download_path, f"{filename_list[0]}.xlsx")
save_path2 = os.path.join(download_path, f"{filename_list[1]}.xlsx")
with open(save_path1, 'wb') as f:
    f.write(down_response1.content)
with open(save_path2, 'wb') as f:
    f.write(down_response2.content)
print(save_path1, "4차 처리 파일 저장 완료")
print(save_path2, "4차 처리 파일 저장 완료")


for fname in filename_list:
    shutil.copy(download_path+fname+".xlsx",work_path+today_mmdd+"/")  ## "C:/Users/user/Documents/GitHub/lotteplaza/작업파일/0102/"
    shutil.move(download_path+fname+".xlsx",work_path+today_mmdd+"/결과/")
    zip_path=srp_path+fname
    result_path=work_path+today_mmdd+"/결과/"+fname
    shutil.make_archive(result_path,'zip',zip_path)


editted_fname = f"통합작업본_{today_mmdd}20{today_yy}_작업완료원본.xlsx"


for fname in filename_list:
    if "_"+objdate[1] in fname:
        editted_fname = f"통합작업본_{today_mmdd}20{today_yy}_작업완료원본.xlsx"
        shutil.move(work_path+today_mmdd+"/결과/"+fname+".xlsx", work_path+today_mmdd+"/결과/MARGIN_DIFF/"+editted_fname)
        shutil.move(work_path+today_mmdd+"/결과/"+fname+".zip", work_path+today_mmdd+"/결과/MARGIN_DIFF/")

print("작업중입니다 : ")
print(f"\t - [통합작업본_{today_mmdd}20{today_yy}_작업완료원본.xlsx]의 아래 컬럼들 삭제 후 저장중.. ")
print("\t   [PK] (A), [작업내용] (E), [처리내용] (F), [DupFlag] (AE), [C or E] ~ [department] (AL~AN)")


# 워크북 로드 및 워크시트 선택
wb = load_workbook(work_path+today_mmdd+"/결과/MARGIN_DIFF/"+editted_fname)
ws = wb.active
# 삭제할 열 목록
columns_to_delete = [1, 5, 6, 31, 38, 39, 40]
# 열 삭제
df.delete_columns(ws, columns_to_delete)
# 변경 사항 저장
wb.save(work_path+today_mmdd+"/결과/MARGIN_DIFF/"+editted_fname)

postprocess_path = "C:/Users/user/Documents/GitHub/lotteplaza/postprocess/"
shutil.copy(work_path+today_mmdd+"/결과/MARGIN_DIFF/"+editted_fname,postprocess_path)
df.makedir(postprocess_path+"pb_"+today_mmdd+today_yy)

for fname in filename_list:
    if "_"+objdate[1] in fname:
        shutil.unpack_archive(work_path+today_mmdd+"/결과/MARGIN_DIFF/"+fname+".zip",postprocess_path+"pb_"+today_mmdd+today_yy,'zip')
row_a_length = ws.max_column

print("작업 결과 : ", row_a_length,"개 컬럼 확인")
print()

print("==================================================")
print("==================================================")
print("postprocessing_backup.py 실행...")
exec(open("postprocessing_backup.py", encoding="utf-8").read())