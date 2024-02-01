from openpyxl import load_workbook
import shutil
import dataEntryFunction as df

filename, objdate = df.set_filename()
today_mmdd = filename[2:6]
today_yy=filename[0:2]
filename_list = []

for i in objdate:
    if df.isMonday():
        filename_list.append(filename + "_TOT_" + i)
    else:
        filename_list.append(filename+"_"+i)

md_process_fname=f"C:/Users/user/Documents/GitHub/lotteplaza/postprocess/result/{today_mmdd}{today_yy}_MD"
work_path = "C:/Users/user/Documents/GitHub/lotteplaza/작업파일/"

editted_fname1 = f"통합작업본_{today_mmdd}20{today_yy}_작업완료원본.xlsx"
editted_fname2 = f"통합작업본_{today_mmdd}20{today_yy}.xlsx"
pb_fname=f"PB batch files_{today_mmdd}20{today_yy}.zip"
cpb_fname=f"CPB_처리_결과파일_{today_mmdd}20{today_yy}"


for fname in filename_list:
    if "_"+objdate[0] in fname:
        shutil.move(work_path+today_mmdd+"/결과/"+fname+".xlsx",work_path+today_mmdd+"/결과/"+editted_fname1)
        shutil.copy(work_path+today_mmdd+"/결과/"+editted_fname1, work_path + today_mmdd + "/결과/" + editted_fname2)
        shutil.move(work_path+today_mmdd+"/결과/"+fname+".zip",work_path+today_mmdd+"/결과/"+pb_fname)
shutil.make_archive(work_path+today_mmdd+"/결과/"+cpb_fname,'zip',md_process_fname)

# 워크북 로드 및 워크시트 선택
print(work_path+today_mmdd+"/"+"결과/"+editted_fname1)
wb = load_workbook(work_path+today_mmdd+"/"+"결과/"+editted_fname1)
ws = wb.active
# 삭제할 열 목록
columns_to_delete = [1, 5, 6, 31, 38, 39, 40]
# 열 삭제
df.delete_columns(ws, columns_to_delete)
row_a_length = ws.max_column
print("열 개수:", row_a_length)
# 변경 사항 저장
wb.save(work_path+today_mmdd+"/"+"결과/"+editted_fname1)

print("=======================================================")

print(work_path+today_mmdd+"/"+"결과/"+editted_fname2)
# 워크북 로드 및 워크시트 선택
wb = load_workbook(work_path + today_mmdd + "/" + "결과/" + editted_fname2)
ws = wb.active
# 열 삭제
columns_to_delete = [1, 5, 6, 8, 9, 10, 11, 12, 18, 19, 20, 21, 31, 38, 39, 40]
df.delete_columns(ws, columns_to_delete)
row_a_length = ws.max_column
print("열 개수:", row_a_length)
# 변경 사항 저장
wb.save(work_path + today_mmdd + "/" + "결과/" + editted_fname2)

print("==================================================")
print("==================================================")
print("finalProcessing.py까지 실행되었습니다.")
print("이후 작업량 계산 프로세스 진행해 주세요.")



