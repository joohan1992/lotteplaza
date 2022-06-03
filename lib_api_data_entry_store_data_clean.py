import re
from openpyxl import load_workbook, Workbook
from os import listdir, mkdir, remove
from os.path import isfile, join, isdir
import traceback
import threading
import time
import pickle


flag_ignore_srp_sticker = True
flag_enable_cp_price = True
flag_enable_depart_vat = True
flag_margin_diff = True


def getFormalUPC(upc, vc):
    if upc is None or upc == '':
        ret_upc = ''
    elif len(upc) == 12:
        ret_upc = '00' + upc[:-1]
    elif len(upc) == 13:
        ret_upc = '0' + upc[:-1]
    elif len(upc) == 11:
        ret_upc = '00' + upc
    else:
        ret_upc = upc

    cand_upc = []
    if upc is not None and len(upc) > 1 and vc is not None:
        if vc == '0179':
            if len(upc) == 14:
                cand_upc.append('0' + upc[1:-1])
        elif vc == '0501':
            if len(upc) == 11:
                cand_upc.append('000' + upc[:-1])
        elif vc == '1004':
            cand_upc.append('00' + upc[:-1])
            cand_upc.append('0' + upc[:-1])
            cand_upc.append('00' + upc)
        elif vc == '1213':
            if len(upc) == 10:
                cand_upc.append('000' + upc)
                cand_upc.append('007' + upc)
        elif vc == '1229':
            if len(upc) == 13:
                cand_upc.append(upc)
            elif len(upc) == 11:
                cand_upc.append('00' + upc)
        elif vc == '1231':
            tmp_len = len(upc)
            suff = ''
            for tmp_idx in range(13 - tmp_len):
                suff += '0'
            cand_upc.append(suff + upc)
        elif vc == '1251':
            cand_upc.append('0' + upc)
        elif vc == '1210':
            cand_upc.append('00028400'+upc)
        elif vc == '1017':
            cand_upc.append('000'+upc)
        '''
        elif vc == '1223':
            cand_upc.append(upc)
        '''

    return ret_upc, cand_upc


def getFormalItemCode(ic, vc, dict_vend):

    tmp_ic = ic
    cand_ic = []
    if vc == '0203' and len(str(tmp_ic)) > 0 and (ord(tmp_ic[-1]) in range(65, 91) or ord(tmp_ic[-1]) in range(97, 123)):
        tmp_ic = tmp_ic[:-1]
    elif vc == '1229':
        while len(str(tmp_ic)) > 0 and (ord(tmp_ic[0]) in range(65, 91) or ord(tmp_ic[0]) in range(97, 123)):
            tmp_ic = tmp_ic[1:]
        while len(str(tmp_ic)) > 0 and (ord(tmp_ic[-1]) in range(65, 91) or ord(tmp_ic[-1]) in range(97, 123)):
            tmp_ic = tmp_ic[:-1]

    if tmp_ic is not None and len(str(tmp_ic)) > 0 and vc is not None:
        if vc == '1205':
            cand_ic.append('CLO'+tmp_ic[-4:])
            cand_ic.append('0000'+tmp_ic[-4:])
        elif vc in dict_vend and dict_vend[vc]['prefix'] is not None and dict_vend[vc]['prefix'] != 'None' \
                and dict_vend[vc]['prefix'] != '':
            if tmp_ic.find(dict_vend[vc]['prefix']) == 0:
                cand_ic.append(tmp_ic)
                cand_ic.append(tmp_ic.split(dict_vend[vc]['prefix'], 1)[1])
            else:
                cand_ic.append(dict_vend[vc]['prefix']+tmp_ic)
                cand_ic.append(tmp_ic)
        else:
            cand_ic.append(tmp_ic)

    return cand_ic


def getFormalAMT(csize, ccost):
    if csize is not None and csize != '' and re.match(r"(?<![-.])\b[0-9]+\b(?!\.[0-9])", str(csize)) \
            and ccost is not None and ccost != '' and re.match(r"(?<![-.])\b[0-9]+\b(?!\.[0-9])", str(ccost)):
        ret_amt = str(float(csize)*float(ccost))
    else:
        ret_amt = '0'
    return ret_amt


def getVAT(vc, sc, dp):
    vat_depart = {
        'CIGARETTES': 0,
        'DAIRY': 12,
        'DELI': 0,
        'EGGS': 0.2,
        'ETHNIC': 2.5,
        'FROZ BAKERY': 8,
        'FROZ FOOD': 21,
        'SUPPLIES': 6,
        'GEN MDSE': 13.5,
        'GROCERY': 11,
        'HBA': 14.5,
        'ICE': 23,
        'MEAT': 0,
        'MILK': 0,
        'PRODUCE': 0
    }
    if vc is not None and vc == '1229':
        if flag_enable_depart_vat and sc == '011':
            return 1.5
        elif flag_enable_depart_vat and dp is not None and dp in vat_depart:
            return vat_depart[dp]
        return 11.5
    elif sc is not None and sc == '011':
        return 5
    else:
        return 0


def check_except_pb(item):
    if item[3] == '1205' and (item[0] in ['0004938710011', '0004938710012', '0004938710013', '0004938710015'] or item[4] in ['CLO1051', 'CLO1307', 'CLO1425', 'CLO1593']):
        return False
    else:
        return True


def load_store_db(fname_store_data):
    f = open('./bin_store_data/'+fname_store_data+'.pkl', 'rb')
    dict_vendor = pickle.load(f)
    f.close()

    print('DB file is loaded')

    return dict_vendor


def load_base_db(g_data):
    file_result = './base_file/'+g_data['fname_base_data']
    wb_r = load_workbook(file_result)

    ws_r = wb_r['점포번호']
    dict_store = dict()
    idx1 = 0
    for row in ws_r.rows:
        # 첫 행(컬럼 명) 통과
        if idx1 == 0:
            idx1 += 1
            continue
        store_no = row[0].value
        store_nm = row[1].value
        store_cd = row[2].value
        store_addr = row[3].value
        dict_store[store_no] = dict()
        dict_store[store_no]['name'] = store_nm
        dict_store[store_no]['code'] = store_cd
        dict_store[store_no]['addr'] = store_addr

    ws_r = wb_r['업체별번호']
    dict_vendor = dict()
    idx1 = 0
    for row in ws_r.rows:
        # 첫 행(컬럼 명) 통과
        if idx1 == 0:
            idx1 += 1
            continue
        vendor_no = str(row[0].value)
        vendor_nm = str(row[1].value)
        vendor_pre = str(row[3].value)
        dict_vendor[vendor_no] = dict()
        dict_vendor[vendor_no]['name'] = vendor_nm
        dict_vendor[vendor_no]['prefix'] = vendor_pre

    wb_r.close()

    # wb_r = load_workbook()
    # dict_item_code = dict()

    print('Base DB file is loaded')

    return dict_store, dict_vendor
    # return dict_store, dict_vendor, dict_item_code


def data_lookup(g_data, tmp_upc, tmp_sc, tmp_vc, tmp_ic, tmp_csize, tmp_ncost, tmp_amt, tmp_upc_cand2, tmp_vat,
                tmp_c_or_e, tmp_flag_processed_ce, dict_error_flag, dict_vendor, dict_vend, param_etc=None):
    result_upc = tmp_upc
    result_ic = tmp_ic
    result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
    if len(result_cand_ic) == 2:
        result_ic = result_cand_ic[0]
    tmp_ic_cand = getFormalItemCode(tmp_ic, tmp_vc, dict_vend)
    if len(tmp_ic_cand) == 2:
        tmp_ic = tmp_ic_cand[1]

    result_desc = ''
    if tmp_csize is not None and tmp_csize != '':
        tmp_csize = float(tmp_csize)
    result_csize = tmp_csize

    tmp_ncost = float(tmp_ncost)
    tmp_amt = float(tmp_amt)
    result_ncost = tmp_ncost
    result_amt = tmp_amt
    result_flag_processed_ce = tmp_flag_processed_ce
    result_upc_check = '0'
    result_upc_same_check = '0'
    result_desc_lookup_base = 'N'
    result_old_case_cost = ''
    result_new_case_cost = ''
    result_price = ''
    result_old_margin = ''
    result_new_margin = ''
    result_srp = ''
    detail_cat = ''
    result_note = ''

    if tmp_csize is None or tmp_csize == '' or float(tmp_csize) <= 0:
        dict_error_flag['lessthanone'] = 0

    # L0 업체 코드 없는 경우
    if tmp_vc is None:
        dict_error_flag['disable'] = 0

    # L0 업체 코드 있고 매장 데이터에 등록되어있는 경우
    elif tmp_vc in dict_vendor:

        # L1 아이템 코드가 인보이스에 없는 경우
        if tmp_ic is None:

            # L2 UPC도 없는 경우
            if tmp_upc is None:
                dict_error_flag['disable'] = 0

            # L2 UPC 코드는 입력된 경우
            else:

                if tmp_sc is None:
                    tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                    # UPC 기준으로 탐색
                    flag_exist_upc = False
                    flag_exist_upc_cand = False
                    flag_exist_exact = False
                    tmp_ic2 = ''
                    last_sc = ''
                    idx_upc_cand = -1
                    for key1 in dict_vendor[tmp_vc]:
                        for key2 in dict_vendor[tmp_vc][key1]:
                            if key2 == 'upc_dict':
                                continue
                            for key3 in dict_vendor[tmp_vc][key1][key2]:
                                if key3 == tmp_upc2:
                                    flag_exist_upc = True
                                    flag_exist_upc_cand = False
                                    tmp_ic2 = key1
                                    last_sc = key2
                                    break
                                elif key3 in tmp_upc_cand:
                                    flag_exist_upc_cand = True
                                    flag_exist_upc = False
                                    idx_upc_cand = tmp_upc_cand.index(key3)
                                    tmp_ic2 = key1
                                    last_sc = key2
                                    break
                            if flag_exist_upc or flag_exist_upc_cand:
                                break
                        if flag_exist_upc or flag_exist_upc_cand:
                            break

                    # UPC가 없는 경우
                    # if not flag_exist_upc and not flag_exist_upc_cand:
                    if True:    # store_cd 없으면 무조건 신상품
                        dict_error_flag['new'] = 1
                        result_ic = tmp_ic
                        result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                        if len(result_cand_ic) == 2:
                            result_ic = result_cand_ic[0]
                    else:
                        result_desc_lookup_base = 'U'
                        if flag_exist_upc:
                            result_upc = tmp_upc2
                        else:
                            result_upc = tmp_upc_cand[idx_upc_cand]
                        result_desc = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['desc']
                        result_csize = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['csize']
                        vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                        if float(vendor_csize) != float(result_csize):
                            detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                        result_csize = vendor_csize
                        if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                            result_csize = 1
                            detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                        result_old_case_cost = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['ccost']
                        result_new_case_cost = tmp_ncost
                        if tmp_flag_processed_ce == 1:
                            result_ncost = tmp_ncost * result_csize
                            result_new_case_cost = tmp_ncost * result_csize
                            result_amt = tmp_amt * result_csize
                            result_flag_processed_ce = 0
                        elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                            result_ncost = (tmp_ncost / tmp_csize) * result_csize
                            result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                            result_amt = (tmp_amt / tmp_csize) * result_csize
                        result_price = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['price']
                        result_ic = tmp_ic2
                        result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                        if len(result_cand_ic) == 2:
                            result_ic = result_cand_ic[0]

                        result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                        result_note = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['note']
                        if flag_ignore_srp_sticker and result_note == 'STICKER':
                            print(tmp_ic + "\t" + tmp_upc)
                            result_srp = ' '

                else:
                    tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                    # UPC 기준으로 탐색
                    flag_exist_upc = False
                    flag_exist_upc_cand = False
                    flag_exist_exact = False
                    tmp_ic2 = ''
                    last_sc = tmp_sc
                    idx_upc_cand = -1
                    for key1 in dict_vendor[tmp_vc]:
                        if tmp_sc in dict_vendor[tmp_vc][key1]:
                            for key3 in dict_vendor[tmp_vc][key1][tmp_sc]:
                                if key3 == tmp_upc2:
                                    flag_exist_upc = True
                                    flag_exist_upc_cand = False
                                    tmp_ic2 = key1
                                    flag_exist_exact = True
                                    break
                                elif key3 in tmp_upc_cand:
                                    flag_exist_upc_cand = True
                                    flag_exist_upc = False
                                    idx_upc_cand = tmp_upc_cand.index(key3)
                                    tmp_ic2 = key1
                                    flag_exist_exact = True
                                    break
                        '''
                        for key2 in dict_vendor[tmp_vc][key1]:
                            if key2 == 'upc_dict':
                                continue
                            for key3 in dict_vendor[tmp_vc][key1][key2]:
                                if key3 == tmp_upc2:
                                    flag_exist_upc = True
                                    flag_exist_upc_cand = False
                                    tmp_ic2 = key1
                                    last_sc = key2
                                    if key2 == tmp_sc:
                                        flag_exist_exact = True
                                        break
                                elif key3 in tmp_upc_cand:
                                    flag_exist_upc_cand = True
                                    flag_exist_upc = False
                                    idx_upc_cand = tmp_upc_cand.index(key3)
                                    tmp_ic2 = key1
                                    last_sc = key2
                                    if key2 == tmp_sc:
                                        flag_exist_exact = True
                                        break
                            if flag_exist_exact:
                                break
                        if flag_exist_exact:
                            break
                        '''

                    # UPC가 없는 경우
                    if not flag_exist_upc and not flag_exist_upc_cand:
                        dict_error_flag['new'] = 1
                        result_ic = tmp_ic
                        result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                        if len(result_cand_ic) == 2:
                            result_ic = result_cand_ic[0]
                    else:
                        result_desc_lookup_base = 'U'
                        if flag_exist_upc:
                            result_upc = tmp_upc2
                        else:
                            result_upc = tmp_upc_cand[idx_upc_cand]
                        result_desc = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['desc']
                        result_csize = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['csize']
                        vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                        if float(vendor_csize) != float(result_csize):
                            detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                        result_csize = vendor_csize
                        if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                            result_csize = 1
                            detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                        result_old_case_cost = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['ccost']
                        result_new_case_cost = tmp_ncost
                        if tmp_flag_processed_ce == 1:
                            result_ncost = tmp_ncost * result_csize
                            result_new_case_cost = tmp_ncost * result_csize
                            result_amt = tmp_amt * result_csize
                            result_flag_processed_ce = 0
                        elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                            result_ncost = (tmp_ncost / tmp_csize) * result_csize
                            result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                            result_amt = (tmp_amt / tmp_csize) * result_csize
                        result_price = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['price']
                        result_ic = tmp_ic2
                        result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                        if len(result_cand_ic) == 2:
                            result_ic = result_cand_ic[0]

                        result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                        result_note = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['note']
                        if flag_ignore_srp_sticker and result_note == 'STICKER':
                            print(tmp_ic + "\t" + tmp_upc)
                            result_srp = ' '

        # L1 prefix를 사용하면서 prefix를 붙인 ic가 매장데이터에 있는 경우
        elif len(tmp_ic_cand) > 1 and tmp_ic_cand[0] in dict_vendor[tmp_vc]:
            org_tmp_ic = tmp_ic
            tmp_ic = tmp_ic_cand[0]

            # L2 Store Code가 없는 경우
            if tmp_sc is None or tmp_sc not in dict_vendor[tmp_vc][tmp_ic]:
                # prefix 적용 안된 ic로 해당 점포 기록이 있는 경우
                if tmp_sc is not None and org_tmp_ic in dict_vendor[tmp_vc] and tmp_sc in dict_vendor[tmp_vc][org_tmp_ic]:
                    tmp_ic = org_tmp_ic
                    fname = param_etc['file_nm']
                    tag_is = param_etc['tag_is']
                    f_item_search = open('item_search/'+str(tag_is)+'.txt', 'a')
                    f_item_search.write(fname+'\t'+tmp_vc+'\t'+str(tmp_ic_cand)+'\n')
                    for tmp_cand_item in tmp_ic_cand:
                        if tmp_cand_item in dict_vendor[tmp_vc]:
                            f_item_search.write(tmp_cand_item+'\t'+str(dict_vendor[tmp_vc][tmp_cand_item])+'\n')
                    f_item_search.write('\n\n')
                    f_item_search.close()

                    # Store Code에 중복없이 하나의 UPC만 있는 경우
                    if len(dict_vendor[tmp_vc][tmp_ic][tmp_sc]) == 1:
                        result_upc = next(iter(dict_vendor[tmp_vc][tmp_ic][tmp_sc]))
                        result_desc = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['desc']
                        result_csize = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['csize']
                        vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                        if float(vendor_csize) != float(result_csize):
                            detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                        result_csize = vendor_csize
                        if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                            result_csize = 1
                            detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                        result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['ccost']
                        result_new_case_cost = tmp_ncost
                        if tmp_flag_processed_ce == 1:
                            result_ncost = tmp_ncost * result_csize
                            result_new_case_cost = tmp_ncost * result_csize
                            result_amt = tmp_amt * result_csize
                            result_flag_processed_ce = 0
                        elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                            result_ncost = (tmp_ncost / tmp_csize) * result_csize
                            result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                            result_amt = (tmp_amt / tmp_csize) * result_csize
                        result_price = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['price']

                        result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                        result_note = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['note']
                        if flag_ignore_srp_sticker and result_note == 'STICKER':
                            print(tmp_ic + "\t" + tmp_upc)
                            result_srp = ' '

                        result_upc_check = '1'
                        result_desc_lookup_base = 'I'

                        tmp_upc_3, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)
                        if tmp_upc_3 == result_upc or result_upc in tmp_upc_cand:
                            result_upc_same_check = '1'

                        result_ic = tmp_ic
                        result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                        if len(result_cand_ic) == 2:
                            result_ic = result_cand_ic[0]

                    elif len(dict_vendor[tmp_vc][tmp_ic][tmp_sc]) > 1:
                        result_upc_check = str(len(dict_vendor[tmp_vc][tmp_ic][tmp_sc]))

                        # UPC가 인보이스상에 있는 경우
                        if tmp_upc is not None:
                            tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                            # UPC 기준으로 탐색
                            flag_exist_upc = False
                            flag_exist_upc_cand = False
                            flag_exist_exact = False
                            tmp_ic2 = ''
                            last_sc = ''
                            idx_upc_cand = -1
                            for key1 in dict_vendor[tmp_vc]:
                                if tmp_sc in dict_vendor[tmp_vc][key1]:
                                    last_sc = tmp_sc
                                    for key3 in dict_vendor[tmp_vc][key1][tmp_sc]:
                                        if key3 == tmp_upc2:
                                            flag_exist_upc = True
                                            flag_exist_upc_cand = False
                                            tmp_ic2 = key1
                                            flag_exist_exact = True
                                            break
                                        if key3 in tmp_upc_cand:
                                            flag_exist_upc_cand = True
                                            flag_exist_upc = False
                                            idx_upc_cand = tmp_upc_cand.index(key3)
                                            tmp_ic2 = key1
                                            flag_exist_exact = True
                                            break
                                '''
                                for key2 in dict_vendor[tmp_vc][key1]:
                                    if key2 == 'upc_dict':
                                        continue
                                    for key3 in dict_vendor[tmp_vc][key1][key2]:
                                        if key3 == tmp_upc2:
                                            flag_exist_upc = True
                                            flag_exist_upc_cand = False
                                            tmp_ic2 = key1
                                            last_sc = key2
                                            if key2 == tmp_sc:
                                                flag_exist_exact = True
                                                break
                                        if key3 in tmp_upc_cand:
                                            flag_exist_upc_cand = True
                                            flag_exist_upc = False
                                            idx_upc_cand = tmp_upc_cand.index(key3)
                                            tmp_ic2 = key1
                                            last_sc = key2
                                            if key2 == tmp_sc:
                                                flag_exist_exact = True
                                                break
                                    if flag_exist_exact:
                                        break
                                if flag_exist_exact:
                                    break
                                '''

                            # UPC가 Store 내에 없는 경우
                            if not flag_exist_upc and not flag_exist_upc_cand:
                                dict_error_flag['new'] = 1
                                '''
                                tmp_upc_cand_list = []
                                for cand in dict_vendor[tmp_vc][tmp_ic][tmp_sc]:
                                    tmp_upc_cand_list.append(cand)
                                tmp_upc_cand2.extend(tmp_upc_cand_list)
                                if len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']) > 0:

                                    # UPC is sole
                                    if len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']) == 1:
                                        result_upc_check = '1'
                                        itr = iter(dict_vendor[tmp_vc][tmp_ic])
                                        next(itr)
                                        tmp_sc2 = next(itr)
                                        result_upc = next(iter(dict_vendor[tmp_vc][tmp_ic][tmp_sc2]))
                                        result_desc_lookup_base = 'I'
                                        result_desc = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['desc']
                                        result_csize = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['csize']
                                        vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                                        if float(vendor_csize) != float(result_csize):
                                            detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                                        result_csize = vendor_csize
                                        if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                                            result_csize = 1
                                            detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                                        result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['ccost']
                                        result_new_case_cost = tmp_ncost
                                        if tmp_flag_processed_ce == 1:
                                            result_ncost = tmp_ncost * result_csize
                                            result_new_case_cost = tmp_ncost * result_csize
                                            result_amt = tmp_amt * result_csize
                                            result_flag_processed_ce = 0
                                        elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                                            result_ncost = (tmp_ncost / tmp_csize) * result_csize
                                            result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                                            result_amt = (tmp_amt / tmp_csize) * result_csize
                                        result_price = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['price']

                                        result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                                        result_note = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['note']
                                        if flag_ignore_srp_sticker and result_note == 'STICKER':
                                            print(tmp_ic + "\t" + tmp_upc)
                                            result_srp = ' '

                                        tmp_upc3, tmp_upc_cand = getFormalUPC(result_upc, tmp_vc)
                                        if tmp_upc == result_upc or result_upc in tmp_upc_cand:
                                            result_upc_same_check = 1
                                        else:
                                            result_upc_same_check = 0

                                    # UPC is duplicated
                                    elif len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']) > 1:

                                        if tmp_upc is None:
                                            dict_error_flag['duplicate'] = 0
                                        else:
                                            tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                                            # UPC 기준으로 탐색
                                            flag_exist_upc = False
                                            flag_exist_upc_cand = False
                                            last_sc = ''
                                            idx_upc_cand = -1
                                            for key1 in dict_vendor[tmp_vc][tmp_ic]:
                                                if key1 == 'upc_dict':
                                                    continue
                                                for key2 in dict_vendor[tmp_vc][tmp_ic][key1]:
                                                    if key2 == tmp_upc2:
                                                        flag_exist_upc = True
                                                        flag_exist_upc_cand = False
                                                        last_sc = key1
                                                        break
                                                    if key2 in tmp_upc_cand:
                                                        flag_exist_upc_cand = True
                                                        flag_exist_upc = False
                                                        idx_upc_cand = tmp_upc_cand.index(key2)
                                                        last_sc = key1
                                                        break
                                                if flag_exist_upc or flag_exist_upc_cand:
                                                    break

                                            # UPC가 없는 경우
                                            if not flag_exist_upc and not flag_exist_upc_cand:
                                                tmp_upc_cand_list = []
                                                for cand in dict_vendor[tmp_vc][tmp_ic]['upc_dict']:
                                                    tmp_upc_cand_list.append(cand)
                                                tmp_upc_cand2.extend(tmp_upc_cand_list)
                                                dict_error_flag['new'] = 1
                                            else:
                                                result_desc_lookup_base = 'U'
                                                if flag_exist_upc:
                                                    result_upc = tmp_upc2
                                                else:
                                                    result_upc = tmp_upc_cand[idx_upc_cand]
                                                result_desc = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['desc']
                                                result_csize = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['csize']
                                                vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                                                if float(vendor_csize) != float(result_csize):
                                                    detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                                                result_csize = vendor_csize
                                                if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                                                    result_csize = 1
                                                    detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                                                result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc][
                                                    'ccost']
                                                result_new_case_cost = tmp_ncost
                                                if tmp_flag_processed_ce == 1:
                                                    result_ncost = tmp_ncost * result_csize
                                                    result_new_case_cost = tmp_ncost * result_csize
                                                    result_amt = tmp_amt * result_csize
                                                    result_flag_processed_ce = 0
                                                elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                                                    result_ncost = (tmp_ncost / tmp_csize) * result_csize
                                                    result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                                                    result_amt = (tmp_amt / tmp_csize) * result_csize
                                                result_price = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['price']

                                                result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                                                result_note = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['note']
                                                if flag_ignore_srp_sticker and result_note == 'STICKER':
                                                    print(tmp_ic + "\t" + tmp_upc)
                                                    result_srp = ' '

                                    else:
                                        dict_error_flag['new'] = 1

                                result_ic = tmp_ic
                                result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                                if len(result_cand_ic) == 2:
                                    result_ic = result_cand_ic[0]
                                '''
                            else:
                                result_desc_lookup_base = 'U'
                                if flag_exist_upc:
                                    result_upc = tmp_upc2
                                else:
                                    result_upc = tmp_upc_cand[idx_upc_cand]
                                result_desc = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['desc']
                                result_csize = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['csize']
                                vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                                if float(vendor_csize) != float(result_csize):
                                    detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                                result_csize = vendor_csize
                                if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                                    result_csize = 1
                                    detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                                result_old_case_cost = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['ccost']
                                result_new_case_cost = tmp_ncost
                                if tmp_flag_processed_ce == 1:
                                    result_ncost = tmp_ncost * result_csize
                                    result_new_case_cost = tmp_ncost * result_csize
                                    result_amt = tmp_amt * result_csize
                                    result_flag_processed_ce = 0
                                elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                                    result_ncost = (tmp_ncost / tmp_csize) * result_csize
                                    result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                                    result_amt = (tmp_amt / tmp_csize) * result_csize
                                result_price = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['price']

                                result_ic = tmp_ic2
                                result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                                if len(result_cand_ic) == 2:
                                    result_ic = result_cand_ic[0]

                                result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                                result_note = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['note']
                                if flag_ignore_srp_sticker and result_note == 'STICKER':
                                    print(tmp_ic + "\t" + tmp_upc)
                                    result_srp = ' '

                        else:
                            result_ic = tmp_ic
                            result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                            if len(result_cand_ic) == 2:
                                result_ic = result_cand_ic[0]
                            dict_error_flag['duplicate'] = 1

                    else:
                        result_ic = tmp_ic
                        result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                        if len(result_cand_ic) == 2:
                            result_ic = result_cand_ic[0]

                # UPC is sole
                elif len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']) == 1 and tmp_sc in dict_vendor[tmp_vc][tmp_ic]:
                    result_upc_check = '1'
                    # itr = iter(dict_vendor[tmp_vc][tmp_ic])
                    # next(itr)
                    # tmp_sc2 = next(itr)
                    result_upc = next(iter(dict_vendor[tmp_vc][tmp_ic][tmp_sc]))
                    result_desc_lookup_base = 'I'
                    result_desc = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['desc']
                    result_csize = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['csize']
                    vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                    if float(vendor_csize) != float(result_csize):
                        detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                    result_csize = vendor_csize
                    if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                        result_csize = 1
                        detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                    result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['ccost']
                    result_new_case_cost = tmp_ncost
                    if tmp_flag_processed_ce == 1:
                        result_ncost = tmp_ncost * result_csize
                        result_new_case_cost = tmp_ncost * result_csize
                        result_amt = tmp_amt * result_csize
                        result_flag_processed_ce = 0
                    elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                        result_ncost = (tmp_ncost / tmp_csize) * result_csize
                        result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                        result_amt = (tmp_amt / tmp_csize) * result_csize
                    result_price = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['price']

                    result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                    result_note = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['note']
                    if flag_ignore_srp_sticker and result_note == 'STICKER':
                        print(tmp_ic + "\t" + tmp_upc)
                        result_srp = ' '

                    tmp_upc_3, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)
                    if result_upc == tmp_upc_3 or result_upc in tmp_upc_cand:
                        result_upc_same_check = 1
                    else:
                        result_upc_same_check = 0

                    result_ic = tmp_ic
                    result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                    if len(result_cand_ic) == 2:
                        result_ic = result_cand_ic[0]

                # UPC is duplicated
                elif len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']) > 1 and tmp_sc in dict_vendor[tmp_vc][tmp_ic]:
                    result_upc_check = str(len(dict_vendor[tmp_vc][tmp_ic][tmp_sc].keys()))

                    if tmp_upc is None:
                        dict_error_flag['duplicate'] = 1

                    else:
                        tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                        # UPC 기준으로 탐색
                        flag_exist_upc = False
                        flag_exist_upc_cand = False
                        last_sc = tmp_sc
                        last_sc_cand = tmp_sc
                        idx_upc_cand = -1
                        if tmp_sc in dict_vendor[tmp_vc][tmp_ic]:
                            for key2 in dict_vendor[tmp_vc][tmp_ic][tmp_sc]:
                                if key2 == tmp_upc2:
                                    flag_exist_upc = True
                                    flag_exist_upc_cand = False
                                    break
                                if key2 in tmp_upc_cand:
                                    flag_exist_upc_cand = True
                                    flag_exist_upc = False
                                    idx_upc_cand = tmp_upc_cand.index(key2)
                                    break
                        '''
                        for key1 in dict_vendor[tmp_vc][tmp_ic]:
                            if key1 == 'upc_dict':
                                continue
                            for key2 in dict_vendor[tmp_vc][tmp_ic][key1]:
                                if key2 == tmp_upc2:
                                    flag_exist_upc = True
                                    flag_exist_upc_cand = False
                                    last_sc = key1
                                    break
                                if key2 in tmp_upc_cand:
                                    flag_exist_upc_cand = True
                                    flag_exist_upc = False
                                    idx_upc_cand = tmp_upc_cand.index(key2)
                                    last_sc_cand = key1
                                    break
                            if flag_exist_upc or flag_exist_upc_cand:
                                break
                        '''

                        # UPC가 없는 경우
                        if not flag_exist_upc and not flag_exist_upc_cand:
                            tmp_upc_cand_list = []
                            for cand in dict_vendor[tmp_vc][tmp_ic]['upc_dict']:
                                tmp_upc_cand_list.append(cand)
                            tmp_upc_cand2.extend(tmp_upc_cand_list)
                            dict_error_flag['new'] = 1
                        else:
                            result_desc_lookup_base = 'U'
                            if flag_exist_upc:
                                result_upc = tmp_upc2
                            else:
                                result_upc = tmp_upc_cand[idx_upc_cand]
                                last_sc = last_sc_cand
                            result_desc = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['desc']
                            result_csize = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['csize']
                            vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                            if float(vendor_csize) != float(result_csize):
                                detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                            result_csize = vendor_csize
                            if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                                result_csize = 1
                                detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                            result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['ccost']
                            result_new_case_cost = tmp_ncost
                            if tmp_flag_processed_ce == 1:
                                result_ncost = tmp_ncost * result_csize
                                result_new_case_cost = tmp_ncost * result_csize
                                result_amt = tmp_amt * result_csize
                                result_flag_processed_ce = 0
                            elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                                result_ncost = (tmp_ncost / tmp_csize) * result_csize
                                result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                                result_amt = (tmp_amt / tmp_csize) * result_csize
                            result_price = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['price']

                            result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                            result_note = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['note']
                            if flag_ignore_srp_sticker and result_note == 'STICKER':
                                print(tmp_ic + "\t" + tmp_upc)
                                result_srp = ' '
                    result_ic = tmp_ic
                    result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                    if len(result_cand_ic) == 2:
                        result_ic = result_cand_ic[0]

                else:
                    result_ic = tmp_ic
                    result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                    if len(result_cand_ic) == 2:
                        result_ic = result_cand_ic[0]

            # L2 Store Code가 있는 경우
            elif tmp_sc is not None and tmp_sc in dict_vendor[tmp_vc][tmp_ic]:

                # Store Code에 중복없이 하나의 UPC만 있는 경우
                if len(dict_vendor[tmp_vc][tmp_ic][tmp_sc]) == 1:
                    result_upc = next(iter(dict_vendor[tmp_vc][tmp_ic][tmp_sc]))
                    result_desc = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['desc']
                    result_csize = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['csize']
                    vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                    if float(vendor_csize) != float(result_csize):
                        detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                    result_csize = vendor_csize
                    if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                        result_csize = 1
                        detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                    result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['ccost']
                    result_new_case_cost = tmp_ncost
                    if tmp_flag_processed_ce == 1:
                        result_ncost = tmp_ncost * result_csize
                        result_new_case_cost = tmp_ncost * result_csize
                        result_amt = tmp_amt * result_csize
                        result_flag_processed_ce = 0
                    elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                        result_ncost = (tmp_ncost / tmp_csize) * result_csize
                        result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                        result_amt = (tmp_amt / tmp_csize) * result_csize
                    result_price = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['price']

                    result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                    result_note = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['note']
                    if flag_ignore_srp_sticker and result_note == 'STICKER':
                        print(tmp_ic + "\t" + tmp_upc)
                        result_srp = ' '

                    result_upc_check = '1'
                    result_desc_lookup_base = 'I'

                    tmp_upc_3, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)
                    if tmp_upc_3 == result_upc or result_upc in tmp_upc_cand:
                        result_upc_same_check = '1'

                    result_ic = tmp_ic
                    result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                    if len(result_cand_ic) == 2:
                        result_ic = result_cand_ic[0]

                elif len(dict_vendor[tmp_vc][tmp_ic][tmp_sc]) > 1:
                    result_upc_check = str(len(dict_vendor[tmp_vc][tmp_ic][tmp_sc]))

                    if tmp_upc is not None:
                        tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                        # UPC 기준으로 탐색
                        flag_exist_upc = False
                        flag_exist_upc_cand = False
                        flag_exist_exact = False
                        tmp_ic2 = ''
                        last_sc = tmp_sc
                        idx_upc_cand = -1
                        for key1 in dict_vendor[tmp_vc]:
                            if tmp_sc in dict_vendor[tmp_vc][key1]:
                                for key3 in dict_vendor[tmp_vc][key1][tmp_sc]:
                                    if key3 == tmp_upc2:
                                        flag_exist_upc = True
                                        flag_exist_upc_cand = False
                                        tmp_ic2 = key1
                                        flag_exist_exact = True
                                        break
                                    if key3 in tmp_upc_cand:
                                        flag_exist_upc_cand = True
                                        flag_exist_upc = False
                                        idx_upc_cand = tmp_upc_cand.index(key3)
                                        tmp_ic2 = key1
                                        flag_exist_exact = True
                                        break
                            if flag_exist_exact:
                                break
                            '''
                            for key2 in dict_vendor[tmp_vc][key1]:
                                if key2 == 'upc_dict':
                                    continue
                                for key3 in dict_vendor[tmp_vc][key1][key2]:
                                    if key3 == tmp_upc2:
                                        flag_exist_upc = True
                                        flag_exist_upc_cand = False
                                        tmp_ic2 = key1
                                        last_sc = key2
                                        if key2 == tmp_sc:
                                            flag_exist_exact = True
                                            break
                                    if key3 in tmp_upc_cand:
                                        flag_exist_upc_cand = True
                                        flag_exist_upc = False
                                        idx_upc_cand = tmp_upc_cand.index(key3)
                                        tmp_ic2 = key1
                                        last_sc = key2
                                        if key2 == tmp_sc:
                                            flag_exist_exact = True
                                            break
                                if flag_exist_exact:
                                    break
                            if flag_exist_exact:
                                break
                            '''

                        # UPC가 Store 내에 없는 경우
                        if not flag_exist_upc and not flag_exist_upc_cand:
                            tmp_upc_cand_list = []
                            for cand in dict_vendor[tmp_vc][tmp_ic][tmp_sc]:
                                tmp_upc_cand_list.append(cand)
                            tmp_upc_cand2.extend(tmp_upc_cand_list)
                            if False and len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']) > 0:
                                # UPC is sole
                                if len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']) == 1:
                                    result_upc_check = '1'
                                    itr = iter(dict_vendor[tmp_vc][tmp_ic])
                                    next(itr)
                                    tmp_sc2 = next(itr)
                                    result_upc = next(iter(dict_vendor[tmp_vc][tmp_ic][tmp_sc2]))
                                    result_desc_lookup_base = 'I'
                                    result_desc = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['desc']
                                    result_csize = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['csize']
                                    vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                                    if float(vendor_csize) != float(result_csize):
                                        detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                                    result_csize = vendor_csize
                                    if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                                        result_csize = 1
                                        detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                                    result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['ccost']
                                    result_new_case_cost = tmp_ncost
                                    if tmp_flag_processed_ce == 1:
                                        result_ncost = tmp_ncost * result_csize
                                        result_new_case_cost = tmp_ncost * result_csize
                                        result_amt = tmp_amt * result_csize
                                        result_flag_processed_ce = 0
                                    elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                                        result_ncost = (tmp_ncost / tmp_csize) * result_csize
                                        result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                                        result_amt = (tmp_amt / tmp_csize) * result_csize
                                    result_price = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['price']

                                    result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                                    result_note = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['note']
                                    if flag_ignore_srp_sticker and result_note == 'STICKER':
                                        print(tmp_ic + "\t" + tmp_upc)
                                        result_srp = ' '

                                    tmp_upc3, tmp_upc_cand = getFormalUPC(result_upc, tmp_vc)
                                    if tmp_upc == result_upc or result_upc in tmp_upc_cand:
                                        result_upc_same_check = 1
                                    else:
                                        result_upc_same_check = 0

                                # UPC is duplicated
                                elif len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']) > 1:

                                    if tmp_upc is None:
                                        dict_error_flag['duplicate'] = 0

                                    else:
                                        tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                                        # UPC 기준으로 탐색
                                        flag_exist_upc = False
                                        flag_exist_upc_cand = False
                                        last_sc = ''
                                        idx_upc_cand = -1
                                        for key1 in dict_vendor[tmp_vc][tmp_ic]:
                                            if key1 == 'upc_dict':
                                                continue
                                            for key2 in dict_vendor[tmp_vc][tmp_ic][key1]:
                                                if key2 == tmp_upc2:
                                                    flag_exist_upc = True
                                                    flag_exist_upc_cand = False
                                                    last_sc = key1
                                                    break
                                                if key2 in tmp_upc_cand:
                                                    flag_exist_upc_cand = True
                                                    flag_exist_upc = False
                                                    idx_upc_cand = tmp_upc_cand.index(key2)
                                                    last_sc = key1
                                                    break
                                            if flag_exist_upc or flag_exist_upc_cand:
                                                break

                                        # UPC가 없는 경우
                                        if not flag_exist_upc and not flag_exist_upc_cand:
                                            tmp_upc_cand_list = []
                                            for cand in dict_vendor[tmp_vc][tmp_ic]['upc_dict']:
                                                tmp_upc_cand_list.append(cand)
                                            tmp_upc_cand2.extend(tmp_upc_cand_list)
                                            dict_error_flag['new'] = 1
                                        else:
                                            result_desc_lookup_base = 'U'
                                            if flag_exist_upc:
                                                result_upc = tmp_upc2
                                            else:
                                                result_upc = tmp_upc_cand[idx_upc_cand]
                                            result_desc = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['desc']
                                            result_csize = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['csize']
                                            vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                                            if float(vendor_csize) != float(result_csize):
                                                detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                                            result_csize = vendor_csize
                                            if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                                                result_csize = 1
                                                detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                                            result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc][
                                                'ccost']
                                            result_new_case_cost = tmp_ncost
                                            if tmp_flag_processed_ce == 1:
                                                result_ncost = tmp_ncost * result_csize
                                                result_new_case_cost = tmp_ncost * result_csize
                                                result_amt = tmp_amt * result_csize
                                                result_flag_processed_ce = 0
                                            elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                                                result_ncost = (tmp_ncost / tmp_csize) * result_csize
                                                result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                                                result_amt = (tmp_amt / tmp_csize) * result_csize
                                            result_price = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['price']

                                            result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                                            result_note = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['note']
                                            if flag_ignore_srp_sticker and result_note == 'STICKER':
                                                print(tmp_ic + "\t" + tmp_upc)
                                                result_srp = ' '

                                else:
                                    dict_error_flag['new'] = 1
                            result_ic = tmp_ic
                            result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                            if len(result_cand_ic) == 2:
                                result_ic = result_cand_ic[0]
                        else:
                            result_desc_lookup_base = 'U'
                            if flag_exist_upc:
                                result_upc = tmp_upc2
                            else:
                                result_upc = tmp_upc_cand[idx_upc_cand]
                            result_desc = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['desc']
                            result_csize = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['csize']
                            vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                            if float(vendor_csize) != float(result_csize):
                                detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                            result_csize = vendor_csize
                            if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                                result_csize = 1
                                detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                            result_old_case_cost = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['ccost']
                            result_new_case_cost = tmp_ncost
                            if tmp_flag_processed_ce == 1:
                                result_ncost = tmp_ncost * result_csize
                                result_new_case_cost = tmp_ncost * result_csize
                                result_amt = tmp_amt * result_csize
                                result_flag_processed_ce = 0
                            elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                                result_ncost = (tmp_ncost / tmp_csize) * result_csize
                                result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                                result_amt = (tmp_amt / tmp_csize) * result_csize
                            result_price = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['price']

                            result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                            result_note = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['note']
                            if flag_ignore_srp_sticker and result_note == 'STICKER':
                                print(tmp_ic + "\t" + tmp_upc)
                                result_srp = ' '
                            result_ic = tmp_ic2
                            result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                            if len(result_cand_ic) == 2:
                                result_ic = result_cand_ic[0]

                    else:
                        dict_error_flag['duplicate'] = 1
                        result_ic = tmp_ic
                        result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                        if len(result_cand_ic) == 2:
                            result_ic = result_cand_ic[0]

                # Store Code에는 없고 Item Code로는 UPC가 조회된 경우 <- 이런 경우는 없으니 삭제

            else:
                result_ic = tmp_ic
                result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                if len(result_cand_ic) == 2:
                    result_ic = result_cand_ic[0]

        # L1 기 등록된 상품 (Vendor Code & Item Code 기준)
        elif tmp_ic in dict_vendor[tmp_vc]:

            # L2 Store Code가 없는 경우
            if tmp_sc is None or tmp_sc not in dict_vendor[tmp_vc][tmp_ic]:

                # UPC is sole
                if False and len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']) == 1:
                    result_upc_check = '1'
                    itr = iter(dict_vendor[tmp_vc][tmp_ic])
                    next(itr)
                    tmp_sc2 = next(itr)
                    result_upc = next(iter(dict_vendor[tmp_vc][tmp_ic][tmp_sc2]))
                    result_desc_lookup_base = 'I'
                    result_desc = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['desc']
                    result_csize = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['csize']
                    vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                    if float(vendor_csize) != float(result_csize):
                        detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                    result_csize = vendor_csize
                    if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                        result_csize = 1
                        detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                    result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['ccost']
                    result_new_case_cost = tmp_ncost
                    if tmp_flag_processed_ce == 1:
                        result_ncost = tmp_ncost * result_csize
                        result_new_case_cost = tmp_ncost * result_csize
                        result_amt = tmp_amt * result_csize
                        result_flag_processed_ce = 0
                    elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                        result_ncost = (tmp_ncost / tmp_csize) * result_csize
                        result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                        result_amt = (tmp_amt / tmp_csize) * result_csize
                    result_price = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['price']

                    result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                    result_note = dict_vendor[tmp_vc][tmp_ic][tmp_sc2][result_upc]['note']
                    if flag_ignore_srp_sticker and result_note == 'STICKER':
                        print(tmp_ic + "\t" + tmp_upc)
                        result_srp = ' '

                    tmp_upc_3, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)
                    if result_upc == tmp_upc_3 or result_upc in tmp_upc_cand:
                        result_upc_same_check = 1
                    else:
                        result_upc_same_check = 0

                # UPC is duplicated
                elif False and len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']) > 1:
                    result_upc_check = str(len(dict_vendor[tmp_vc][tmp_ic]['upc_dict']))

                    if tmp_upc is None:
                        dict_error_flag['duplicate'] = 1

                    else:
                        tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                        # UPC 기준으로 탐색
                        flag_exist_upc = False
                        flag_exist_upc_cand = False
                        last_sc = ''
                        last_sc_cand = ''
                        idx_upc_cand = -1
                        for key1 in dict_vendor[tmp_vc][tmp_ic]:
                            if key1 == 'upc_dict':
                                continue
                            for key2 in dict_vendor[tmp_vc][tmp_ic][key1]:
                                if key2 == tmp_upc2:
                                    flag_exist_upc = True
                                    flag_exist_upc_cand = False
                                    last_sc = key1
                                    break
                                if key2 in tmp_upc_cand:
                                    flag_exist_upc_cand = True
                                    flag_exist_upc = False
                                    idx_upc_cand = tmp_upc_cand.index(key2)
                                    last_sc_cand = key1
                                    break
                            if flag_exist_upc or flag_exist_upc_cand:
                                break

                        # UPC가 없는 경우
                        if not flag_exist_upc and not flag_exist_upc_cand:
                            tmp_upc_cand_list = []
                            for cand in dict_vendor[tmp_vc][tmp_ic]['upc_dict']:
                                tmp_upc_cand_list.append(cand)
                            tmp_upc_cand2.extend(tmp_upc_cand_list)
                            dict_error_flag['new'] = 1
                        else:
                            result_desc_lookup_base = 'U'
                            if flag_exist_upc:
                                result_upc = tmp_upc2
                            else:
                                result_upc = tmp_upc_cand[idx_upc_cand]
                                last_sc = last_sc_cand
                            result_desc = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['desc']
                            result_csize = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['csize']
                            vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                            if float(vendor_csize) != float(result_csize):
                                detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                            result_csize = vendor_csize
                            if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                                result_csize = 1
                                detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                            result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['ccost']
                            result_new_case_cost = tmp_ncost
                            if tmp_flag_processed_ce == 1:
                                result_ncost = tmp_ncost * result_csize
                                result_new_case_cost = tmp_ncost * result_csize
                                result_amt = tmp_amt * result_csize
                                result_flag_processed_ce = 0
                            elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                                result_ncost = (tmp_ncost / tmp_csize) * result_csize
                                result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                                result_amt = (tmp_amt / tmp_csize) * result_csize
                            result_price = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['price']

                            result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                            result_note = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['note']
                            if flag_ignore_srp_sticker and result_note == 'STICKER':
                                print(tmp_ic + "\t" + tmp_upc)
                                result_srp = ' '

                result_ic = tmp_ic
                result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                if len(result_cand_ic) == 2:
                    result_ic = result_cand_ic[0]

            # L2 Store Code가 있는 경우
            elif tmp_sc is not None and tmp_sc in dict_vendor[tmp_vc][tmp_ic]:

                # Store Code에 중복없이 하나의 UPC만 있는 경우
                if len(dict_vendor[tmp_vc][tmp_ic][tmp_sc]) == 1:
                    result_upc = next(iter(dict_vendor[tmp_vc][tmp_ic][tmp_sc]))
                    result_desc = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['desc']
                    result_csize = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['csize']
                    vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                    if float(vendor_csize) != float(result_csize):
                        detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                    result_csize = vendor_csize
                    if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                        result_csize = 1
                        detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                    result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['ccost']
                    result_new_case_cost = tmp_ncost
                    if tmp_flag_processed_ce == 1:
                        result_ncost = tmp_ncost * result_csize
                        result_new_case_cost = tmp_ncost * result_csize
                        result_amt = tmp_amt * result_csize
                        result_flag_processed_ce = 0
                    elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                        result_ncost = (tmp_ncost / tmp_csize) * result_csize
                        result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                        result_amt = (tmp_amt / tmp_csize) * result_csize
                    result_price = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['price']

                    result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                    result_note = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['note']
                    if flag_ignore_srp_sticker and result_note == 'STICKER':
                        print(tmp_ic + "\t" + tmp_upc)
                        result_srp = ' '

                    result_upc_check = '1'
                    result_desc_lookup_base = 'I'

                    tmp_upc_3, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)
                    if tmp_upc_3 == result_upc or result_upc in tmp_upc_cand:
                        result_upc_same_check = '1'
                    result_ic = tmp_ic
                    result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                    if len(result_cand_ic) == 2:
                        result_ic = result_cand_ic[0]

                elif len(dict_vendor[tmp_vc][tmp_ic][tmp_sc]) > 1:
                    # print('#2-1-2-2')
                    result_upc_check = str(len(dict_vendor[tmp_vc][tmp_ic][tmp_sc]))

                    if tmp_upc is not None:
                        tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                        # UPC 기준으로 탐색
                        flag_exist_upc = False
                        flag_exist_upc_cand = False
                        flag_exist_exact = False
                        tmp_ic2 = ''
                        last_sc = tmp_sc
                        idx_upc_cand = -1
                        for key1 in dict_vendor[tmp_vc]:
                            if tmp_sc in dict_vendor[tmp_vc][key1]:
                                for key3 in dict_vendor[tmp_vc][key1][tmp_sc]:
                                    if key3 == tmp_upc2:
                                        flag_exist_upc = True
                                        flag_exist_upc_cand = False
                                        tmp_ic2 = key1
                                        flag_exist_exact = True
                                        break
                                    if key3 in tmp_upc_cand:
                                        flag_exist_upc_cand = True
                                        flag_exist_upc = False
                                        idx_upc_cand = tmp_upc_cand.index(key3)
                                        tmp_ic2 = key1
                                        flag_exist_exact = True
                                        break
                            if flag_exist_exact:
                                break
                            '''
                            for key2 in dict_vendor[tmp_vc][key1]:
                                if key2 == 'upc_dict':
                                    continue
                                for key3 in dict_vendor[tmp_vc][key1][key2]:
                                    if key3 == tmp_upc2:
                                        flag_exist_upc = True
                                        flag_exist_upc_cand = False
                                        tmp_ic2 = key1
                                        last_sc = key2
                                        if key2 == tmp_sc:
                                            flag_exist_exact = True
                                            break
                                    if key3 in tmp_upc_cand:
                                        flag_exist_upc_cand = True
                                        flag_exist_upc = False
                                        idx_upc_cand = tmp_upc_cand.index(key3)
                                        tmp_ic2 = key1
                                        last_sc = key2
                                        if key2 == tmp_sc:
                                            flag_exist_exact = True
                                            break
                                if flag_exist_exact:
                                    break
                            if flag_exist_exact:
                                break
                            '''

                        # UPC가 Store 내에 없는 경우
                        if not flag_exist_upc and not flag_exist_upc_cand:
                            tmp_upc_cand_list = []
                            for cand in dict_vendor[tmp_vc][tmp_ic][tmp_sc]:
                                tmp_upc_cand_list.append(cand)
                            tmp_upc_cand2.extend(tmp_upc_cand_list)
                            if len(dict_vendor[tmp_vc][tmp_ic][tmp_sc].keys()) > 0:

                                # UPC is sole
                                if len(dict_vendor[tmp_vc][tmp_ic][tmp_sc].keys()) == 1:
                                    result_upc_check = '1'
                                    # itr = iter(dict_vendor[tmp_vc][tmp_ic])
                                    # next(itr)
                                    result_upc = next(iter(dict_vendor[tmp_vc][tmp_ic][tmp_sc]))
                                    result_desc_lookup_base = 'I'
                                    result_desc = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['desc']
                                    result_csize = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['csize']
                                    vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                                    if float(vendor_csize) != float(result_csize):
                                        detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                                    result_csize = vendor_csize
                                    if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                                        result_csize = 1
                                        detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                                    result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['ccost']
                                    result_new_case_cost = tmp_ncost
                                    if tmp_flag_processed_ce == 1:
                                        result_ncost = tmp_ncost * result_csize
                                        result_new_case_cost = tmp_ncost * result_csize
                                        result_amt = tmp_amt * result_csize
                                        result_flag_processed_ce = 0
                                    elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                                        result_ncost = (tmp_ncost / tmp_csize) * result_csize
                                        result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                                        result_amt = (tmp_amt / tmp_csize) * result_csize
                                    result_price = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['price']

                                    result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                                    result_note = dict_vendor[tmp_vc][tmp_ic][tmp_sc][result_upc]['note']
                                    if flag_ignore_srp_sticker and result_note == 'STICKER':
                                        print(tmp_ic + "\t" + tmp_upc)
                                        result_srp = ' '

                                    tmp_upc3, tmp_upc_cand = getFormalUPC(result_upc, tmp_vc)
                                    if tmp_upc == result_upc or result_upc in tmp_upc_cand:
                                        result_upc_same_check = 1
                                    else:
                                        result_upc_same_check = 0

                                # UPC is duplicated
                                elif len(dict_vendor[tmp_vc][tmp_ic][tmp_sc].keys()) > 1:

                                    if tmp_upc is None:
                                        dict_error_flag['duplicate'] = 0

                                    else:
                                        tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                                        # UPC 기준으로 탐색
                                        flag_exist_upc = False
                                        flag_exist_upc_cand = False
                                        last_sc = tmp_sc
                                        idx_upc_cand = -1
                                        if tmp_sc in dict_vendor[tmp_vc][tmp_ic]:
                                            for key2 in dict_vendor[tmp_vc][tmp_ic][tmp_sc]:
                                                if key2 == tmp_upc2:
                                                    flag_exist_upc = True
                                                    flag_exist_upc_cand = False
                                                    break
                                                if key2 in tmp_upc_cand:
                                                    flag_exist_upc_cand = True
                                                    flag_exist_upc = False
                                                    idx_upc_cand = tmp_upc_cand.index(key2)
                                                    break
                                        '''
                                        for key1 in dict_vendor[tmp_vc][tmp_ic]:
                                            if key1 == 'upc_dict':
                                                continue
                                            for key2 in dict_vendor[tmp_vc][tmp_ic][key1]:
                                                if key2 == tmp_upc2:
                                                    flag_exist_upc = True
                                                    flag_exist_upc_cand = False
                                                    last_sc = key1
                                                    break
                                                if key2 in tmp_upc_cand:
                                                    flag_exist_upc_cand = True
                                                    flag_exist_upc = False
                                                    idx_upc_cand = tmp_upc_cand.index(key2)
                                                    last_sc = key1
                                                    break
                                            if flag_exist_upc or flag_exist_upc_cand:
                                                break
                                        '''

                                        # UPC가 없는 경우
                                        if not flag_exist_upc and not flag_exist_upc_cand:
                                            tmp_upc_cand_list = []
                                            for cand in dict_vendor[tmp_vc][tmp_ic]['upc_dict']:
                                                tmp_upc_cand_list.append(cand)
                                            tmp_upc_cand2.extend(tmp_upc_cand_list)
                                            dict_error_flag['new'] = 1
                                        else:
                                            result_desc_lookup_base = 'U'
                                            if flag_exist_upc:
                                                result_upc = tmp_upc2
                                            else:
                                                result_upc = tmp_upc_cand[idx_upc_cand]
                                            result_desc = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['desc']
                                            result_csize = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['csize']
                                            vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                                            if float(vendor_csize) != float(result_csize):
                                                detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                                            result_csize = vendor_csize
                                            if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                                                result_csize = 1
                                                detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                                            result_old_case_cost = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc][
                                                'ccost']
                                            result_new_case_cost = tmp_ncost
                                            if tmp_flag_processed_ce == 1:
                                                result_ncost = tmp_ncost * result_csize
                                                result_new_case_cost = tmp_ncost * result_csize
                                                result_amt = tmp_amt * result_csize
                                                result_flag_processed_ce = 0
                                            elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                                                result_ncost = (tmp_ncost / tmp_csize) * result_csize
                                                result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                                                result_amt = (tmp_amt / tmp_csize) * result_csize
                                            result_price = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['price']

                                            result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                                            result_note = dict_vendor[tmp_vc][tmp_ic][last_sc][result_upc]['note']
                                            if flag_ignore_srp_sticker and result_note == 'STICKER':
                                                print(tmp_ic + "\t" + tmp_upc)
                                                result_srp = ' '

                                else:
                                    dict_error_flag['new'] = 1

                            result_ic = tmp_ic
                            result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                            if len(result_cand_ic) == 2:
                                result_ic = result_cand_ic[0]
                        else:
                            result_desc_lookup_base = 'U'
                            if flag_exist_upc:
                                result_upc = tmp_upc2
                            else:
                                result_upc = tmp_upc_cand[idx_upc_cand]
                            result_desc = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['desc']
                            result_csize = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['csize']
                            vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                            if float(vendor_csize) != float(result_csize):
                                detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                            result_csize = vendor_csize
                            if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                                result_csize = 1
                                detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                            result_old_case_cost = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['ccost']
                            result_new_case_cost = tmp_ncost
                            if tmp_flag_processed_ce == 1:
                                result_ncost = tmp_ncost * result_csize
                                result_new_case_cost = tmp_ncost * result_csize
                                result_amt = tmp_amt * result_csize
                                result_flag_processed_ce = 0
                            elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                                result_ncost = (tmp_ncost / tmp_csize) * result_csize
                                result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                                result_amt = (tmp_amt / tmp_csize) * result_csize
                            result_price = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['price']

                            result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                            result_note = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['note']
                            if flag_ignore_srp_sticker and result_note == 'STICKER':
                                print(tmp_ic + "\t" + tmp_upc)
                                result_srp = ' '
                            result_ic = tmp_ic2
                            result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                            if len(result_cand_ic) == 2:
                                result_ic = result_cand_ic[0]

                    else:
                        dict_error_flag['duplicate'] = 1
                        result_ic = tmp_ic
                        result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                        if len(result_cand_ic) == 2:
                            result_ic = result_cand_ic[0]

                # Store Code에는 없고 Item Code로는 UPC가 조회된 경우 <= 이런 경우는 없으니깐 지움

        # L1 미 등록된 상품 중 인보이스에 UPC가 있는 경우 (신상품)
        elif tmp_upc is not None:

            if tmp_sc is None:
                tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                # UPC 기준으로 탐색
                flag_exist_upc = False
                flag_exist_upc_cand = False
                tmp_ic2 = ''
                last_sc = ''
                idx_upc_cand = -1
                for key1 in dict_vendor[tmp_vc]:
                    for key2 in dict_vendor[tmp_vc][key1]:
                        if key2 == 'upc_dict':
                            continue
                        for key3 in dict_vendor[tmp_vc][key1][key2]:
                            if key3 == tmp_upc2:
                                flag_exist_upc = True
                                flag_exist_upc_cand = False
                                tmp_ic2 = key1
                                last_sc = key2
                                break
                            if key3 in tmp_upc_cand:
                                flag_exist_upc_cand = True
                                flag_exist_upc = False
                                idx_upc_cand = tmp_upc_cand.index(key3)
                                tmp_ic2 = key1
                                last_sc = key2
                                break
                        if flag_exist_upc or flag_exist_upc_cand:
                            break
                    if flag_exist_upc or flag_exist_upc_cand:
                        break

                # UPC가 없는 경우
                if not flag_exist_upc and not flag_exist_upc_cand:
                    dict_error_flag['new'] = 1
                    result_ic = tmp_ic
                    result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                    if len(result_cand_ic) == 2:
                        result_ic = result_cand_ic[0]
                else:
                    result_desc_lookup_base = 'U'
                    if flag_exist_upc:
                        result_upc = tmp_upc2
                    else:
                        result_upc = tmp_upc_cand[idx_upc_cand]
                    result_desc = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['desc']
                    result_csize = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['csize']
                    vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                    if float(vendor_csize) != float(result_csize):
                        detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                    result_csize = vendor_csize
                    if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                        result_csize = 1
                        detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                    result_old_case_cost = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['ccost']
                    result_new_case_cost = tmp_ncost
                    if tmp_flag_processed_ce == 1:
                        result_ncost = tmp_ncost * result_csize
                        result_new_case_cost = tmp_ncost * result_csize
                        result_amt = tmp_amt * result_csize
                        result_flag_processed_ce = 0
                    elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                        result_ncost = (tmp_ncost / tmp_csize) * result_csize
                        result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                        result_amt = (tmp_amt / tmp_csize) * result_csize
                    result_price = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['price']
                    result_ic = tmp_ic2
                    result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                    if len(result_cand_ic) == 2:
                        result_ic = result_cand_ic[0]

                    result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                    result_note = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['note']
                    if flag_ignore_srp_sticker and result_note == 'STICKER':
                        print(tmp_ic + "\t" + tmp_upc)
                        result_srp = ' '

            else:
                tmp_upc2, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)

                # UPC 기준으로 탐색
                flag_exist_upc = False
                flag_exist_upc_cand = False
                flag_exist_exact = False
                tmp_ic2 = ''
                last_sc = tmp_sc
                tmp_ic2_cand = ''
                last_sc_cand = tmp_sc
                idx_upc_cand = -1
                # print('#2-2-1')
                for key1 in dict_vendor[tmp_vc]:
                    if tmp_sc in dict_vendor[tmp_vc][key1]:
                        for key3 in dict_vendor[tmp_vc][key1][tmp_sc]:
                            if key3 == tmp_upc2:
                                flag_exist_upc = True
                                flag_exist_upc_cand = False
                                tmp_ic2 = key1
                                flag_exist_exact = True
                                break
                            if key3 in tmp_upc_cand:
                                flag_exist_upc_cand = True
                                flag_exist_upc = False
                                idx_upc_cand = tmp_upc_cand.index(key3)
                                tmp_ic2_cand = key1
                                flag_exist_exact = True
                                break
                        if flag_exist_exact:
                            break
                    '''
                    for key2 in dict_vendor[tmp_vc][key1]:
                        if key2 == 'upc_dict':
                            continue
                        for key3 in dict_vendor[tmp_vc][key1][key2]:
                            if key3 == tmp_upc2:
                                flag_exist_upc = True
                                flag_exist_upc_cand = False
                                tmp_ic2 = key1
                                last_sc = key2
                                if key2 == tmp_sc:
                                    flag_exist_exact = True
                                    break
                            if key3 in tmp_upc_cand:
                                flag_exist_upc_cand = True
                                flag_exist_upc = False
                                idx_upc_cand = tmp_upc_cand.index(key3)
                                tmp_ic2_cand = key1
                                last_sc_cand = key2
                                if key2 == tmp_sc:
                                    flag_exist_exact = True
                                    break
                        if flag_exist_exact:
                            break
                    if flag_exist_exact:
                        break
                    '''

                # UPC가 없는 경우
                if not flag_exist_upc and not flag_exist_upc_cand:
                    dict_error_flag['new'] = 1
                    result_ic = tmp_ic
                    result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                    if len(result_cand_ic) == 2:
                        result_ic = result_cand_ic[0]
                else:
                    result_desc_lookup_base = 'U'
                    if flag_exist_upc:
                        result_upc = tmp_upc2
                    else:
                        result_upc = tmp_upc_cand[idx_upc_cand]
                        tmp_ic2 = tmp_ic2_cand
                        last_sc = last_sc_cand
                    result_desc = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['desc']
                    result_csize = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['csize']
                    vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
                    if float(vendor_csize) != float(result_csize):
                        detail_cat = '업체 데이터로 Lookup해온 F19 대체'
                    result_csize = vendor_csize
                    if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
                        result_csize = 1
                        detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'
                    result_old_case_cost = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['ccost']
                    result_new_case_cost = tmp_ncost
                    if tmp_flag_processed_ce == 1:
                        result_ncost = tmp_ncost * result_csize
                        result_new_case_cost = tmp_ncost * result_csize
                        result_amt = tmp_amt * result_csize
                        result_flag_processed_ce = 0
                    elif param_etc['step'] == 4 and tmp_c_or_e in ['e', 'E']:
                        result_ncost = (tmp_ncost / tmp_csize) * result_csize
                        result_new_case_cost = (tmp_ncost / tmp_csize) * result_csize
                        result_amt = (tmp_amt / tmp_csize) * result_csize
                    result_price = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['price']
                    result_ic = tmp_ic2
                    result_cand_ic = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                    if len(result_cand_ic) == 2:
                        result_ic = result_cand_ic[0]

                    result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat)
                    result_note = dict_vendor[tmp_vc][tmp_ic2][last_sc][result_upc]['note']
                    if flag_ignore_srp_sticker and result_note == 'STICKER':
                        print(tmp_ic + "\t" + tmp_upc)
                        result_srp = ' '

    if result_desc_lookup_base == 'N':
        vendor_csize = lookup_vendor_ic(g_data, result_csize, tmp_vc, tmp_sc, tmp_ic_cand, tmp_upc)
        if param_etc['step'] == 4 and float(result_csize) != float(vendor_csize):
            detail_cat = '업체 데이터로 직접 입력한 F19 대체'
        result_csize = vendor_csize
        if param_etc['step'] == 4 and tmp_vc in ['1205', '1210']:
            result_csize = 1
            detail_cat = 'Frito-ray 및 Cloverland F19 1로 고정'

    return result_upc, result_ic, result_desc, result_csize, result_ncost, result_amt, result_flag_processed_ce, result_upc_check, result_upc_same_check, result_desc_lookup_base, result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp, tmp_upc_cand2, detail_cat, result_note


def getSrpSet(result_csize, result_old_case_cost, result_new_case_cost, result_price, tmp_vat):

    vat_depart = {
        'CIGARETTES': 0,
        'DAIRY': 12,
        'DELI': 0,
        'EGGS': 0.2,
        'ETHNIC': 2.5,
        'FROZ BAKERY': 8,
        'FROZ FOOD': 21,
        'SUPPLIES': 6,
        'GEN MDSE': 13.5,
        'GROCERY': 11,
        'HBA': 14.5,
        'ICE': 23,
        'MEAT': 0,
        'MILK': 0,
        'PRODUCE': 0
    }

    result_old_margin = ''
    result_new_margin = ''
    result_srp = ''
    if result_csize is None or result_csize == '' or result_old_case_cost is None or result_old_case_cost == '' or result_new_case_cost is None or result_new_case_cost == '' or result_price is None or result_price == '':
        result_old_case_cost = ''
        result_new_case_cost = ''
        result_price = ''
    else:
        if float(result_csize) > 0.0 and float(result_price) > 0.0:
            result_old_margin = (float(result_price) - (
                    (float(result_old_case_cost) / float(result_csize)) * (
                    1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
            result_new_margin = (float(result_price) - (
                    (float(result_new_case_cost) / float(result_csize)) * (
                    1.0 + (float(tmp_vat) / 100.0)))) / float(result_price)
            if float(result_old_case_cost) < float(result_new_case_cost):
                if float(result_old_margin) == 1.0:
                    converted_price = get_converted_price(float(result_price))
                    if flag_enable_cp_price and converted_price != result_price:
                        if flag_margin_diff:
                            cp_margin = get_margin(converted_price, result_new_case_cost, result_csize, tmp_vat)
                            margin_diff = float(cp_margin) - float(result_old_margin)
                            if -0.03 <= margin_diff <= 0.03:
                                result_srp = ' '
                            else:
                                result_srp = converted_price
                        else:
                            result_srp = converted_price
                    else:
                        result_srp = ' '
                else:
                    if flag_margin_diff:
                        converted_price = get_converted_price(float(result_price))
                        cp_margin = get_margin(converted_price, result_new_case_cost, result_csize, tmp_vat)
                        margin_diff = float(cp_margin) - float(result_old_margin)
                        if -0.03 <= margin_diff <= 0.03:
                            result_srp = ' '
                        else:
                            result_srp = round(((float(result_new_case_cost) / float(result_csize)) * (
                                1 + (float(tmp_vat) / 100.0))) / (1.0 - float(result_old_margin)), 1) - 0.01
                    else:
                        result_srp = round(((float(result_new_case_cost) / float(result_csize)) * (
                            1 + (float(tmp_vat) / 100.0))) / (1.0 - float(result_old_margin)), 1) - 0.01
            else:
                converted_price = get_converted_price(float(result_price))
                if flag_enable_cp_price and converted_price != result_price:
                    if flag_margin_diff:
                        cp_margin = get_margin(converted_price, result_new_case_cost, result_csize, tmp_vat)
                        margin_diff = float(cp_margin) - float(result_old_margin)
                        if -0.03 <= margin_diff <= 0.03:
                            result_srp = ' '
                        else:
                            result_srp = converted_price
                    else:
                        result_srp = converted_price
                else:
                    result_srp = ' '
    return result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp


def convert_store_data(f):

    dict_vendor = dict()
    print('converting...'+f)

    file_dir = './org_store_data'

    file_dir_targ = './processed_store_data'
    file_dir_targ_bin = './bin_store_data'

    file = join(file_dir, f)
    if isfile(file):
        file_targ = join(file_dir_targ, f)
        file_targ_bin = join(file_dir_targ_bin, f)

        wb_t = load_workbook(file)
        # 각 점포(시트) 별 순회
        for ws_t in wb_t:
            dict_vendor[ws_t.title] = dict()
            idx1 = 0
            for row in ws_t.rows:
                # 첫 행(컬럼 명) 통과
                if idx1 == 0:
                    idx1 += 1
                    continue

                # 아이템 코드 부분이 공백인 경우 통과
                if row[0].value is None or row[0].value.replace(' ', '') == '':
                    continue

                # 벤더 코드 부분이 공백인 경우 통과
                if row[4].value is None or row[4].value.replace(' ', '') == '':
                    continue

                # UPC 공백인 경우 통과
                if row[1].value is None or row[1].value.replace(' ', '') == '':
                    continue

                # Global all vendors 1500번 무시
                if row[4].value == '1500':
                    continue

                last_ic = row[0].value
                last_vc = row[4].value

                '''
                print(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[7].value, row[8].value, row[9].value)
                print(last_ic)
                print(row[4].value in dict_vendor)
                if row[4].value in dict_vendor:
                    print(last_ic in dict_vendor[row[4].value])
                    if last_ic in dict_vendor[row[4].value]:
                        print(dict_vendor[row[4].value][last_ic])

                input('!')
                '''

                # Vendor가 이전에 등장했는지 확인
                # 재등장 시 update, 최초로 등장 시 insert
                if last_vc not in dict_vendor[ws_t.title]:
                    dict_vendor[ws_t.title][last_vc] = dict()
                    dict_vendor[ws_t.title][last_vc][last_ic] = dict()
                    dict_vendor[ws_t.title][last_vc][last_ic][row[1].value] = dict()
                    dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['desc'] = row[11].value
                    dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['csize'] = row[2].value
                    dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['ccost'] = row[8].value
                    dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['price'] = row[5].value
                    if len(row) >= 16:
                        dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['note'] = row[15].value
                    else:
                        dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['note'] = ''
                else:
                    # 기존 벤더에 last_ic가 있는 경우 UPC 중복 여부를 확인해서 같으면 값이 있는 부분만 갱신하고 다르면 추가(중복)
                    if last_ic in dict_vendor[ws_t.title][last_vc]:
                        if row[1].value not in dict_vendor[ws_t.title][last_vc][last_ic]:
                            dict_vendor[ws_t.title][last_vc][last_ic][row[1].value] = dict()
                            dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['desc'] = row[11].value
                            dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['csize'] = row[2].value
                            dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['ccost'] = row[8].value
                            dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['price'] = row[5].value
                            if len(row) >= 16:
                                dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['note'] = row[15].value
                            else:
                                dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['note'] = ''
                        else:
                            if row[11].value is not None and row[11].value.replace(' ', '') != '':
                                dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['desc'] = row[11].value
                            if row[2].value is not None and str(row[2].value).replace(' ', '') != '':
                                dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['csize'] = row[2].value
                            if row[8].value is not None and str(row[8].value).replace(' ', '') != '':
                                dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['ccost'] = row[8].value
                            if row[5].value is not None and str(row[5].value).replace(' ', '') != '':
                                dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['price'] = row[5].value
                            if len(row) >= 16:
                                dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['note'] = row[15].value
                            else:
                                dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['note'] = ''
                    else:
                        dict_vendor[ws_t.title][last_vc][last_ic] = dict()
                        dict_vendor[ws_t.title][last_vc][last_ic][row[1].value] = dict()
                        dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['desc'] = row[11].value
                        dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['csize'] = row[2].value
                        dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['ccost'] = row[8].value
                        dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['price'] = row[5].value
                        if len(row) >= 16:
                            dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['note'] = row[15].value
                        else:
                            dict_vendor[ws_t.title][last_vc][last_ic][row[1].value]['note'] = ''
        wb_t.close()

        print('building dictionary...')
        wb_n = Workbook()
        ws_n = wb_n.active
        ws_n.title = 'Sheet1'
        ws_n.append(
            ['Store Code', 'Vendor Code', 'Item Code', 'UPC', 'HELPER', 'Description', 'Case Size', 'Case Cost',
             'Price', 'Note'])
        dict_vendor_bin = dict()

        for sc in dict_vendor:
            for vc in dict_vendor[sc]:
                for ic in dict_vendor[sc][vc]:
                    for upc in dict_vendor[sc][vc][ic]:
                        ws_n.append([sc, vc, ic, upc, vc + upc, dict_vendor[sc][vc][ic][upc]['desc'],
                                     dict_vendor[sc][vc][ic][upc]['csize'], dict_vendor[sc][vc][ic][upc]['ccost'],
                                     dict_vendor[sc][vc][ic][upc]['price'], dict_vendor[sc][vc][ic][upc]['note']])

                        tmp_vc = vc
                        tmp_ic = ic
                        tmp_sc = sc
                        tmp_upc = upc
                        tmp_desc = dict_vendor[sc][vc][ic][upc]['desc']
                        tmp_csize = dict_vendor[sc][vc][ic][upc]['csize']
                        tmp_ccost = dict_vendor[sc][vc][ic][upc]['ccost']
                        tmp_price = dict_vendor[sc][vc][ic][upc]['price']
                        tmp_note = dict_vendor[sc][vc][ic][upc]['note']

                        if tmp_vc not in dict_vendor_bin:
                            dict_vendor_bin[tmp_vc] = dict()
                        if tmp_ic not in dict_vendor_bin[tmp_vc]:
                            dict_vendor_bin[tmp_vc][tmp_ic] = dict()
                            dict_vendor_bin[tmp_vc][tmp_ic]['upc_dict'] = dict()
                        if tmp_sc not in dict_vendor_bin[tmp_vc][tmp_ic]:
                            dict_vendor_bin[tmp_vc][tmp_ic][tmp_sc] = dict()
                        if tmp_upc not in dict_vendor_bin[tmp_vc][tmp_ic][tmp_sc]:
                            dict_vendor_bin[tmp_vc][tmp_ic][tmp_sc][tmp_upc] = dict()
                            dict_vendor_bin[tmp_vc][tmp_ic][tmp_sc][tmp_upc]['desc'] = tmp_desc
                            dict_vendor_bin[tmp_vc][tmp_ic][tmp_sc][tmp_upc]['csize'] = tmp_csize
                            dict_vendor_bin[tmp_vc][tmp_ic][tmp_sc][tmp_upc]['ccost'] = tmp_ccost
                            dict_vendor_bin[tmp_vc][tmp_ic][tmp_sc][tmp_upc]['price'] = tmp_price
                            dict_vendor_bin[tmp_vc][tmp_ic][tmp_sc][tmp_upc]['note'] = tmp_note
                            dict_vendor_bin[tmp_vc][tmp_ic]['upc_dict'][tmp_upc] = 1

        wb_n.save(file_targ)
        wb_n.close()

        f = open(file_targ_bin+'.pkl', "wb")
        pickle.dump(dict_vendor_bin, f)
        f.close()

        return file_targ


def thread_convert_dict_vendor(thread_id, g_data):
    g_data['status_dict_vendor'] = 2
    convert_store_data(g_data['fname_dict_vendor'])
    g_data['status_dict_vendor'] = 3
    load_dict_vendor(g_data)


def thread_setup_dict_vendor(thread_id, fname_dict_vendor, g_data):
    g_data['dict_vendor'] = load_store_db(fname_dict_vendor)
    g_data['dict_store'], g_data['dict_vend'] = load_base_db(g_data)
    # g_data['dict_store'], g_data['dict_vend'], g_data['dict_item_code'] = load_base_db(g_data)
    write_config(g_data)
    g_data['status_dict_vendor'] = 9


def convert_dict_vendor(g_data):
    th = threading.Thread(target=thread_convert_dict_vendor, args=(0, g_data,))
    th.start()


def load_dict_vendor(g_data):
    print('Loading '+g_data['fname_dict_vendor'])
    th = threading.Thread(target=thread_setup_dict_vendor, args=(0, g_data['fname_dict_vendor'], g_data,))
    th.start()
    th.join()


def write_config(g_data):
    f_config = open('config.txt', 'w')
    f_config.write(g_data['fname_dict_vendor'])
    f_config.close()


def read_config(g_data):
    if not isfile('config.txt'):
        f_config = open('config.txt', 'w')
        f_config.write('\n')
        f_config.close()
    f_config = open('config.txt', 'r')
    content = f_config.readlines()
    if len(content) > 0:
        g_data['fname_dict_vendor'] = content[0]
    else:
        g_data['fname_dict_vendor'] = ''
    #print(g_data['fname_dict_vendor'])
    if g_data['fname_dict_vendor'] != '':
        g_data['status_dict_vendor'] = 3
        load_dict_vendor(g_data)
    f_config.close()


def dict_to_file(dict_link, dict_vendor_link, ws_link):

    for u_key in dict_link:
        arr2 = []
        idx3 = -1
        for u_key2 in dict_link[u_key]:
            idx3 += 1
            if idx3 == 1:
                arr2.append(u_key2)
                if dict_link[u_key][16] == 'N':
                    if int(dict_link[u_key][3]) > 1 and (dict_link[u_key][6] is None or dict_link[u_key][6] == ''):
                        arr2.append('업체+제품 검색 시 다수이나 인보이스 내 UPC 없음')
                        arr2.append(dict_link[u_key][33])
                        arr2.append('Description 및 F19 입력')
                        arr2.append('')
                    elif (dict_link[u_key][12] is None or dict_link[u_key][12] == '') and (
                            dict_link[u_key][6] is None or dict_link[u_key][6] == ''):
                        arr2.append('인보이스에 제품코드와 UPC 둘 다 없음')
                        arr2.append(dict_link[u_key][33])
                        arr2.append('제품 코드, UPC 없는지 확인 후 Description 및 F19 입력')
                        arr2.append('')
                    elif dict_link[u_key][10] not in dict_vendor_link:
                        arr2.append('매장 데이터에 등록되지 않은 업체 코드')
                        arr2.append(dict_link[u_key][33])
                        arr2.append('업체 코드 재확인 후 Description 및 F19 입력')
                        arr2.append('')
                    elif int(dict_link[u_key][3]) == 0:
                        arr2.append('신상품')
                        arr2.append(dict_link[u_key][33])
                        arr2.append('제품 코드와 UPC 및 업체 코드 재확인 후 Description 및 F19 입력')
                        arr2.append('')
                    else:
                        arr2.append('업체+제품 검색 시 다수이나 인보이스의 UPC와 상이함')
                        arr2.append(dict_link[u_key][33])
                        arr2.append('UPC 재확인 후 Description 및 F19 입력')
                        arr2.append('')
                elif dict_link[u_key][16] == 'I' and dict_link[u_key][5] == 0 and dict_link[u_key][6] is not None and \
                        dict_link[u_key][6] != '':
                    arr2.append('업체+제품으로 검색된 UPC가 단일이지만 인보이스의 UPC와 상이')
                    arr2.append(dict_link[u_key][33])
                    arr2.append('UPC 재확인')
                    arr2.append('')
                elif dict_link[u_key][16] == 'U' and dict_link[u_key][15] == 1:
                    arr2.append('UPC로 검색된 제품코드가 인보이스의 제품코드와 상이')
                    arr2.append(dict_link[u_key][33])
                    arr2.append('제품코드 재확인')
                    arr2.append('')
                else:
                    arr2.append('')
                    arr2.append(dict_link[u_key][33])
                    arr2.append('')
                    arr2.append('')
            elif idx3 == 20:
                if dict_link[u_key][26] > 0:
                    arr2.append(str(dict_link[u_key][24] / dict_link[u_key][23]))
                    result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp = \
                        getSrpSet(dict_link[u_key][19], dict_link[u_key][27], arr2[-1], dict_link[u_key][29], dict_link[u_key][25])
                    dict_link[u_key][27] = result_old_case_cost
                    dict_link[u_key][28] = result_new_case_cost
                    dict_link[u_key][29] = result_price
                    dict_link[u_key][30] = result_old_margin
                    dict_link[u_key][31] = result_new_margin
                    dict_link[u_key][32] = result_srp
                else:
                    arr2.append(u_key2)
            elif idx3 == 23 or idx3 == 24:
                arr2.append(str(u_key2))
            elif u_key2 is None:
                arr2.append('')
            elif idx3 < 33:
                arr2.append(u_key2)
            elif idx3 > 33:
                arr2.append(u_key2)
        ws_link.append(arr2)


def scheduler_step_two(g_data):
    list_th = []
    for step_unit in g_data['list_step_two']:
        if g_data['list_step_two'][step_unit] == 0:
            th = threading.Thread(target=process_step_two, args=(0, g_data, step_unit,))
            g_data['list_step_two'][step_unit] = 1
            th.start()
            print('thread started!')
            list_th.append(th)

    for th_unit in list_th:
        th_unit.join()


def scheduler_step_four(g_data):
    list_th = []
    for step_unit in g_data['list_step_four']:
        if g_data['list_step_four'][step_unit] == 0:
            th = threading.Thread(target=process_step_four, args=(0, g_data, step_unit,))
            g_data['list_step_four'][step_unit] = 1
            th.start()
            list_th.append(th)

    for th_unit in list_th:
        th_unit.join()


def process_step_two(thread_id, g_data, param1):
    len_last = 0
    len_tmp = 0

    last_ic = ''
    last_vc = ''

    flag_error = 0

    count_step_one = 0

    file_dir_target = './input_step_2'
    if not isdir(file_dir_target):
        mkdir(file_dir_target)
    if not isdir(file_dir_target+'/'+param1):
        mkdir(file_dir_target+'/'+param1)
    file_error_log = './error/'+param1+'.txt'
    if not isdir('./error'):
        mkdir('./error')
    output_file_nm = param1

    f_error_log = open(file_error_log, 'a')
    f_error_log.write('\nSTART\n')
    f_error_log.close()

    dict_vendor = g_data['dict_vendor']
    dict_vend = g_data['dict_vend']

    wb_n = Workbook()
    ws_n = wb_n.active
    ws_n.title = 'Sheet1'

    param_etc = {}
    param_etc['tag_is'] = str(time.time())
    param_etc['step'] = 2

    title_flag = 1
    primary_key = 0
    file_no = 0
    total_file = len(listdir(file_dir_target+'/'+param1))
    # 여러 파일일 경우 하나씩 순회
    for f in listdir(file_dir_target+'/'+param1):
        file_targ = join(file_dir_target+'/'+param1, f)
        file_no += 1
        # print('processing {0} of {1}\n'.format(file_no, total_file))
        if isfile(file_targ):
            print('step 2 processing... -> '+str(f))
            file_name = f.rsplit('.', 1)[0]
            file_ext = f.rsplit('.', 1)[1]

            dict_error_flag = dict()

            dict_unique = dict()

            wb_t = load_workbook(file_targ)
            # print(file_targ)
            ws_t = wb_t['Sheet1']

            last_date = ''

            idx1 = -1
            try:
                for row in ws_t.rows:
                    idx1 += 1

                    if idx1 == 0:
                        if title_flag == 1:
                            arr1 = ['PK', '파일명', '분류', '세부 분류', '작업 내용', '처리 내용']
                            idx2 = 0
                            for col in row:
                                arr1.append(col.value)
                                if idx2 == 0:
                                    arr1.append('UPC check')
                                    arr1.append('lookup UPC')
                                    arr1.append('UPC 동일 체크')
                                    arr1.append('원본 UPC')
                                    arr1.append('후보 UPC')
                                if idx2 == 5:
                                    arr1.append('Lookup F26')
                                    arr1.append('원본 F26')
                                    arr1.append('Lookup F26 동일 체크')
                                    arr1.append('Description Lookup 기준')
                                if col.value == 'F1122':
                                    break
                                idx2 += 1
                            arr1.append('DupFlag')
                            arr1.append('OLD CASE COST')
                            arr1.append('NEW CASE COST')
                            arr1.append('PRICE')
                            arr1.append('OLD MARGIN')
                            arr1.append('NEW MARGIN')
                            arr1.append('SRP')
                            arr1.append('C or E')
                            arr1.append('flagProcessedCE')
                            arr1.append('department')
                            ws_n.append(arr1)
                            title_flag = 0
                        continue
                    elif idx1 == 1:
                        continue

                    tmp_upc = str(row[0].value)
                    if row[0].value is None:
                        tmp_upc = ''
                    tmp_sc = str(row[2].value)
                    if row[2].value is None:
                        tmp_sc = ''
                    tmp_vc = str(row[3].value)
                    if row[3].value is None:
                        tmp_vc = ''
                    tmp_ic = str(row[5].value)
                    if row[5].value is None:
                        tmp_ic = ''
                    tmp_desc = row[6].value
                    tmp_csize = ''
                    tmp_qty = row[12].value
                    tmp_ncost = row[9].value
                    tmp_amt = row[13].value
                    tmp_c_or_e = row[15].value
                    if tmp_c_or_e is None:
                        tmp_c_or_e = ''
                    tmp_date = row[10].value
                    tmp_upc_cand2 = []
                    if last_date != tmp_date:
                        last_date = tmp_date
                        dict_to_file(dict_unique, dict_vendor, ws_n)
                        dict_unique = dict()
                    if (tmp_vc is None or tmp_vc == '') and (tmp_ic is None or tmp_ic == '') and (tmp_sc is None or tmp_sc == '') and (tmp_upc is None or tmp_upc == ''):
                        break
                    if tmp_vc in ('0202', '0203') and tmp_c_or_e not in ['c', 'C', 'e', 'E']:
                        f_error_log = open(file_error_log, 'a')
                        flag_error = 1
                        f_error_log.write(f+'\t'+str(idx1+1)+'\t'+'All rows of file step 1 of Wismettac(Nishimotto) and JFC should be filled with C or E in case_yn\n\n')
                        f_error_log.close()
                        break
                    elif tmp_vc not in ['0202', '0203'] and tmp_c_or_e in ['c', 'C', 'e', 'E']:
                        f_error_log = open(file_error_log, 'a')
                        flag_error = 1
                        f_error_log.write(f+'\t'+str(idx1+1)+'\t'+'If the vendor is neither Wismettac(Nishimotto) or JFC, the column case_yn should be empty\n\n')
                        f_error_log.close()
                        break
                    tmp_vat = getVAT(tmp_vc, tmp_sc, '')
                    valid_cost = float(tmp_ncost)
                    valid_cost = float(tmp_amt)
                    valid_qty = float(tmp_qty)
                    if tmp_upc != '':
                        valid_upc = int(tmp_upc)
                    if (tmp_upc is None or tmp_upc == '') and (tmp_ic is None or tmp_ic == ''):
                        f_error_log = open(file_error_log, 'a')
                        flag_error = 1
                        f_error_log.write(f+'\t'+str(idx1+1)+'\t'+'item code and UPC no exist error\n\n')
                        f_error_log.close()
                        break
                    valid_date = tmp_date.split('/')
                    if int(valid_date[0]) < 1 or int(valid_date[0]) > 12 or int(valid_date[1]) < 1 or int(valid_date[1]) > 31 or len(valid_date[2]) != 4:
                        f_error_log = open(file_error_log, 'a')
                        flag_error = 1
                        f_error_log.write(f+'\t'+str(idx1+1)+'\t'+'date format error in '+str(tmp_date)+'\n\n')
                        f_error_log.close()
                        break
                    if tmp_sc is None or len(tmp_sc) != 3 or int(tmp_sc) > 13 or int(tmp_sc) < 1:
                        f_error_log = open(file_error_log, 'a')
                        flag_error = 1
                        f_error_log.write(f+'\t'+str(idx1+1)+'\t'+'store code format error in '+str(tmp_sc)+'\n\n')
                        f_error_log.close()
                        break
                    if tmp_vc is None or len(tmp_vc) != 4:
                        f_error_log = open(file_error_log, 'a')
                        flag_error = 1
                        f_error_log.write(f+'\t'+str(idx1+1)+'\t'+'vendor code format error in '+str(tmp_vc)+'\n\n')
                        f_error_log.close()
                        break

                    param_etc['file_nm'] = f
                    param_etc['department'] = ' '
                    result_upc, result_ic, result_desc, result_csize, _, _, result_flag_processed_ce, result_upc_check, \
                        result_upc_same_check, result_desc_lookup_base, result_old_case_cost, result_new_case_cost, \
                        result_price, result_old_margin, result_new_margin, result_srp, tmp_upc_cand2, detail_cat, result_note = \
                        data_lookup(g_data, tmp_upc, tmp_sc, tmp_vc, tmp_ic, tmp_csize, tmp_ncost, tmp_amt,
                                    tmp_upc_cand2, tmp_vat, tmp_c_or_e, 0, dict_error_flag, dict_vendor, dict_vend,
                                    param_etc)

                    # arr1 = [PK, 파일명,
                    # 0       F01, UPC check, lookup UPC, UPC 동일여부, 원본 UPC, 후보 UPC,
                    # 1~3     F902, F1000, F27
                    # 4       F1184(CASE)
                    # 5       F26, Description Lookup 기준
                    # 6       Description
                    # 7       F1001
                    # 8       F19
                    # 9~10    F38, F39
                    # 11      F90(1)
                    # 12, 13  QTY, Amount
                    # 14      F1122]
                    count_step_one += 1
                    arr1 = [primary_key, file_name]
                    primary_key += 1
                    idx2 = 0
                    if len(row) != 16 and row[15]:
                        f_error_log = open(file_error_log, 'a')
                        flag_error = 1
                        f_error_log.write(f+'\t'+str(idx1+1)+'\t'+'format column error on '+f+'\n')
                        f_error_log.write('it seems that there are '+str(len(row))+' columns\n')
                        f_error_log.write('check there are 16 columns on that file\n')
                        for test_idx in range(len(row) - 16):
                            if row[test_idx + 16].value is not None and row[test_idx + 16].value != 'None' and row[test_idx + 16].value != '':
                                f_error_log.write('column '+str(test_idx+17)+' caused error including '+str(row[test_idx+16].value)+'\n')
                        f_error_log.write('\n')
                        f_error_log.close()
                        break
                    for col in row:
                        if idx2 == 0:
                            if result_desc_lookup_base == 'N':
                                tmp_upc_3, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)
                                arr1.append(tmp_upc_3)
                            else:
                                arr1.append(result_upc)
                            arr1.append(result_upc_check)
                            arr1.append(result_upc)
                            tmp_upc_3, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)
                            if tmp_upc_3 == result_upc or result_upc in tmp_upc_cand:
                                arr1.append(1)
                            else:
                                arr1.append(0)
                            arr1.append(tmp_upc)
                            arr1.append(str(tmp_upc_cand2))
                        elif idx2 == 4:
                            arr1.append('CASE')
                        elif idx2 == 5:
                            ic_cand = getFormalItemCode(tmp_ic, tmp_vc, dict_vend)
                            ic_cand2 = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                            if len(ic_cand2) == 0:
                                ic_cand2.append(tmp_ic)
                            if tmp_ic == '':
                                arr1.append(ic_cand2[0])
                            else:
                                arr1.append(ic_cand[0])
                            arr1.append(ic_cand2[0])
                            arr1.append(tmp_ic)
                            if tmp_ic != result_ic and result_ic not in ic_cand:
                                if tmp_ic != '':
                                    arr1.append(1)
                                else:
                                    arr1.append(0)
                            else:
                                arr1.append(0)
                            arr1.append(result_desc_lookup_base)
                        elif idx2 == 6:
                            if result_desc_lookup_base != 'N':
                                arr1.append(result_desc)
                            else:
                                arr1.append(tmp_desc)
                        elif idx2 == 7:
                            arr1.append('1')
                        elif idx2 == 8:
                            if result_desc_lookup_base != 'N':
                                arr1.append(result_csize)
                            else:
                                arr1.append(tmp_csize)
                        elif idx2 == 10:
                            token_date = col.value.split('/')
                            arr1.append(str(int(token_date[0]))+'/'+str(int(token_date[1]))+'/'+token_date[2])
                        elif idx2 == 11:
                            arr1.append('1')
                        elif idx2 == 12 or idx2 == 13:
                            if col.value is None:
                                arr1.append(0.0)
                            else:
                                arr1.append(float(str(col.value).replace(',', '')))
                            if idx2 == 13:
                                arr1.append(tmp_vat)
                        elif idx2 != 14:
                            arr1.append(col.value)
                        idx2 += 1
                    # print('#2-3')
                    u_vc = tmp_vc
                    if u_vc is None:
                        u_vc = ''
                    u_ic = tmp_ic
                    if u_ic is None:
                        u_ic = ''
                    u_upc = tmp_upc
                    if u_upc is None:
                        u_upc = ''
                    u_tmp_key = u_vc + u_ic + u_upc
                    if (u_ic is None or u_ic == '') and (u_upc is None or u_upc == ''):
                        u_tmp_key = 'nokey_' + str(idx1)
                    if u_tmp_key in dict_unique:
                        # if int(arr1[21]) == 0:
                        # QTY
                        dict_unique[u_tmp_key][23] = dict_unique[u_tmp_key][23] + arr1[23]
                        # Amount
                        dict_unique[u_tmp_key][24] = dict_unique[u_tmp_key][24] + arr1[24]
                        dict_unique[u_tmp_key][26] += 1
                        '''
                        else:
                            idx_u_key = 0
                            u_tmp_key2 = u_tmp_key+'_'+str(idx_u_key)
                            while u_tmp_key2 in dict_unique:
                                idx_u_key += 1
                                u_tmp_key2 = u_tmp_key+'_'+str(idx_u_key)
                            dict_unique[u_tmp_key2] = arr1
                            dict_unique[u_tmp_key2].append(2)
                            dict_unique[u_tmp_key2].append(result_old_case_cost)
                            dict_unique[u_tmp_key2].append(result_new_case_cost)
                            dict_unique[u_tmp_key2].append(result_price)
                            dict_unique[u_tmp_key2].append(result_old_margin)
                            dict_unique[u_tmp_key2].append(result_new_margin)
                            dict_unique[u_tmp_key2].append(result_srp)                        '''

                    else:
                        dict_unique[u_tmp_key] = arr1[:26]
                        dict_unique[u_tmp_key].append(0)
                        dict_unique[u_tmp_key].append(result_old_case_cost)
                        dict_unique[u_tmp_key].append(result_new_case_cost)
                        dict_unique[u_tmp_key].append(result_price)
                        dict_unique[u_tmp_key].append(result_old_margin)
                        dict_unique[u_tmp_key].append(result_new_margin)
                        dict_unique[u_tmp_key].append(result_srp)
                        dict_unique[u_tmp_key].append(detail_cat)
                        dict_unique[u_tmp_key].append(tmp_c_or_e)
                        if tmp_c_or_e in ['e', 'E']:
                            dict_unique[u_tmp_key].append(1)
                        else:
                            dict_unique[u_tmp_key].append(0)

                # print('#3')
                dict_to_file(dict_unique, dict_vendor, ws_n)
                dict_unique = dict()
                wb_t.close()

                # print('#4')

            except Exception as ex:
                f_error_log = open(file_error_log, 'a')
                flag_error = 1
                f_error_log.write(f+'\t'+str(idx1+1)+'\t'+str(traceback.format_exc())+'\n')
                f_error_log.close()
            print('finished! total line: '+str(idx1))

    print('saving result of step 2...')
    if not isdir('./result_step_2'):
        mkdir('./result_step_2')
    wb_n.save('./result_step_2/'+output_file_nm+'.xlsx')
    print('saving count file of step 2...')
    count_record = open('./result_step_2/'+output_file_nm+'.txt', 'w')
    count_record.write(str(count_step_one))
    count_record.close()

    if flag_error == 0:
        remove(file_error_log)


def process_step_four(thread_id, g_data, param1):
    len_last = 0
    len_tmp = 0

    last_ic = ''
    last_vc = ''

    flag_error = 0

    count_step_one = 0

    file_dir_target = './input_step_4'
    if not isdir(file_dir_target):
        mkdir(file_dir_target)
    if not isdir(file_dir_target+'/'+param1):
        mkdir(file_dir_target+'/'+param1)
    file_error_log = './error_step_4/'+param1+'.txt'
    if not isdir('./error_step_4'):
        mkdir('./error_step_4')
    output_file_nm = param1

    f_error_log = open(file_error_log, 'a')
    f_error_log.write('\nSTART\n')
    f_error_log.close()

    param_etc = {}
    param_etc['tag_is'] = str(time.time())
    param_etc['step'] = 4

    dict_vendor = g_data['dict_vendor']
    dict_vend = g_data['dict_vend']
    dict_store = g_data['dict_store']

    list_item = []

    dict_srp_store_date = dict()
    header = []

    file_dir_target = './input_step_4/'+param1
    idx0 = -1
    for f in listdir(file_dir_target):
        idx0 += 1
        file_targ = join(file_dir_target, f)
        if isfile(file_targ):
            wb_t = load_workbook(file_targ)
            ws_t = wb_t['Sheet1']
            idx1 = -1
            dict_item = {}
            for row in ws_t.rows:
                idx1 += 1
                # print(idx1, sep=' ')
                if idx1 == 0:
                    if idx0 == 0:
                        for col in row:
                            header.append(col.value)
                    continue
                arr1 = []
                for col in row:
                    arr1.append(col.value)

                tmp_upc = str(row[10].value)
                if row[10].value is None:
                    tmp_upc = ''
                tmp_sc = str(row[13].value)
                if row[13].value is None:
                    tmp_sc = ''
                tmp_vc = str(row[14].value)
                if row[14].value is None:
                    tmp_vc = ''
                tmp_ic = str(row[18].value)
                if row[18].value is None:
                    tmp_ic = ''
                tmp_csize = row[23].value if tmp_vc not in ['1210', '1205'] else 1
                tmp_qty = row[27].value
                tmp_ncost = row[24].value
                tmp_amt = row[28].value
                tmp_date = row[25].value
                tmp_upc_cand2 = []
                tmp_srp = row[36].value
                tmp_c_or_e = row[37].value
                tmp_depart = row[39].value
                tmp_vat = getVAT(tmp_vc, tmp_sc, tmp_depart)
                if tmp_c_or_e is None:
                    tmp_c_or_e = ''
                tmp_flag_processed_ce = row[38].value
                valid_cost = float(tmp_ncost)
                valid_cost = float(tmp_amt)
                valid_qty = float(tmp_qty)
                if tmp_upc != '':
                    valid_upc = int(tmp_upc)
                if (tmp_upc is None or tmp_upc == '') and (tmp_ic is None or tmp_ic == ''):
                    f_error_log = open(file_error_log, 'a')
                    flag_error = 1
                    f_error_log.write(f+'\t'+str(idx1+1)+'\t'+'item code and UPC no exist error\n\n')
                    f_error_log.close()
                    break
                valid_date = tmp_date.split('/')
                if int(valid_date[0]) < 1 or int(valid_date[0]) > 12 or int(valid_date[1]) < 1 or int(valid_date[1]) > 31 or len(valid_date[2]) != 4:
                    f_error_log = open(file_error_log, 'a')
                    flag_error = 1
                    f_error_log.write(f+'\t'+str(idx1+1)+'\t'+'date format error in '+str(tmp_date)+'\n\n')
                    f_error_log.close()
                    break
                if tmp_sc is None or len(tmp_sc) != 3 or int(tmp_sc) > 13 or int(tmp_sc) < 1:
                    f_error_log = open(file_error_log, 'a')
                    flag_error = 1
                    f_error_log.write(f+'\t'+str(idx1+1)+'\t'+'store code format error in '+str(tmp_sc)+'\n\n')
                    f_error_log.close()
                    break
                if tmp_vc is None or len(tmp_vc) != 4:
                    f_error_log = open(file_error_log, 'a')
                    flag_error = 1
                    f_error_log.write(f+'\t'+str(idx1+1)+'\t'+'vendor code format error in '+str(tmp_vc)+'\n\n')
                    f_error_log.close()
                    break

                # 처리 내용이 있으면 다시 lookup
                # if arr1[5] is not None and arr1[5] != '':
                try:
                    dict_error_flag = dict()
                    param_etc['file_nm'] = arr1[1]
                    result_upc, result_ic, result_desc, result_csize, result_ncost, result_amt, result_flag_processed_ce, result_upc_check, result_upc_same_check, result_desc_lookup_base, result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp, tmp_upc_cand2, detail_cat, result_note = \
                        data_lookup(g_data, tmp_upc, tmp_sc, tmp_vc, tmp_ic, tmp_csize, tmp_ncost, tmp_amt, tmp_upc_cand2, tmp_vat, tmp_c_or_e, tmp_flag_processed_ce, dict_error_flag, dict_vendor, dict_vend, param_etc)
                    if result_desc_lookup_base == 'N':
                        tmp_upc_3, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)
                        arr1[6] = tmp_upc_3
                    else:
                        arr1[6] = result_upc
                    arr1[7] = result_upc_check
                    arr1[8] = result_upc
                    tmp_upc_3, tmp_upc_cand = getFormalUPC(tmp_upc, tmp_vc)
                    if tmp_upc_3 == result_upc or result_upc in tmp_upc_cand:
                        result_upc_same_check = 1
                        arr1[9] = 1
                    else:
                        result_upc_same_check = 0
                        arr1[9] = 0
                    arr1[10] = tmp_upc
                    arr1[11] = str(tmp_upc_cand2)
                    arr1[13] = tmp_sc
                    arr1[14] = tmp_vc
                    ic_cand = getFormalItemCode(tmp_ic, tmp_vc, dict_vend)
                    ic_cand2 = getFormalItemCode(result_ic, tmp_vc, dict_vend)
                    if len(ic_cand2) == 0:
                        ic_cand2.append(tmp_ic)
                    if tmp_ic == '':
                        arr1[16] = ic_cand2[0]
                    else:
                        arr1[16] = ic_cand[0]
                    arr1[17] = ic_cand2[0]
                    arr1[18] = tmp_ic
                    if tmp_ic != result_ic and result_ic not in ic_cand:
                        if tmp_ic != '':
                            arr1[19] = 1
                        else:
                            arr1[19] = 0
                    else:
                        arr1[19] = 0
                    arr1[24] = result_ncost
                    arr1[28] = result_amt
                    arr1[29] = tmp_vat

                    arr1[20] = result_desc_lookup_base
                    if result_desc is not None and result_desc != '':
                        arr1[21] = result_desc
                    if result_csize is not None and result_csize != '':
                        arr1[23] = result_csize
                    if len(arr1) < 40:
                        for idx3 in range(40-len(arr1)):
                            arr1.append(' ')
                    arr1[31] = result_old_case_cost
                    arr1[32] = result_new_case_cost
                    arr1[33] = result_price
                    arr1[34] = result_old_margin
                    arr1[35] = result_new_margin
                    arr1[36] = result_srp
                    arr1[38] = result_flag_processed_ce

                    if result_desc_lookup_base == 'N':
                        if int(result_upc_check) > 1 and (tmp_upc is None or tmp_upc == ''):
                            arr1[2] = '업체+제품 검색 시 다수이나 인보이스 내 UPC 없음'
                            arr1[3] = detail_cat
                            arr1[4] = 'Description 및 F19 입력'
                        elif (tmp_ic is None or tmp_ic == '') and (tmp_upc is None or tmp_upc == ''):
                            arr1[2] = '인보이스에 제품코드와 UPC 둘 다 없음'
                            arr1[3] = detail_cat
                            arr1[4] = '제품 코드, UPC 없는지 확인 후 Description 및 F19 입력'
                        elif tmp_vc not in dict_vendor:
                            arr1[2] = '매장 데이터에 등록되지 않은 업체 코드'
                            arr1[3] = detail_cat
                            arr1[4] = '업체 코드 재확인 후 Description 및 F19 입력'
                        elif int(result_upc_check) == 0:
                            arr1[2] = '신상품'
                            arr1[3] = detail_cat
                            arr1[4] = '제품 코드와 UPC 및 업체 코드 재확인 후 Description 및 F19 입력'
                        else:
                            arr1[2] = '업체+제품 검색 시 다수이나 인보이스의 UPC와 상이함'
                            arr1[3] = detail_cat
                            arr1[4] = 'UPC 재확인 후 Description 및 F19 입력'
                    else:
                        if result_desc_lookup_base == 'I' and result_upc_same_check == 0 and tmp_upc is not None and tmp_upc != '':
                            arr1[2] = '업체+제품으로 검색된 UPC가 단일이지만 인보이스의 UPC와 상이'
                            arr1[3] = detail_cat
                            arr1[4] = 'UPC 재확인'
                        elif result_desc_lookup_base == 'U' and arr1[19] == 1:
                            arr1[2] = 'UPC로 검색된 제품코드가 인보이스의 제품코드와 상이'
                            arr1[3] = detail_cat
                            arr1[4] = '제품코드 재확인'
                        else:
                            arr1[2] = ''
                            arr1[3] = detail_cat
                            arr1[4] = ''
                    if str(tmp_sc)+'_'+str(tmp_date) not in dict_srp_store_date:
                        dict_srp_store_date[str(tmp_sc)+'_'+str(tmp_date)] = []
                    # 유니크 키를 만들어 중복되는 제품을 가려냄
                    key_item = str(arr1[1])+str(arr1[13])+'_+_'+str(tmp_upc)+'_+_'+str(tmp_ic)+'_+_'+str(tmp_date)
                    # 중복되는 제품이 있는 경우
                    if key_item in dict_item:
                        dict_item[key_item]['qty'] += float(tmp_qty)
                        dict_item[key_item]['amount'] += float(result_amt)
                        list_item[dict_item[key_item]['idx']][30] += 10
                        list_item[dict_item[key_item]['idx']][27] = dict_item[key_item]['qty']
                        list_item[dict_item[key_item]['idx']][28] = dict_item[key_item]['amount']
                        list_item[dict_item[key_item]['idx']][24] = dict_item[key_item]['amount'] / dict_item[key_item]['qty']
                        result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp \
                            = getSrpSet(list_item[dict_item[key_item]['idx']][23], list_item[dict_item[key_item]['idx']][31], list_item[dict_item[key_item]['idx']][24], result_price, list_item[dict_item[key_item]['idx']][29])
                        if flag_ignore_srp_sticker and result_note == 'STICKER':
                            print(tmp_ic + "\t" + tmp_upc)
                            result_srp = ' '
                        list_item[dict_item[key_item]['idx']][31] = result_old_case_cost
                        list_item[dict_item[key_item]['idx']][32] = result_new_case_cost
                        list_item[dict_item[key_item]['idx']][33] = result_price
                        list_item[dict_item[key_item]['idx']][34] = result_old_margin
                        list_item[dict_item[key_item]['idx']][35] = result_new_margin
                        list_item[dict_item[key_item]['idx']][36] = result_srp

                        flag_only_prefix = False
                        if tmp_vc in dict_vend and dict_vend[tmp_vc]['prefix'] is not None and dict_vend[tmp_vc]['prefix'] != 'None' and dict_vend[tmp_vc]['prefix'] != '' and result_ic == dict_vend[tmp_vc]['prefix']:
                            flag_only_prefix = True

                        if not flag_only_prefix:
                            # 계산된 srp가 있는 경우
                            if result_srp is not None and result_srp != '' and result_srp != ' ':
                                # 기존에 srp가 없었으면 새로 추가
                                if dict_item[key_item]['key_srp_2'] == -1:
                                    dict_item[key_item]['key_srp_2'] = len(dict_srp_store_date[str(tmp_sc)+'_'+str(tmp_date)])
                                    dict_srp_store_date[str(tmp_sc)+'_'+str(tmp_date)].append([result_upc, '', tmp_sc, tmp_vc, result_ic, 'CASE', '1', '1', result_csize, result_new_case_cost, tmp_date, '1', tmp_vat, result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp, True])
                                # 기존에 srp가 있었으면 새로운 srp로 덮어쓰기
                                else:
                                    dict_srp_store_date[str(tmp_sc)+'_'+str(tmp_date)][dict_item[key_item]['key_srp_2']] = [result_upc, '', tmp_sc, tmp_vc, result_ic, 'CASE', '1', '1', result_csize, result_new_case_cost, tmp_date, '1', tmp_vat, result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp, True]
                            # 계산된 srp가 없는 경우 non은 추가 나머지 세가지는 모두 제거
                            else:
                                if dict_item[key_item]['key_srp_2'] > -1:
                                    dict_srp_store_date[str(tmp_sc)+'_'+str(tmp_date)][dict_item[key_item]['key_srp_2']][-1] = False
                    # 중복되는 제품이 없는 경우 dict에 넣어 줌
                    else:
                        dict_item[key_item] = {
                            'idx': len(list_item)
                            , 'key_srp_1': str(tmp_sc)+'_'+str(tmp_date)
                            , 'key_srp_2': len(dict_srp_store_date[str(tmp_sc)+'_'+str(tmp_date)]) if result_srp and result_srp != '' and result_srp != ' ' else -1
                            , 'file_name': arr1[1]
                            , 'qty': float(tmp_qty)
                            , 'amount': float(result_amt)
                        }
                        list_item.append(arr1)

                        flag_only_prefix = False
                        if tmp_vc in dict_vend and dict_vend[tmp_vc]['prefix'] is not None and dict_vend[tmp_vc]['prefix'] != 'None' and dict_vend[tmp_vc]['prefix'] != '' and result_ic == dict_vend[tmp_vc]['prefix']:
                            flag_only_prefix = True

                        if result_srp is not None and result_srp != '' and result_srp != ' ' and not flag_only_prefix:
                            dict_srp_store_date[str(tmp_sc)+'_'+str(tmp_date)].append([result_upc, '', tmp_sc, tmp_vc, result_ic, 'CASE', '1', '1', result_csize, result_new_case_cost, tmp_date, '1', tmp_vat, result_old_case_cost, result_new_case_cost, result_price, result_old_margin, result_new_margin, result_srp, True])

                except Exception as ex:
                    f_error_log = open(file_error_log, 'a')
                    flag_error = 1
                    f_error_log.write(str(arr1)+'\n')
                    f_error_log.write(f + '\t' + str(idx1 + 1) + '\t' + str(traceback.format_exc()) + '\n')
                    f_error_log.close()

    wb_n = Workbook()
    ws_n = wb_n.active
    ws_n.title = 'Sheet1'
    ws_n.append(header)
    wb_chk = Workbook()
    ws_chk = wb_chk.active
    ws_chk.title = 'Sheet1'
    ws_chk.append(['#']+split_for_chk_meta(header))
    for idx1 in range(len(list_item)):
        ws_n.append(list_item[idx1])
        if float(list_item[idx1][27]) != 0.0 and float(list_item[idx1][24]) != 0.0:
            cost_val = float(list_item[idx1][28])/float(list_item[idx1][27])/float(list_item[idx1][24])
            if cost_val < 0.999 or cost_val >= 1.001:
                ws_chk.append([str(idx1)] + split_for_chk_meta(list_item[idx1]))
                continue
        if list_item[idx1][31] != '' and float(list_item[idx1][31]) != 0.0:
            case_cost_val = list_item[idx1][32]/list_item[idx1][31]
            if case_cost_val <= 0.5 or case_cost_val >= 2.0:
                ws_chk.append([str(idx1)] + list_item[idx1])
                continue
        if list_item[idx1][31] != '' and list_item[idx1][35] != '':
            margin_val = float(list_item[idx1][35])
            if margin_val < 0.1:
                ws_chk.append([str(idx1)] + list_item[idx1])
                continue

    if not isdir('./result_step_4/'):
        mkdir('./result_step_4/')
    wb_n.save('./result_step_4/'+output_file_nm+'.xlsx')
    wb_chk.save('./result_step_4/CHK_FILE_'+output_file_nm+'.xlsx')

    if not isdir('./srp_result/'):
        mkdir('./srp_result/')

    if not isdir('./srp_result/'+param1):
        mkdir('./srp_result/'+param1)
    elif len(listdir('./srp_result/'+param1)):
        for file_to_rm in listdir('./srp_result/'+param1):
            remove(join('./srp_result/'+param1, file_to_rm))

    for idx_s in dict_srp_store_date:
        tmp_sc, tmp_date = idx_s.split('_')
        tmp_date = tmp_date.split('/')
        for idx_d in range(2):
            if len(tmp_date[idx_d]) == 1:
                tmp_date[idx_d] = '0'+tmp_date[idx_d]
        tmp_date[2] = tmp_date[2][-2:]
        file_nm = tmp_date[0]+tmp_date[1]+tmp_date[2]+'_'+dict_store[tmp_sc]['code']+'_PB.xlsx'

        wb_n = Workbook()
        ws_n = wb_n.active
        ws_n.title = 'Sheet1'
        ws_n.append(['F01', 'F902', 'F1000', 'F27', 'F26', 'F1184', 'F126', 'F1001', 'F19', 'F38', 'F39', 'F90', 'F1122', 'Old Case Cost', 'New Case Cost', 'Active Price', 'Old Margin', 'New Margin', 'F30'])
        cnt_true = 0
        for item in dict_srp_store_date[idx_s]:
            if check_except_pb(item) and item[-2] and item[-1]:
                ws_n.append(item[:-1])
                cnt_true += 1

        if cnt_true > 0:
            wb_n.save('./srp_result/'+param1+'/'+file_nm)
        wb_n.close()

    if not isdir('./srp_analysis/'):
        mkdir('./srp_analysis/')

    if flag_error == 0:
        remove(file_error_log)


def split_for_chk_meta(targ_list):
    return [targ_list[1]]+targ_list[4:6]+targ_list[10:12]+targ_list[14:16]+targ_list[18:20]+targ_list[21:23]+targ_list[24:27]+targ_list[28:30]+[targ_list[32]]+[targ_list[36]]+targ_list[38:40]


def lookup_vendor_ic(g_data, org_csize, vc, sc, ic_cand, upc):
    targ_file = ''
    if vc and sc and vc in g_data['dict_vendor_item_code']:

        tmp_upc, tmp_upc_cand = getFormalUPC(upc, vc)

        refer_idata = g_data['dict_vendor_item_code'][vc][sc]['IC']
        idx_ic = -1
        idx_tmp = 0
        for item in ic_cand:
            if item in refer_idata:
                idx_ic = idx_tmp
                break
            idx_tmp += 1
        if idx_ic > -1:
            if tmp_upc in refer_idata[ic_cand[idx_ic]]:
                return float(refer_idata[ic_cand[idx_ic]][tmp_upc]['csize'])
            else:
                idx_upc = -1
                idx_tmp = 0
                for item in tmp_upc_cand:
                    if item in refer_idata[ic_cand[idx_ic]]:
                        idx_upc = idx_tmp
                        break
                    idx_tmp += 1
                if idx_upc > -1:
                    return float(refer_idata[ic_cand[idx_ic]][tmp_upc_cand[idx_upc]]['csize'])

        refer_udata = g_data['dict_vendor_item_code'][vc][sc]['UPC']
        if tmp_upc in refer_udata:
            idx_ic = -1
            idx_tmp = 0
            for item in ic_cand:
                if item in refer_udata[tmp_upc]:
                    idx_ic = idx_tmp
                    break
                idx_tmp += 1
            if idx_ic > -1:
                return float(refer_udata[tmp_upc][ic_cand[idx_ic]]['csize'])
        idx_upc = -1
        idx_tmp = 0
        for item in tmp_upc_cand:
            if item in refer_udata:
                idx_upc = idx_tmp
                break
            idx_tmp += 1
        if idx_upc > -1:
            idx_ic = -1
            idx_tmp = 0
            for item in ic_cand:
                if item in refer_udata[tmp_upc_cand[idx_upc]]:
                    idx_ic = idx_tmp
                    break
                idx_tmp += 1
            if idx_ic > -1:
                return float(refer_udata[tmp_upc_cand[idx_upc]][ic_cand[idx_ic]]['csize'])

    if org_csize is not None and org_csize != '':
        org_csize = float(org_csize)

    return org_csize


def get_converted_price(prc):
    if prc <= 0.0:
        return 0.0
    int_prc = int(prc)
    und_float = prc - int_prc
    result = 0.0 + int_prc
    if und_float == 0.00:
        result -= 1
        result += 0.99
    elif und_float < 0.31:
        result += 0.29
    elif und_float < 0.51:
        result += 0.49
    elif und_float < 0.81:
        result += 0.79
    elif und_float < 1.00:
        result += 0.99
    return result


def get_margin(price, case_price, case_size, vat):
    new_price = (float(case_price) / float(case_size)) * (1.0 + (float(vat)/100))
    diff_price = float(price) - new_price
    return diff_price / price
