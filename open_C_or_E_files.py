from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import datetime
import os


STORE_LIST = {'001' : 'MD300', '002': 'VA111', '003': 'MD91', '004': 'VA77', '005': 'MD100', '006': 'MD600', '007': 'VA900',
     '008': 'VA500', '009': 'MD700', '010': 'MD21', '011': 'FL01', '012': 'VA05', '013': 'NJ01', '014': 'VA06', '015': 'FL02'}

# now = datetime.datetime.now() - datetime.timedelta(days=2)
now = datetime.datetime.now() - datetime.timedelta(days=1)
now = datetime.datetime.now()
print(now)
now_yymmdd= now.strftime("%y%m%d")
now_mmdd= now.strftime("%m%d")

BASE_DIR = "C:/Users/user/Documents/GitHub/lotteplaza/작업파일"
todayFilePath = os.path.join(BASE_DIR, now_mmdd, "4차", f"TOT_{now_yymmdd}_01.xlsx")
print(todayFilePath)
print()

def column_index(letter):
    return column_index_from_string(letter) - 1

def open_file(path):
    try:
        os.system(path)
    except Exception as e:
        print(f"오류발생: {e}")

def add_file_list(file_path_list, row):
    b_column_index = column_index_from_string('B') - 1
    n_column_index = column_index_from_string('N') - 1
    b_value = row[b_column_index].value
    n_value = row[n_column_index].value
    store = STORE_LIST[n_value]
    file_path = os.path.join(store,b_value)
    if file_path not in file_path_list:
        file_path_list.append(file_path)
        print(file_path)
    return file_path_list

def get_data_range_workbook(filepath):
    workbook = load_workbook(filename=filepath)
    sheet = workbook.active
    # 데이터가 저장된 행의 수 계산
    total_rows = sheet.max_row
    data_range = sheet['A1':'AN' + str(total_rows)]
    return data_range

def open_xlsxs(file_list):
    for item in file_list:
        desktop_path = r"C:\Users\user\Desktop"
        file_path = os.path.join(desktop_path, now_mmdd, item)
        open_file(file_path)

def open_files_based_on_criteria(filepath, filter_columns, filter_values):
    data_range = get_data_range_workbook(filepath)
    open_file_path_list = []
    firstRow = True
    for row in data_range:
        if firstRow:
            firstRow = False
        elif all(row[column_index(letter)].value in values for letter, values in zip(filter_columns, filter_values)):
            open_file_path_list = add_file_list(open_file_path_list, row)
    open_file_path_list.reverse()
    open_xlsxs(open_file_path_list)

def openFileNameForCorE(filepath):
    filter_columns = ['AL', 'AM']
    filter_values = [['C', 'E'], [1, '1']]
    open_files_based_on_criteria(filepath, filter_columns, filter_values)

def openFileNameForDupFlag(filepath):
    filter_columns = ['AL', 'AE']
    filter_values = [['C', 'E'], ['10', 10, '20', 20, '30', 30, '40', 40, '50', 50, '60', 60]] ## dupFlag는 10~60 까지만 있다고 가정
    open_files_based_on_criteria(filepath, filter_columns, filter_values)

def openFileNameForSupervalu(filepath):
    filter_columns = ['O', 'N']
    filter_values = [['1229', 1229], ['011', '015']]
    open_files_based_on_criteria(filepath, filter_columns, filter_values)

print("DupFlag")
# file open
openFileNameForDupFlag(todayFilePath)
print('----------------------------------------------')
print("C or E")
# file open
openFileNameForCorE(todayFilePath)
# print("SuperValu")
# # file open
# openFileNameForSupervalu(todayFilePath)


