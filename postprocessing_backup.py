from openpyxl import load_workbook, Workbook
from os import listdir, mkdir, remove
from os.path import isfile, join, isdir
import dataEntryFunction as df

def build_cb_item(row):
    dict_item = dict()
    file_nm = row[0].value
    dict_item['file_nm'] = file_nm
    cat = row[1].value
    if cat is None:
        cat = ''
    dict_item['cat'] = cat
    sub_cat = row[2].value
    if sub_cat is None:
        sub_cat = ''
    dict_item['sub_cat'] = sub_cat
    f01 = row[3].value
    if f01 is None:
        f01 = ''
    dict_item['f01'] = f01
    f902 = ''
    dict_item['f902'] = f902
    f1000 = row[10].value
    dict_item['f1000'] = f1000
    f27 = row[11].value
    dict_item['f27'] = f27
    f1184 = row[12].value
    dict_item['f1184'] = f1184
    f26 = row[13].value
    if f26 is None:
        f26 = ''
    dict_item['f26'] = f26
    lookup_f26 = row[14].value
    if lookup_f26 is None:
        lookup_f26 = ''
    dict_item['lookup_f26'] = lookup_f26
    desc = row[18].value
    if desc is None:
        desc = ''
    dict_item['desc'] = desc
    f1001 = '1'
    dict_item['f1001'] = f1001
    f19 = row[20].value
    dict_item['f19'] = f19
    f38 = row[21].value
    dict_item['f38'] = f38
    f39 = row[22].value
    dict_item['f39'] = f39
    f90 = '1'
    dict_item['f90'] = f90
    qty = row[24].value
    dict_item['qty'] = qty
    amt = row[25].value
    dict_item['amt'] = amt
    f1122 = row[26].value
    dict_item['f1122'] = f1122
    old_ccost = row[27].value
    if old_ccost is None:
        old_ccost = ''
    dict_item['old_ccost'] = old_ccost
    new_ccost = row[28].value
    if new_ccost is None:
        new_ccost = ''
    dict_item['new_ccost'] = new_ccost
    price = row[29].value
    if price is None:
        price = ''
    dict_item['price'] = price
    old_margin = row[30].value
    if old_margin is None:
        old_margin = ''
    dict_item['old_margin'] = old_margin
    new_margin = row[31].value
    if new_margin is None:
        new_margin = ''
    dict_item['new_margin'] = new_margin
    srp = row[32].value
    if srp is None:
        srp = ''
    dict_item['srp'] = srp

    return dict_item


def build_pb_item(row):
    dict_item = dict()
    f01 = row[0].value
    if f01 is None:
        f01 = ''
    dict_item['f01'] = f01
    f902 = ''
    dict_item['f902'] = f902
    f1000 = row[2].value
    dict_item['f1000'] = f1000
    f27 = row[3].value
    dict_item['f27'] = f27
    f26 = row[4].value
    if f26 is None:
        f26 = ''
    dict_item['f26'] = f26
    f1184 = row[5].value
    dict_item['f1184'] = f1184
    f126 = '1'
    dict_item['f126'] = f126
    f1001 = '1'
    dict_item['f1001'] = f1001
    f19 = row[8].value
    dict_item['f19'] = f19
    f38 = row[9].value
    dict_item['f38'] = f38
    f39 = row[10].value
    dict_item['f39'] = f39
    f90 = '1'
    dict_item['f90'] = f90
    f1122 = row[12].value
    dict_item['f1122'] = f1122
    old_ccost = row[13].value
    if old_ccost is None:
        old_ccost = ''
    dict_item['old_ccost'] = old_ccost
    new_ccost = row[14].value
    if new_ccost is None:
        new_ccost = ''
    dict_item['new_ccost'] = new_ccost
    price = row[15].value
    if price is None:
        price = ''
    dict_item['price'] = price
    old_margin = row[16].value
    if old_margin is None:
        old_margin = ''
    dict_item['old_margin'] = old_margin
    new_margin = row[17].value
    if new_margin is None:
        new_margin = ''
    dict_item['new_margin'] = new_margin
    srp = row[18].value
    if srp is None:
        srp = ''
    dict_item['srp'] = srp

    return dict_item


def str_margin_to_percent(str_margin):
    if str_margin == ' ' or str_margin == '':
        return str_margin
    else:
        return float(str_margin)


def get_header_from_row(header_row):
    arr = []
    for cell in header_row:
        arr.append(cell.value)
    return arr


def is_clover_one_gallon(vc, upc, ic):
    if vc == '1205' and (upc in ['0004938710011', '0004938710012', '0004938710013', '0004938710015'] or ic in ['CLO1051', 'CLO1307', 'CLO1425', 'CLO1593']):
        return True
    else:
        return False


def get_converted_price(prc):
    if prc <= 0.0:
        return 0.0
    int_prc = int(prc)
    und_float = prc - int_prc
    result = 0.0 + int_prc
    if und_float == 0.00:
        result -= 1
        result += 0.99
    elif und_float < 0.31:
        result += 0.29
    elif und_float < 0.51:
        result += 0.49
    elif und_float < 0.81:
        result += 0.79
    elif und_float < 1.00:
        result += 0.99
    return result


def get_margin(price, case_price, case_size, vat):
    new_price = (float(case_price) / float(case_size)) * (1.0 + (float(vat)/100))
    diff_price = float(price) - new_price
    return diff_price / price


dict_store = {
    '001': 'MD300',
    '002': 'VA111',
    '003': 'MD91',
    '004': 'VA77',
    '005': 'MD100',
    '006': 'MD600',
    '007': 'VA900',
    '008': 'VA500',
    '009': 'MD700',
    '010': 'MD21',
    '011': 'FL01',
    '012': 'VA05',
    '013': 'NJ01',
    '014': 'VA06',
    '015': 'FL02'
}
flag_add_cp_margin = True
flag_dup_cb_udf = False
add_file_nm = '_ORG'
add_file_nm = '_MD'
date_output = df.postprocessing_date_process() ## str : mmddyy
print()
print()
print("date_output : ",date_output)
print()
print()
date_process = date_output[:4]+'20'+date_output[4:]
exceptional = ['0140', '0154']  # RB, SW
test_mode = False

file_cb = './postprocess/통합작업본_'+date_process+'_작업완료원본.xlsx'
pb_dir = './postprocess/pb_'+date_output
wb_r = load_workbook(file_cb)
ws_r = wb_r['Sheet1']
dict_sc = dict()
cb_header = None
pb_header = None
idx1 = 0
for cur_row in ws_r.rows:
    # 첫 행(컬럼 명) 통과
    if idx1 == 0:
        cb_header = get_header_from_row(cur_row)
    else:
        test = []
        item = build_cb_item(cur_row)
        item['idx'] = idx1
        if item['f1000'] not in dict_sc:
            dict_sc[item['f1000']] = {
                'ORG': {},
                'ORG_PB': {},
                'CB_EX_IDX': {},
                'CB_UDF_IDX': {},
                'PB_UDF_IDX': {},
                'PB_IGNORE_DICT': {},
                'OMIT_DUPLICATION': {},
                'OMIT_DUPLICATION_PB': {},
                'UPC': {},
                'UPC_PB': {},
                'RB_FLAG': False,
                'RB_IDX': {},
                'SW_FLAG': False,
                'SW_IDX': {}
            }

        dict_sc[item['f1000']]['ORG'][idx1] = item
        if is_clover_one_gallon(item['f27'], item['f01'], item['f26']):
            dict_sc[item['f1000']]['PB_IGNORE_DICT'][str(item['f01'])+'_+_'+str(item['f27'])+'_+_'+str(item['lookup_f26'])] = 1
            test.append(1)
            if item['old_margin'] == '' or item['new_margin'] == '':
                dict_sc[item['f1000']]['CB_UDF_IDX'][idx1] = 1
                if dict_sc[item['f1000']]['ORG'][idx1]['cat'] == '':
                    test.append(8)
                    dict_sc[item['f1000']]['ORG'][idx1]['cat'] = '고마진/저마진 확인 요망'
        else:
            if item['f27'] in exceptional:
                if item['f27'] == '0140':
                    dict_sc[item['f1000']]['RB_FLAG'] = True
                    dict_sc[item['f1000']]['RB_IDX'][idx1] = 1
                    test.append(2)
                if item['f27'] == '0154':
                    dict_sc[item['f1000']]['SW_FLAG'] = True
                    dict_sc[item['f1000']]['SW_IDX'][idx1] = 1
                    test.append(3)
            elif item['cat'] != '':
                dict_sc[item['f1000']]['CB_UDF_IDX'][idx1] = 1
                test.append(4)
                if item['srp'] != ' ':
                    dict_sc[item['f1000']]['PB_IGNORE_DICT'][str(item['f01'])+'_+_'+str(item['f27'])+'_+_'+str(item['lookup_f26'])] = 1
                    test.append(5)
            elif item['srp'] != ' ':
                dict_sc[item['f1000']]['CB_EX_IDX'][idx1] = 1
                test.append(6)
            else:
                test.append(7)

            if item['f27'] not in exceptional:
                if item['old_margin'] == '' or item['new_margin'] == '':
                    dict_sc[item['f1000']]['CB_UDF_IDX'][idx1] = 1
                    if dict_sc[item['f1000']]['ORG'][idx1]['cat'] == '':
                        test.append(8)
                        dict_sc[item['f1000']]['ORG'][idx1]['cat'] = '고마진/저마진 확인 요망'
                    else:
                        test.append(9)
                elif idx1 not in dict_sc[item['f1000']]['CB_EX_IDX']:
                    if float(item['old_margin']) < 0.01 or float(item['old_margin']) > 0.59:
                        dict_sc[item['f1000']]['CB_UDF_IDX'][idx1] = 1
                        if dict_sc[item['f1000']]['ORG'][idx1]['cat'] == '':
                            dict_sc[item['f1000']]['ORG'][idx1]['cat'] = '고마진/저마진 확인 요망'
                            test.append(10)
                        else:
                            test.append(11)
                    else:
                        test.append(12)
                    if float(item['new_margin']) < 0.01 or float(item['new_margin']) > 0.59:
                        dict_sc[item['f1000']]['CB_UDF_IDX'][idx1] = 1
                        if dict_sc[item['f1000']]['ORG'][idx1]['cat'] == '':
                            dict_sc[item['f1000']]['ORG'][idx1]['cat'] = '고마진/저마진 확인 요망'
                            test.append(13)
                        else:
                            test.append(14)
                    else:
                        test.append(15)
                else:
                    test.append(16)
            else:
                test.append(17)

        if item['f26'] == '0':
            dict_sc[item['f1000']]['CB_UDF_IDX'][idx1] = 1
            if item['srp'] != ' ':
                dict_sc[item['f1000']]['PB_IGNORE_DICT'][str(item['f01'])+'_+_'+str(item['f27'])+'_+_'+str(item['lookup_f26'])] = 1
                test.append(18)
            else:
                test.append(19)
        else:
            test.append(20)

        # if item['f1000'] == "007":
        #     print(item)
        #     print(test)

        # UPC 중복 항목 처리
        if item['f01'] != '' and item['f26'] != '' and item['f27'] not in exceptional:
            test.append(21)
            if item['f01'] not in dict_sc[item['f1000']]['UPC']:
                dict_sc[item['f1000']]['UPC'][item['f01']] = {}
            # 제품 코드+업체 코드가 해당 UPC 이력에 없으면 추가
            if str(item['f27'])+'_+_'+str(item['f26']) not in dict_sc[item['f1000']]['UPC'][item['f01']]:
                dict_sc[item['f1000']]['UPC'][item['f01']][str(item['f27'])+'_+_'+str(item['f26'])] = []
            # 해당 upc, 업체, 제품이 몇번 추가 되었었는지를 기억, idx와 순번
            dict_sc[item['f1000']]['UPC'][item['f01']][str(item['f27'])+'_+_'+str(item['f26'])].append([idx1, 0])
            # 특정 UPC로 등록된 항목이 중복제외하고 여러개인 경우 모두 UDF
            if len(dict_sc[item['f1000']]['UPC'][item['f01']]) > 1:
                # 해당 UPC로 등록된 업체코드+제품코드 모두 순회
                for udf_ic in dict_sc[item['f1000']]['UPC'][item['f01']]:
                    # 해당 UPC+업체코드+제품코드로 등록된 모든 org_idx의 인덱스, udf_idx 순회
                    for udf_idx in range(len(dict_sc[item['f1000']]['UPC'][item['f01']][udf_ic])):
                        item2 = dict_sc[item['f1000']]['ORG'][dict_sc[item['f1000']]['UPC'][item['f01']][udf_ic][udf_idx][0]]
                        if dict_sc[item['f1000']]['UPC'][item['f01']][udf_ic][udf_idx][1] != -1:
                            dict_sc[item['f1000']]['CB_UDF_IDX'][dict_sc[item['f1000']]['UPC'][item['f01']][udf_ic][udf_idx][0]] = 1
                            dict_sc[item['f1000']]['PB_IGNORE_DICT'][str(item2['f01'])+'_+_'+str(item2['f27'])+'_+_'+str(item2['lookup_f26'])] = 1
                            if dict_sc[item['f1000']]['UPC'][item['f01']][udf_ic][udf_idx][0] in dict_sc[item['f1000']]['CB_EX_IDX']:
                                dict_sc[item['f1000']]['CB_EX_IDX'].pop(dict_sc[item['f1000']]['UPC'][item['f01']][udf_ic][udf_idx][0], None)
                            dict_sc[item['f1000']]['UPC'][item['f01']][udf_ic][udf_idx][1] = -1
            # 한 종류가 여러개인 경우 처음거 남기고 다 UDF
            elif len(dict_sc[item['f1000']]['UPC'][item['f01']][str(item['f27'])+'_+_'+str(item['f26'])]) > 1:
                udf_ic = str(item['f27'])+'_+_'+str(item['f26'])
                for udf_idx in range(len(dict_sc[item['f1000']]['UPC'][item['f01']][udf_ic])):
                    item2 = dict_sc[item['f1000']]['ORG'][dict_sc[item['f1000']]['UPC'][item['f01']][udf_ic][udf_idx][0]]
                    if dict_sc[item['f1000']]['UPC'][item['f01']][udf_ic][udf_idx][1] not in [-2, -3]:
                        if udf_idx == 0:
                            # 2020-10-26 va900에서 다 udf로 분류된 오류와 관련있음
                            if item2['srp'] != ' ':
                                dict_sc[item['f1000']]['CB_UDF_IDX'][dict_sc[item['f1000']]['UPC'][item['f01']][udf_ic][udf_idx][0]] = 1
                            dict_sc[item['f1000']]['PB_IGNORE_DICT'][str(item2['f01'])+'_+_'+str(item2['f27'])+'_+_'+str(item2['lookup_f26'])] = 1
                            dict_sc[item['f1000']]['OMIT_DUPLICATION'][dict_sc[item['f1000']]['UPC'][item['f01']][udf_ic][udf_idx][0]] = 1
                            dict_sc[item['f1000']]['UPC'][item['f01']][udf_ic][udf_idx][1] = -2
                        else:
                            dict_sc[item['f1000']]['CB_UDF_IDX'][dict_sc[item['f1000']]['UPC'][item['f01']][udf_ic][udf_idx][0]] = 1
                            dict_sc[item['f1000']]['PB_IGNORE_DICT'][str(item2['f01']) + '_+_' + str(item2['f27']) + '_+_' + str(item2['lookup_f26'])] = 1
                            if dict_sc[item['f1000']]['UPC'][item['f01']][udf_ic][udf_idx][0] in dict_sc[item['f1000']]['CB_EX_IDX']:
                                dict_sc[item['f1000']]['CB_EX_IDX'].pop(dict_sc[item['f1000']]['UPC'][item['f01']][udf_ic][udf_idx][0], None)
                            dict_sc[item['f1000']]['UPC'][item['f01']][udf_ic][udf_idx][1] = -3
        # if item['f01'] == '0001143311279':
        #     print(item)
        #     print(test)
    idx1 += 1
wb_r.close()

#print(dict_sc['006']['UPC'])

pb_f_list = listdir(pb_dir)
idx1 = 0
for pb_f in pb_f_list:
    wb_r = load_workbook(pb_dir+'/'+pb_f)
    ws_r = wb_r['Sheet1']
    idx2 = 0
    for cur_row in ws_r.rows:
        # 첫 행(컬럼 명) 통과
        if idx2 == 0:
            if idx1 == 0:
                pb_header = get_header_from_row(cur_row)
        else:
            test = []
            item = build_pb_item(cur_row)

            if item['f1000'] not in dict_sc:
                dict_sc[item['f1000']] = {
                    'ORG': {},
                    'ORG_PB': {},
                    'CB_EX_IDX': {},
                    'CB_UDF_IDX': {},
                    'PB_UDF_IDX': {},
                    'PB_IGNORE_DICT': {},
                    'OMIT_DUPLICATION': {},
                    'OMIT_DUPLICATION_PB': {},
                    'UPC': {},
                    'UPC_PB': {},
                    'RB_FLAG': False,
                    'RB_IDX': {},
                    'SW_FLAG': False,
                    'SW_IDX': {}
                }

            dict_sc[item['f1000']]['ORG_PB'][idx1] = item

            if str(item['f01'])+'_+_'+str(item['f27'])+'_+_'+str(item['f26']) not in dict_sc[item['f1000']]['PB_IGNORE_DICT']:
                if float(item['old_margin']) < 0.01 or float(item['old_margin']) > 0.49:
                    dict_sc[item['f1000']]['PB_UDF_IDX'][idx1] = 1
                    test.append(1)
                else:
                    test.append(2)
                if float(item['new_margin']) < 0.01 or float(item['new_margin']) > 0.49:
                    dict_sc[item['f1000']]['PB_UDF_IDX'][idx1] = 1
                    test.append(3)
                else:
                    test.append(4)

                if item['f26'] == '0':
                    dict_sc[item['f1000']]['PB_UDF_IDX'][idx1] = 1
                    test.append(5)
                else:
                    test.append(6)

                # UPC 중복 항목 처리
                if item['f01'] != '' and item['f26'] != '':
                    if item['f01'] not in dict_sc[item['f1000']]['UPC_PB']:
                        dict_sc[item['f1000']]['UPC_PB'][item['f01']] = {}
                    if str(item['f27'])+'_+_'+str(item['f26']) not in dict_sc[item['f1000']]['UPC_PB'][item['f01']]:
                        dict_sc[item['f1000']]['UPC_PB'][item['f01']][str(item['f27'])+'_+_'+str(item['f26'])] = []
                    dict_sc[item['f1000']]['UPC_PB'][item['f01']][str(item['f27'])+'_+_'+str(item['f26'])].append([idx1, 0])
                    # 특정 UPC로 등록된 항목이 중복제외하고 여러개인 경우 모두 UDF
                    if len(dict_sc[item['f1000']]['UPC_PB'][item['f01']]) > 1:
                        for udf_ic in dict_sc[item['f1000']]['UPC_PB'][item['f01']]:
                            for udf_idx in range(len(dict_sc[item['f1000']]['UPC_PB'][item['f01']][udf_ic])):
                                if dict_sc[item['f1000']]['UPC_PB'][item['f01']][udf_ic][udf_idx][1] != -1:
                                    dict_sc[item['f1000']]['PB_UDF_IDX'][dict_sc[item['f1000']]['UPC_PB'][item['f01']][udf_ic][udf_idx][0]] = 1
                                    dict_sc[item['f1000']]['UPC_PB'][item['f01']][udf_ic][udf_idx][1] = -1
                    # 한 종류가 여러개인 경우 처음거 남기고 다 UDF
                    elif len(dict_sc[item['f1000']]['UPC_PB'][item['f01']][str(item['f27'])+'_+_'+str(item['f26'])]) > 1:
                        udf_ic = str(item['f27'])+'_+_'+str(item['f26'])
                        for udf_idx in range(len(dict_sc[item['f1000']]['UPC_PB'][item['f01']][udf_ic])):
                            if dict_sc[item['f1000']]['UPC_PB'][item['f01']][udf_ic][udf_idx][1] not in [-2, -3]:
                                if udf_idx == 0:
                                    dict_sc[item['f1000']]['UPC_PB'][item['f01']][udf_ic][udf_idx][1] = -2
                                else:
                                    dict_sc[item['f1000']]['PB_UDF_IDX'][dict_sc[item['f1000']]['UPC_PB'][item['f01']][udf_ic][udf_idx][0]] = 1
                                    dict_sc[item['f1000']]['UPC_PB'][item['f01']][udf_ic][udf_idx] = -3
        idx1 += 1
        idx2 += 1
    wb_r.close()

if not isdir('./postprocess/result'):
    mkdir('./postprocess/result')
if not isdir('./postprocess/result/'+date_output+add_file_nm):
    mkdir('./postprocess/result/'+date_output+add_file_nm)

tot = 0
cb_header_reformat = cb_header[:4]+cb_header[9:14]+cb_header[18:]
print(dict_sc.keys())
flag_false = True
for sc in dict_sc:
    print(date_output+' '+dict_store[sc]+' CPB.xlsx')
    file_result = ''
    wb_n = Workbook()
    ws_n_1 = wb_n.active
    ws_n_1.title = 'CB'
    temp_header = cb_header_reformat[3:-1]
    temp_header.append('Converted Price')
    if flag_add_cp_margin:
        temp_header.append('CP Margin')
    ws_n_1.append(temp_header)
    # ws_n_1.append(cb_header_reformat[3:-1])
    ws_n_2 = wb_n.create_sheet('PB')
    temp_header = pb_header[:]
    temp_header[-1] = 'BEFORE'
    temp_header.append('F30')
    if flag_add_cp_margin:
        temp_header.append('CP Margin')
    ws_n_2.append(temp_header)
    # ws_n_2.append(pb_header)
    ws_n_3 = wb_n.create_sheet('UDF')
    ws_n_3.append(cb_header_reformat)
    ws_n_4 = wb_n.create_sheet('UDF-PB')
    ws_n_4.append(pb_header)
    ws_n_5 = wb_n.create_sheet(sc)
    ws_n_5.append(cb_header_reformat)
    ws_n_6 = wb_n.create_sheet('EXCEPT')
    ws_n_6.append(cb_header_reformat)
    ws_n_8 = wb_n.create_sheet('IGNORE-PB')
    ws_n_8.append(pb_header)
    ws_n_9 = None
    if dict_sc[sc]['RB_FLAG']:
        if dict_sc[sc]['SW_FLAG']:
            ws_n_9 = wb_n.create_sheet('S-R')
        else:
            ws_n_9 = wb_n.create_sheet('R')
    else:
        ws_n_9 = wb_n.create_sheet('S')
    ws_n_9.append(cb_header_reformat)
    ws_n_10 = wb_n.create_sheet('OMIT_DUPLICATION')
    ws_n_10.append(cb_header_reformat)
    idx1 = 0
    idx3 = 0
    idx5 = 0
    idx6 = 0
    idx9 = 0
    idx10 = 0
    cnt_sr_srp = 0
    cnt_cb_ex = 0
    cnt_dup_cb_udf = 0
    for key in dict_sc[sc]['ORG']:
        item = dict_sc[sc]['ORG'][key]
        arr_item = [item['file_nm'], item['cat'], item['sub_cat'], item['f01'], item['f902'], item['f1000'], item['f27'], item['f1184'], item['f26'], item['desc'], item['f1001'], item['f19'], item['f38'], item['f39'], item['f90'], item['qty'], item['amt'], item['f1122'], item['old_ccost'], item['new_ccost'], item['price'], item['old_margin'], item['new_margin'], item['srp']]
        arr_item_floor = [item['file_nm'], item['cat'], item['sub_cat'], item['f01'], item['f902'], item['f1000'], item['f27'], item['f1184'], item['f26'], item['desc'], item['f1001'], item['f19'], item['f38'], item['f39'], item['f90'], item['qty'], item['amt'], item['f1122'], item['old_ccost'], item['new_ccost'], item['price'], str_margin_to_percent(item['old_margin']), str_margin_to_percent(item['new_margin']), item['srp']]
        if key in dict_sc[sc]['RB_IDX'] or key in dict_sc[sc]['SW_IDX']:
            # if item['f01'] == '0001143311279':
            #     print(item)
            if arr_item_floor[1] == '고마진/저마진 확인 요망':
                arr_item_floor[1] = ''
            ws_n_9.append(arr_item_floor)
            idx9 += 1
            ws_n_9.cell(idx9+1, 22).number_format = '0%'
            ws_n_9.cell(idx9+1, 23).number_format = '0%'
            if item['srp'] != ' ' and item['srp'] != '':
                cnt_sr_srp += 1
        else:
            if key in dict_sc[sc]['CB_UDF_IDX']:
                if key in dict_sc[sc]['OMIT_DUPLICATION']:
                    idx10 += 1
                    ws_n_10.append(arr_item_floor)
                    ws_n_10.cell(idx10+1, 22).number_format = '0%'
                    ws_n_10.cell(idx10+1, 23).number_format = '0%'
                else:
                    # print(arr_item_floor)
                    ws_n_3.append(arr_item_floor)
                    idx3 += 1
                    ws_n_3.cell(idx3+1, 22).number_format = '0%'
                    ws_n_3.cell(idx3+1, 23).number_format = '0%'

                    if flag_dup_cb_udf:
                        # 2021-04-27 부 추가 기능: UDF 중 신상품 등 일부 분류에 속하는 항목 조건에 따라 CB에도 입력
                        cat = arr_item_floor[1]
                        upc = arr_item_floor[3]
                        if cat == "신상품" and upc != '':
                            new_arr_item_floor = arr_item_floor[3:-7]
                            ws_n_1.append(new_arr_item_floor)
                            cnt_dup_cb_udf += 1
                            ws_n_1.cell(idx1+cnt_dup_cb_udf+1, 19).number_format = '0%'
                            ws_n_1.cell(idx1+cnt_dup_cb_udf+1, 20).number_format = '0%'
                            if flag_add_cp_margin:
                                ws_n_1.cell(idx1+cnt_dup_cb_udf+1, 22).number_format = '0%'
                        elif cat == "UPC로 검색된 제품코드가 인보이스의 제품코드와 상이":
                            new_arr_item_floor = arr_item_floor[3:-1]
                            if new_arr_item_floor[17] != '':
                                tmp_converted_price = get_converted_price(new_arr_item_floor[17])
                                new_arr_item_floor.append(tmp_converted_price)
                                if flag_add_cp_margin:
                                    new_arr_item_floor.append(
                                        get_margin(tmp_converted_price, new_arr_item_floor[16], new_arr_item_floor[8],
                                                   new_arr_item_floor[14]))
                            ws_n_1.append(new_arr_item_floor)
                            cnt_dup_cb_udf += 1
                            ws_n_1.cell(idx1+cnt_dup_cb_udf+1, 19).number_format = '0%'
                            ws_n_1.cell(idx1+cnt_dup_cb_udf+1, 20).number_format = '0%'
                            if flag_add_cp_margin:
                                ws_n_1.cell(idx1+cnt_dup_cb_udf+1, 22).number_format = '0%'
                        elif cat == "고마진/저마진 확인 요망":
                            new_arr_item_floor = arr_item_floor[3:-1]
                            if new_arr_item_floor[17] != '':
                                tmp_converted_price = get_converted_price(new_arr_item_floor[17])
                                new_arr_item_floor.append(tmp_converted_price)
                                if flag_add_cp_margin:
                                    new_arr_item_floor.append(
                                        get_margin(tmp_converted_price, new_arr_item_floor[16], new_arr_item_floor[8],
                                                   new_arr_item_floor[14]))
                            ws_n_1.append(new_arr_item_floor)
                            cnt_dup_cb_udf += 1
                            ws_n_1.cell(idx1+cnt_dup_cb_udf+1, 19).number_format = '0%'
                            ws_n_1.cell(idx1+cnt_dup_cb_udf+1, 20).number_format = '0%'
                            if flag_add_cp_margin:
                                ws_n_1.cell(idx1+cnt_dup_cb_udf+1, 22).number_format = '0%'
            elif key in dict_sc[sc]['CB_EX_IDX']:
                idx6 += 1
                if test_mode:
                    ws_n_6.append(arr_item_floor)
                    ws_n_6.cell(idx6+1, 22).number_format = '0%'
                    ws_n_6.cell(idx6+1, 23).number_format = '0%'
                cnt_cb_ex += 1
            else:
                new_arr_item_floor = arr_item_floor[3:-1]
                tmp_converted_price = get_converted_price(new_arr_item_floor[17])
                new_arr_item_floor.append(tmp_converted_price)
                if flag_add_cp_margin:
                    new_arr_item_floor.append(get_margin(tmp_converted_price, new_arr_item_floor[16], new_arr_item_floor[8], new_arr_item_floor[14]))
                ws_n_1.append(new_arr_item_floor)
                # ws_n_1.append(arr_item_floor[3:-1])
                idx1 += 1
                ws_n_1.cell(idx1+cnt_dup_cb_udf+1, 19).number_format = '0%'
                ws_n_1.cell(idx1+cnt_dup_cb_udf+1, 20).number_format = '0%'
                if flag_add_cp_margin:
                    ws_n_1.cell(idx1+cnt_dup_cb_udf+1, 22).number_format = '0%'
            if arr_item_floor[1] == '고마진/저마진 확인 요망':
                arr_item_floor[1] = ''
            ws_n_5.append(arr_item_floor)
            idx5 += 1
            ws_n_5.cell(idx5+1, 22).number_format = '0%'
            ws_n_5.cell(idx5+1, 23).number_format = '0%'
    idx2 = 0
    idx4 = 0
    idx7 = 0
    idx8 = 0
    for key in dict_sc[sc]['ORG_PB']:
        item = dict_sc[sc]['ORG_PB'][key]
        arr_item = [item['f01'], item['f902'], item['f1000'], item['f27'], item['f26'], item['f1184'], item['f126'], item['f1001'], item['f19'], item['f38'], item['f39'], item['f90'], item['f1122'], item['old_ccost'], item['new_ccost'], item['price'], item['old_margin'], item['new_margin'], item['srp']]
        arr_item_floor = [item['f01'], item['f902'], item['f1000'], item['f27'], item['f26'], item['f1184'], item['f126'], item['f1001'], item['f19'], item['f38'], item['f39'], item['f90'], item['f1122'], item['old_ccost'], item['new_ccost'], item['price'], str_margin_to_percent(item['old_margin']), str_margin_to_percent(item['new_margin']), item['srp']]
        if str(item['f01'])+'_+_'+str(item['f27'])+'_+_'+str(item['f26']) in dict_sc[sc]['PB_IGNORE_DICT']:
            idx8 += 1
            if test_mode:
                ws_n_8.append(arr_item_floor)
                ws_n_8.cell(idx8+1, 17).number_format = '0%'
                ws_n_8.cell(idx8+1, 18).number_format = '0%'
        elif key in dict_sc[sc]['PB_UDF_IDX']:
            ws_n_4.append(arr_item_floor)
            idx4 += 1
            ws_n_4.cell(idx4+1, 17).number_format = '0%'
            ws_n_4.cell(idx4+1, 18).number_format = '0%'
        else:
            new_arr_item_floor = arr_item_floor[:]
            tmp_converted_price = ''
            if len(new_arr_item_floor) >= 19 and new_arr_item_floor[18] is not None and new_arr_item_floor[18] != '':
                tmp_converted_price = get_converted_price(new_arr_item_floor[18])
            new_arr_item_floor.append(tmp_converted_price)
            if flag_add_cp_margin:
                new_arr_item_floor.append(get_margin(tmp_converted_price, new_arr_item_floor[14], new_arr_item_floor[8], new_arr_item_floor[12]))
            ws_n_2.append(new_arr_item_floor)
            # ws_n_2.append(arr_item_floor)
            idx2 += 1
            ws_n_2.cell(idx2+1, 17).number_format = '0%'
            ws_n_2.cell(idx2+1, 18).number_format = '0%'
            if flag_add_cp_margin:
                ws_n_2.cell(idx2+1, 21).number_format = '0%'

    ws_n_1.freeze_panes = 'A2'
    ws_n_1.auto_filter.ref = ws_n_1.dimensions
    ws_n_2.freeze_panes = 'A2'
    ws_n_2.auto_filter.ref = ws_n_2.dimensions
    ws_n_3.freeze_panes = 'A2'
    ws_n_3.auto_filter.ref = ws_n_3.dimensions
    ws_n_4.freeze_panes = 'A2'
    ws_n_4.auto_filter.ref = ws_n_4.dimensions
    ws_n_5.freeze_panes = 'A2'
    ws_n_5.auto_filter.ref = ws_n_5.dimensions
    ws_n_9.freeze_panes = 'A2'
    ws_n_9.auto_filter.ref = ws_n_9.dimensions

    print(ws_n_1.title+': '+str(ws_n_1.max_row - 1 - cnt_dup_cb_udf))
    print(ws_n_2.title+': '+str(ws_n_2.max_row - 1))
    print(ws_n_3.title+': '+str(ws_n_3.max_row - 1))
    print(ws_n_4.title+': '+str(ws_n_4.max_row - 1))
    print(ws_n_5.title+': '+str(ws_n_5.max_row - 1)) # 1247
    #print(ws_n_6.title+': '+str(ws_n_6.max_row - 1))
    #print(ws_n_8.title+': '+str(ws_n_8.max_row - 1))
    print(ws_n_9.title+': '+str(ws_n_9.max_row - 1))
    print(ws_n_10.title+': '+str(ws_n_10.max_row - 1)) # OMIT_DUPLICATION
    print('sr_srp: '+str(cnt_sr_srp))
    #print('cb_ex: '+str(cnt_cb_ex))
    print((ws_n_1.max_row + ws_n_2.max_row + ws_n_3.max_row + ws_n_4.max_row - 4 - cnt_sr_srp - cnt_dup_cb_udf)) # 1250
    validation_result = (ws_n_5.max_row - 1) == (
                ws_n_1.max_row + ws_n_2.max_row + ws_n_3.max_row + ws_n_4.max_row + ws_n_10.max_row - 5 - cnt_sr_srp - cnt_dup_cb_udf)
    if not validation_result:
        flag_false = False
    print(validation_result)
    tot += (ws_n_5.max_row + ws_n_9.max_row - 3)
    if ws_n_1.max_row == 1:
        wb_n.remove_sheet(ws_n_1)
    if ws_n_2.max_row == 1:
        wb_n.remove_sheet(ws_n_2)
    if ws_n_3.max_row == 1:
        wb_n.remove_sheet(ws_n_3)
    if ws_n_4.max_row == 1:
        wb_n.remove_sheet(ws_n_4)
    if ws_n_5.max_row == 1:
        wb_n.remove_sheet(ws_n_5)
    if ws_n_6.max_row == 1:
        wb_n.remove_sheet(ws_n_6)
    if ws_n_8.max_row == 1:
        wb_n.remove_sheet(ws_n_8)
    if ws_n_9.max_row == 1:
        wb_n.remove_sheet(ws_n_9)
    if ws_n_10.max_row == 1:
        wb_n.remove_sheet(ws_n_10)
    wb_n.save('./postprocess/result/'+date_output+add_file_nm+'/'+date_output+' '+dict_store[sc]+' CPB.xlsx')

print("total: "+str(tot))
print("validation_result: "+str(flag_false))
print("postprocessing_backup.py까지 실행되었습니다.")
print("==================================================")
print("==================================================")
print("finalProcessing.py 실행...")
exec(open("finalProcessing.py", encoding="utf-8").read())