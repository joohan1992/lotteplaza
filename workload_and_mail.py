import dataEntryFunction as df
import datetime
import openpyxl
import pyperclip
import shutil
import os
###날짜 설정
now = datetime.datetime.now()
now_yymmdd = now.strftime("%y%m%d")
now_mmdd = now.strftime("%m%d")
now_yyyymmdd = now.strftime("%Y-%m-%d")
 ## R1, R2 입력 후 출력 결과물 복사해서 그대로 붙여넣으면 됩니다.
 ## 문자열로 입력시 공백이나 앞에 0 입력해도 됩니다. (int로 입력해도 됨)

 ## 구글시트 출력순서 그대로 복붙하시면 됩니다.
 ## 순서대로 출력됩니다.
구글시트출력순서 = '''

임미화
이지원
서지민
윤수연
최진호
BI

'''
#@todo : ##move_files_with_rename(PATH1,PATH2) 이용해서 제외파일 합치기
## 제외파일합치기

##df.move_files_with_rename(r"C:\Users\user\Documents\GitHub\lotteplaza\작업파일\1203\제외",r"C:\Users\user\Documents\GitHub\lotteplaza\작업파일\1204\제외")
##df.move_files_with_rename(r"C:\Users\user\Desktop\1203",r"C:\Users\user\Desktop\1204")  #원본만들어놔야됨
# 최진호님 없는버전
검수_1차 = {
        ###     R1   R2
      "서지민" : [ "1088" , "112"],
         "최진호" : [ " 194" , "148" ],
            "임미화" : [ " 487" , "056"],
               "윤수연" : [ "1018" , "138" ],
                  "이지원" : [ " 558" , "045" ],

}
## 자리다시맞춰야됨

검수1차ocr ='''
아래에 네이버 ocr로 단톡방에 올라온 검수 이미지파일 복붙해서 넣으면 됩니다.
뭔가 에러나거나 잘 안되면 위의 검수_1차 dictionary 이용하세요

서지민 R1: 588 / R2: 132
임미화 R1: 401 / R2: 207
윤수연 R1: 663 / R2:81
이지원 R1: 504 / R2: 172

'''
검수_1차 = df.get_first_workdata (검수1차ocr)
검수_2차 = df.get_second_workdata ()

# 검수_1차 = {
#         ###     R1   R2
#       "서지민" : [ " 720" , "128"],
#          "최진호" : [ " 000" , "000" ],
#             "임미화" : [ " 548" , " 63"],
#                "윤수연" : [ " 651" , "148" ],
#                   "이지원" : [ " 498" , "109" ],
#
# }
# # ## 자리다시맞춰야됨
# 검수_2차 = {
#         ###     R1   R2
#       "서지민" : [ " 720" , "128"],
#          "최진호" : [ " 000" , "000" ],
#             "임미화" : [ " 546" , " 63"],
#                "윤수연" : [ " 655" , "148" ],
#                   "이지원" : [ " 565" , "114" ],
# }

objdate = ["01", "02"]

def convert_to_int(value):
    if type(value) == int:
        return value
    return int(value.strip())

# 딕셔너리의 값들을 정수형으로 변환
for key, value in 검수_1차.items():
    검수_1차[key] = [convert_to_int(val) for val in value]
for key, value in 검수_2차.items():
    검수_2차[key] = [convert_to_int(val) for val in value]

검수_1차_변환={}
검수_2차_변환={}

for key in 검수_1차:
    검수_1차_변환[key] = [검수_1차[key][0]+검수_1차[key][1],검수_1차[key][1]]

for key in 검수_2차:
    검수_2차_변환[key] = [검수_2차[key][0]+검수_2차[key][1],검수_2차[key][1]]

print(검수_1차_변환)
print(검수_2차_변환)

R3변화 = 0
R2변화 = 0


for key in 검수_2차_변환:
    if 검수_2차_변환[key][0]-검수_1차_변환.get(key, [0,0])[0]>0:
        R3변화 +=검수_2차_변환[key][0]-검수_1차_변환[key][0]
    if 검수_2차_변환[key][1]-검수_1차_변환.get(key, [0,0])[1]>0:
        R2변화 += 검수_2차_변환[key][1] - 검수_1차_변환[key][1]

print("R3 변화 :",R3변화)
print("R2 변화 :",R2변화)
검수_최종=검수_1차_변환.copy()
검수_최종['이지원'][0]+=R3변화
검수_최종['이지원'][1]+=R2변화
print("검수 최종 :",검수_최종)


작업자순서 = ""
구글시트복사용 =""
for 작업자 in 구글시트출력순서.split("\n") :
    if 작업자.strip() != '':
        if 작업자 =="BI":
            pass ## BI가 있더라도 아무것도하지말기
        elif 작업자 in 검수_최종:
            작업자순서 += f"----{작업자}"
            구글시트복사용 += f"{검수_최종[작업자][0]}\t{검수_최종[작업자][1]}\n"
        else:
            작업자순서 += f"----{작업자}"
            구글시트복사용 += f"0\t0\n"
구글시트복사용 += "0\t400"
작업자순서 += "----BI 순서대로"
print(작업자순서)
print(구글시트복사용)


작업량기록여부 = False
if df.isMonday(): ## 월요일이면
    try :
        filepath = "C:/Users/user/Documents/GitHub/lotteplaza/작업파일/" + now_mmdd + "/" + str(int(now_yymmdd) - 1) + "_" + objdate[0] + ".xlsx에서 분류행 개수 세는중.."
        print(filepath)
        분류행개수 = df.getClassifiedRows(filepath)
        print("주말 : 분류된 행 개수는 " + str(분류행개수) + " 개 입니다. 그대로 입력하세요")

        filepath = "C:/Users/user/Documents/GitHub/lotteplaza/작업파일/" + now_mmdd + "/" + now_yymmdd + "_" + objdate[0] + ".xlsx"
        print(filepath)
        분류행개수 = df.getClassifiedRows(filepath)
        print("월요일 : 분류된 행 개수 + 400은 "+str(분류행개수 + 400)+" 개 입니다.")
    except :
        print("ERROR")
else:                   ## 월요일 아닌경우
    try :
        filepath = "C:/Users/user/Documents/GitHub/lotteplaza/작업파일/" + now_mmdd + "/" + now_yymmdd + "_" + objdate[0] + ".xlsx"
        print(filepath)
        분류행개수 = df.getClassifiedRows(filepath)
        print("분류된 행 개수 + 400은 "+str(분류행개수 + 400)+" 개 입니다.")
        작업량기록엑셀파일경로 = "C:/Users/user/Documents/GitHub/lotteplaza/작업파일/" + now_mmdd + "/결과/Data_Entry_작업량_기록_" + now_mmdd + "2024.xlsx"
        workbook = openpyxl.load_workbook(작업량기록엑셀파일경로)
        sheet = workbook.active
        단계1작업수 = sum([values[0] for values in 검수_최종.values()])
        for row in range(1, sheet.max_row + 1):
            b_column =  sheet.cell(row=row, column=2).value
            if type(b_column) == datetime.datetime:
                b_column = b_column.strftime('%Y-%m-%d')
            if b_column == now_yyyymmdd:
                sheet.cell(row=row, column=3).value = 단계1작업수  # C열에 입력할 값
                sheet.cell(row=row, column=4).value = 분류행개수 + 400  # D열에 입력할 값
                break
        # 수정된 내용을 저장
        workbook.save(작업량기록엑셀파일경로)
        작업량기록여부 = True
    except :
        print("ERROR")


all_count = df.count_pdf_files("C:/Users/user/Desktop/" + str(now_mmdd))
except_path = "C:/Users/user/Documents/GitHub/lotteplaza/작업파일/" + str(now_mmdd) + "/제외/"
except_count = df.count_pdf_files(except_path)
# 월요일에 전체 인보이스 count추가, 제외파일 합치기
if df.isMonday():
    print("오늘은 월요일 입니다. 전체 인보이스를 합칩니다.")
    weekend_mmdd = str(df.get_previous_day(now_yymmdd)[2:6])
    weekend_count = df.count_pdf_files("C:/Users/user/Desktop/" + weekend_mmdd)
    print(f"주말 인보이스 : {weekend_count}")
    print(f"월요일 인보이스 : {all_count}")
    all_count = all_count + weekend_count
    print(f"전체 인보이스 : {all_count}\n")
    print("제외 파일을 합칩니다.")
    df.move_files_with_rename(f"C:\\Users\\user\\Documents\\GitHub\\lotteplaza\\작업파일\\{weekend_mmdd}\\제외",
                              f"C:\\Users\\user\\Documents\\GitHub\\lotteplaza\\작업파일\\{now_mmdd}\\제외")

    zip_path = f"C:/Users/user/Documents/GitHub/lotteplaza/작업파일/{now_mmdd}/제외"
    result_path = f"C:/Users/user/Documents/GitHub/lotteplaza/작업파일/{now_mmdd}/결과/제외인보이스_{now_mmdd + str(datetime.datetime.now().year)}"
    print(f"합치기 완료, 기존 제외인보이스_{now_mmdd + str(datetime.datetime.now().year)}.zip 파일을 삭제하고 다시 압축합니다.")
    file_path = 'path/to/your/file.txt'  # 삭제할 파일의 경로
    os.remove(result_path+'.zip')
    print("삭제 완료. 재압축합니다.")
    shutil.make_archive(result_path, 'zip', zip_path)
    print("제외 파일이 합쳐져서 재압축 되었습니다.")
    except_count = df.count_pdf_files(except_path)


pyperclip.copy(구글시트복사용)
print("\n클립보드에 구글시트복사용 저장했습니다.")

if 작업량기록여부:
    print(f"Data_Entry_작업량_기록 엑셀파일에 오늘 날짜로 각각 {단계1작업수}, {분류행개수 + 400} 입력했습니다.")
else :
    print("Data_Entry_작업량_기록 엑셀파일에 오늘 날짜로 작업하지 않았습니다. 직접 입력해주세요.")

next = input("아무키나 입력하고 엔터치면 본문이 복사가 됩니다.")

본문1 = df.본문1()

본문1 = 본문1.replace("#{TOT_CNT}",str(all_count))
본문1 = 본문1.replace("#{EXT_CNT}",str(except_count))
본문1 = 본문1.replace("#{VAL_CNT}",str(all_count - except_count))
reason = ["제품코드, UPC 부재 & 한글 Description & 수기", "입력 대상 제품 없음", "업체코드 검색 불가", "파본"]
reason1 = df.count_pdf_files(except_path + "제품코드, UPC 부재 & 한글 Description & 수기")
reason2 = df.count_pdf_files(except_path + "입력 대상 제품 없음")
reason3 = df.count_pdf_files(except_path + "업체코드 검색 불가")
reason4 = df.count_pdf_files(except_path + "파본")
for i in reason:
    if df.count_pdf_files(except_path + i) > 0 :
        temp_cnt = df.count_pdf_files(except_path + i)
        본문1 += "<b>- "+i+" ("+str(temp_cnt)+"건)</b>\n"
        except_count = except_count - temp_cnt
if except_count != 0:
    print("!!분류되지 않은 제외인보이스가 있습니다!!")

본문1 = 본문1.replace("&","&amp;")
본문1 = 본문1+df.본문2()
# @TODO : 클립보드에 add하는법 찾아보기
print(본문1)
print()
print()
pyperclip.copy(본문1)
print("@"*135)
print("@@@@@@@@@@@@ 클립보드에 저장 되었습니다 : 메일내용 !그대로! 붙여넣기 (띄어쓰기 안된것처럼 보여도 전송하면 제대로 적용이 됩니다.) @@@@@@@@@@@@@")
print("@"*135)
print()
첨부파일 = f'''"CPB_처리_결과파일_{now_mmdd}2024.zip" "통합작업본_{now_mmdd}2024.xlsx" "통합작업본_{now_mmdd}2024_작업완료원본.xlsx" "PB batch files_{now_mmdd}2024.zip" "제외인보이스_{now_mmdd}2024.zip" "Data_Entry_작업량_기록_{now_mmdd}2024.xlsx"'''
print(첨부파일)



